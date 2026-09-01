#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
MODELLO v15 — Riforma pensionistica italiana: transizione a sistema multipilastro
================================================================================
Motore Python che sostituisce il file Excel v14. Produce un Excel con struttura
simile a v14 (Parametri, Modello, Dinamica, Modello_Pess, Canali, Sostituzione,
Output, Validazione_v14, Note).

DUE MOTORI:
1) REPLICA v14 — riproduce cella per cella la logica del foglio "Modello" di v14
   (Leva F con EMA, P0 forfait 13 mld, saldo di cassa con emissioni BRDP come
   entrate, nessun decumulo del fondo). Serve SOLO per validazione: dimostra che
   il motore Python riproduce v14 entro tolleranza numerica.

2) MODELLO v15 — contabilità per coorti, con le correzioni:
   C1  Sentiero status quo per ANCORE ufficiali (DFP 2026 / RGS magg. 2026:
       15.4% 2027, 15.5% 2028-29, picco 17.1% 2041 con plateau al 2044,
       16.2% 2050, 14.0% 2070, poi costante) invece della discesa lineare
       infinita di v14 e del livello 16.0% al 2027 (che veniva dal country
       fiche AR2024, perimetro più ampio e PIL pre-revisione ISTAT set. 2025).
   C2  Spesa divisa in quota contributiva (72.8%) e assistenziale: la riduzione
       di P1 si applica solo alla quota contributiva.
   C3  Coorti esplicite: ogni coorte di nuovi pensionati porta la sua P1
       residua, il suo montante P2, il suo P0 (con taper) e la sua mortalità.
       Partecipazione alla biforcazione coerente col phase-in per età (i
       lavoratori over-45 al 2027 non biforcano mai: coorti pre-2051 hanno P2=0).
   C4  DECUMULO DEL FONDO: le rendite P2 pagate ai pensionati escono dal fondo.
   C5  Canale fiscale onesto: perdita IRPEF sulle P1 ridotte (aliquota media
       pensioni); regime TEE (default, come v14: Leva B tassa i contributi
       biforcati) oppure EET (niente Leva B, rendite P2 tassate).
   C6  P0 calcolato, non forfait: taper sul reddito P1+P2 per profilo di
       reddito, al netto del safety net assorbito (32.5 mld — parametro B112 di
       v14 che era MORTO).
   C7  Saldo PRIMARIO separato dal finanziamento: le emissioni BRDP non sono
       entrate; interessi su BRDP e sul debito di transizione accumulato;
       stock di debito di transizione con dinamica esplicita.
   C8  (storica) Sovrapposizione Leva C / Leva J dedotta — SUPERATA da R2.

RAZIONALIZZAZIONE 2026-07 (successiva a C1-C9, decisione utente):
   R1  DATI AGGIORNATI a luglio 2026: sentiero SQ da DFP 2026/RGS (v. C1);
       pensioni vigenti 21.4 mln e pensionati 16.4 mln (INPS Osservatorio
       1.1.2026: 21.26 mln, +0.6%/anno — v10 assumeva +500k/anno, smentito);
       stock fondi pensione 262 €mld (COVIP Relazione 2026, fine 2025);
       coefficiente di trasformazione a 69 anni 6.024% (DM 20/11/2024,
       divisore 16.600); PIL base 2247 €mld a prezzi ~2025 (post revisione
       ISTAT set. 2025; l'etichetta "prezzi 2027" di v10-v14 era impropria:
       2184*(1+g reali) è PIL in volume a prezzi dell'anno base ~2024-25).
   R2  LEVE RAZIONALIZZATE: Leva E eliminata (costo puro senza ritorno
       demografico endogeno). Leva C eliminata: il baseline RGS a legislazione
       vigente include GIA' l'indicizzazione dell'età alla speranza di vita
       (per MEF/RGS vale >20pp di PIL cumulati al 2060), quindi contare per
       intero "67->69 + deroghe" era doppio conteggio. La parte incrementale
       vera confluisce in: Leva I = wedge di +2 anni OLTRE il sentiero
       indicizzato (rampa 15 anni fino a 9 €mld, poi costante); Leva J =
       chiusura deroghe/anticipazioni (3.5 €mld decadenti a zero in 10 anni,
       le coorti aventi diritto si esauriscono entro il ~2037). Il flag
       baseline_politico (default OFF) aggiunge la vecchia componente
       anti-sterilizzazione della Leva I (2.75 + 0.14/biennio), legittima solo
       dichiarando un baseline in cui le sterilizzazioni post-2019 continuano.
       La replica v14 usa ancora le formule originali (validazione invariata).

CONVENZIONI: tutti i valori in termini reali, prezzi ~2025 (volume, anno base
contabilità nazionale). Flussi positivi = risparmi/entrate per lo Stato;
negativi = costi. Importi in €mld salvo indicato.

Uso:  python3 modello_v15.py [--out FILE.xlsx] [--v14 FILE_V14.xlsx]
================================================================================
"""

from __future__ import annotations
import argparse
import dataclasses
from dataclasses import dataclass, field
import math
import numpy as np

# ------------------------------------------------------------------------------
# 1. PARAMETRI
# ------------------------------------------------------------------------------

@dataclass
class Params:
    # --- Orizzonte -------------------------------------------------------------
    anno0: int = 2027
    anni: int = 54                       # 2027-2080

    # --- Demografia e macro (fonti: RGS 26/2025, ISTAT, DEF 2025, AR2024) ------
    # R1 — INPS Osservatorio 1.1.2026 (incl. GDP): 21.26 mln pensioni (+0.6%/a)
    # -> ~21.4 all'1.1.2027; pensionati fisici ~16.3 mln stabili -> 16.4.
    # (v10 usava 22.3/17.0 assumendo +500k/anno: il flusso netto vero e' ~+120k.)
    pensioni_vigenti0: float = 21.4      # mln
    pensionati0: float = 16.4            # mln pensionati fisici
    occupati0: float = 17.5              # mln dipendenti (FPLD+CTPS)

    # --- PERIMETRO (Task 8, 2026-07) --------------------------------------------
    # "universale" = la biforcazione si applica a TUTTI i lavoratori coperti
    # (dipendenti + autonomi INPS + casse professionali), coerente col fatto che
    # la SPESA H copre l'intero sistema. "L1" = solo dipendenti (vecchio default,
    # incoerente: contributi su 17.5 mln ma risparmi su tutti i pensionati).
    # CORREZIONE dell'ASIMMETRIA: prima i contributi (Leva D/B/Z) erano sul monte
    # salari dei soli dipendenti mentre i risparmi (dP1) erano su H = intero
    # sistema. L'asimmetria SOTTOSTIMAVA il costo di transizione. Ora i
    # contributi sono sull'intero monte salari coperto.
    # CALIBRAZIONE (aggiornata 2026-08, w0 ponderato priv+pub): monte dipendenti
    # 565,5 €mld + autonomi ~168 €mld -> gettito universale 0,33x733 = ~242 €mld
    # contro i ~260 osservati (Itinerari XIII). Il residuo deriva dalla
    # contribuzione figurativa GIAS e da rate contributive eterogenee non
    # pienamente catturate: e' una sottostima PRUDENTE (riduce la base da cui la
    # riforma trae contributi, quindi il costo di transizione Leva D e' semmai
    # un limite inferiore).
    #
    # "dipendenti" (default dal 2026-07, indicazione del prof. Sonno): la
    # biforcazione si applica ai soli lavoratori dipendenti. Motivazione di
    # POLITICA: gli autonomi hanno aliquote reali del 24-26% (non 33%) e le casse
    # professionali sono enti privatizzati fuori dall'INPS — estendere la riforma
    # richiederebbe un intervento legislativo separato. Riformare tutto insieme
    # non e' politicamente praticabile.
    # In questa modalita' il modello lavora sul SOTTO-SISTEMA dipendenti in modo
    # COERENTE: spesa, pensionati, flussi e safety net sono ridotti alla quota
    # dipendenti, cosi' i contributi biforcati e i risparmi stanno sulla stessa
    # popolazione. La spesa delle altre gestioni resta invariata e viene
    # ri-sommata per esprimere la spesa di sistema in %PIL (confrontabile RGS).
    perimetro: str = "dipendenti"
    # Quote del perimetro dipendenti (Itinerari Previdenziali, XIII Rapporto,
    # dati 2024): prestazioni dipendenti privati 150,9 + pubblici 93,6 = 244,5
    # €mld su una spesa pensionistica totale (lordo GIAS) di 330,7 €mld.
    quota_spesa_dip: float = 0.739
    # Quota dei FLUSSI di pensionamento (INPS 2025): FPLD 352.632 + Gestione
    # Dipendenti Pubblici 126.591 + fondi speciali 63.630 su 792.998 totali.
    # Piu' bassa della quota di spesa perche' le pensioni dei dipendenti
    # (soprattutto pubblici) sono mediamente piu' alte.
    quota_teste_dip: float = 0.685
    # --- autonomi (usati SOLO se perimetro = "universale") ---
    occ_autonomi0: float = 6.0           # mln autonomi INPS + gest. sep. + casse prof.
    w_autonomi: float = 28000.0          # imponibile medio blended (casse prof. alzano la media)
    g_occ_autonomi: float = -0.007       # trend autonomi (in calo, ISTAT)

    @property
    def quote_perimetro(self) -> tuple[float, float]:
        """(quota di spesa, quota di teste) del perimetro riformato.
        Universale = tutto il sistema; dipendenti = sotto-sistema FPLD+CTPS."""
        if self.perimetro == "universale":
            return 1.0, 1.0
        return self.quota_spesa_dip, self.quota_teste_dip
    # Stock pensioni vigenti (Task 4.4): cresce con l'uscita dei baby boomer fino
    # al picco (~2045, coerente con ISTAT: 65+ in salita fino a ~2050) e poi
    # DECLINA lievemente (popolazione anziana in calo dopo il plateau). Prima
    # cresceva +0.6%/anno all'infinito -> 29.4 mln al 2080, implausibile.
    g_pensioni: float = 0.006            # crescita reale pensioni vigenti (fino al picco)
    anno_picco_pens: int = 2045          # picco stock pensionati (ISTAT 65+)
    g_pensioni_post: float = -0.003      # declino annuo dopo il picco (ISTAT)
    n_occ: float = -0.005                # crescita occupati al netto migrazione
    # C9 — IMMIGRAZIONE decomposta (in v14: 0.116 = 165k x 70%, incoerente con
    # B16=176k, senza filtro tasso di occupazione, e con un fattore (1-n)^t che
    # faceva CRESCERE il flusso migratorio dello 0.5%/anno senza fonte).
    migranti_netti: float = 0.176        # mln/anno netti (RGS 26/2025 base)
    quota_eta_lav: float = 0.70          # quota in eta' lavorativa (Moressa)
    tasso_occ_migr: float = 0.65         # tasso di occupazione migranti (ISTAT)
    migrazione_v14: float = 0.116        # SOLO replica v14 (B18/B145)
    # R1 — PIL in volume, prezzi ~2025 (post revisione ISTAT set. 2025: PIL 2024
    # ~2200 €mld; 2200*1.005*1.007*1.006 ~ 2247 al 2027). NB: era etichettato
    # "prezzi 2027" ma la catena 2184*(1+g reali) e' a prezzi dell'anno base.
    pil0: float = 2247.0                 # €mld, volume a prezzi ~2025
    g: float = 0.008                     # crescita PIL reale
    # IMPONIBILE MEDIO PONDERATO privati+pubblici (correzione 2026-08).
    # Il vecchio 26.700 era l'imponibile medio FPLD (soli privati), ma il
    # perimetro include anche i pubblici, che guadagnano di piu': il monte
    # salari risultava ~21% sotto l'osservato (era il "LIMITE PRINCIPALE"
    # dell'Audit). Derivazione bottom-up dai monte retributivi osservati:
    #   privati  (INPS Oss. dipendenti privati 2024): 17,7 mln lavoratori
    #            x 24.486 € medi          = 433,4 €mld
    #   pubblici (INPS Oss. lavoratori pubblici 2024): 3,74 mln lavoratori
    #            x 35.350 € medi          = 132,1 €mld
    #   monte imponibile FPLD+GDP        = 565,5 €mld
    # Rapportato alle 17,5 mln posizioni del perimetro (i "lavoratori con
    # almeno una giornata retribuita" INPS sono 21,4 mln, molti part-year:
    # il monte va spalmato sulle posizioni, non sulle teste-anno):
    #   w0 = 565,5 €mld / 17,5 mln = ~32.300 €.
    # Il gettito IVS modellato (0,33 x 565,5 = 186,6 €mld) resta ~11% sotto i
    # 209,7 osservati (Itinerari XIII): il residuo e' contribuzione FIGURATIVA
    # a carico GIAS (CIG, malattia, maternita') che non passa dal monte salari
    # effettivo. Sottostima PRUDENTE dichiarata nell'Audit.
    w0: float = 32300.0                  # imponibile medio ponderato priv+pub €
    g_w: float = 0.002                   # crescita salari reali

    @property
    def migrazione(self) -> float:
        """Occupati aggiuntivi netti da migrazione, mln/anno (C9)."""
        return self.migranti_netti * self.quota_eta_lav * self.tasso_occ_migr

    # --- Status quo PAYG (C1/R1: DFP 2026 + RGS monitoraggio magg. 2026) --------
    aliquota: float = 0.33               # aliquota contributiva IVS (9.19+23.81)
    # Cuneo fiscale totale OCSE per il lavoratore medio (Taxing Wages 2025:
    # 45.8%, 5o piu' alto OCSE). Include IRPEF + contributi lav. + contributi
    # datore. Usato solo per esprimere il cuneo TOTALE implicito dopo la Leva K
    # (che taglia la sola aliquota IVS, trasmessa ~1:1 al cuneo totale).
    cuneo_ocse_base: float = 0.458
    # Sentiero spesa/PIL per ANCORE (anno, ratio), interpolazione lineare tra
    # le ancore e COSTANTE dopo l'ultima. Fonti: DFP 2026 (15.4% 2027, 15.5%
    # 2028-29, picco 17.1% 2041 con plateau fino al 2044, 16.2% 2050) e RGS
    # (convergenza a 14.0% al 2070). L'AR2024 da' 13.7% al 2070 ma su perimetro
    # UE piu' ampio: qui si usa UNA sola famiglia di fonti (RGS nazionale).
    sq_anchors: tuple = ((2027, 0.154), (2029, 0.155), (2041, 0.171),
                         (2044, 0.171), (2050, 0.162), (2070, 0.140))
    # Comportamento del sentiero OLTRE l'ultima ancora ufficiale (2070), dove non
    # esistono proiezioni pubbliche. "costante" congela il 14.0%; "trend"
    # (default, scelta utente 2026-08) prosegue la discesa al tasso medio
    # geometrico osservato nella finestra sq_trend_da..sq_trend_a. Va dichiarato
    # come STIMA propria, non come proiezione ufficiale.
    sq_estensione: str = "trend"
    sq_trend_da: int = 2045
    sq_trend_a: int = 2070
    # Pavimento: la spesa pensionistica pubblica non scende sotto questa soglia.
    # Serve a impedire che l'estrapolazione di un trend di TRANSIZIONE (uscita
    # dei baby boomer + sostituzione retributivo/contributivo, processi che si
    # esauriscono) porti su un secolo a valori privi di senso economico.
    sq_floor: float = 0.09
    # Sentiero SQ v14 (SOLO replica di validazione: lineare fino al picco 2040,
    # poi discesa lineare via 14.1% al 2060 estrapolata oltre senza fonte).
    ratio0_v14: float = 0.160
    ratio_picco_v14: float = 0.171
    anno_picco_v14: int = 2040
    ratio_2060_v14: float = 0.141
    quota_contrib: float = 0.728         # C2: quota contributiva della spesa
    # R1 — DM 20/11/2024 tabella ufficiale: eta' 69 = 6.024% (divisore 16.600).
    # Il vecchio 5.94% era una stima. NB: si cancella su P1 (via kappa) ma
    # prezza direttamente le rendite P2 (+1.4%).
    coeff69: float = 0.06024
    mensilita: int = 13
    carriera: int = 40                   # anni carriera alla partenza (27 -> 67)
    durata_media_pensione: float = 19.5  # speranza di vita a 67 (ISTAT)
    eta_pens: int = 67                   # eta' pensionabile OGGI (L. 214/2011)

    # --- ETA' PENSIONABILE DINAMICA (Task 4.1, 2026-07; corretta su input utente)
    # L'eta' parte dai 67 anni VIGENTI e sale gradualmente a 69 (eta' obiettivo
    # della riforma), +1 anno ogni eta_step_anni (~1/decennio, come il
    # meccanismo di indicizzazione alla speranza di vita della L. 335/1995).
    # VINCOLO DI POLITICA ECONOMICA (scelta utente 2026-07): il cap e' 69 e non
    # oltre — una riforma che portasse a lavorare fino a 71 anni non e'
    # proponibile. Il salire graduale 67->69 E' il "+2 anni" della riforma,
    # spalmato su decenni invece che imposto subito.
    # Carriera e coefficiente di trasformazione seguono l'eta' (40->42 anni;
    # 5.608% -> 6.024% dalla tabella DM 20/11/2024).
    # Applicata SIMMETRICAMENTE a status quo e riforma: il coefficiente si
    # cancella su dP1 via kappa -> nessun doppio conteggio con la Leva I.
    # NB (convenzione dichiarata): il sentiero RGS a legislazione vigente indicizza
    # anch'esso l'eta' oltre i 67; qui il cap a 69 e' la SCELTA DI POLICY della
    # riforma, e il risparmio incrementale rispetto al baseline resta
    # contabilizzato parametricamente nella Leva I. eta_pens_max e' un parametro:
    # alzarlo permette di testare varianti piu' aggressive.
    # 0 = eta' statica a 67 (nessuna dinamica).
    eta_dinamica: int = 1
    eta_pens_max: int = 69               # cap: eta' obiettivo della riforma
    eta_step_anni: int = 11              # anni per +1 anno di eta' pensionabile
    # Coefficienti di trasformazione ufficiali (DM 20/11/2024, tasso sconto 1.5%)
    # per eta' 67-71: usati per prezzare P1 e le rendite P2 all'eta' effettiva.
    coeff_eta_tab: tuple = ((67, 0.05608), (68, 0.05808), (69, 0.06024),
                            (70, 0.06258), (71, 0.06510))

    # --- Flusso nuovi pensionati (Task 4.2: DINAMICO, legato agli occupati) -----
    # In v14/pre-Task4 il flusso era a gradini fisso (0.55/0.60/0.48, piatto per
    # sempre dal 2038): non decresceva con la forza lavoro. Ora il flusso segue
    # flusso(t) = tasso_pens(t) x occ(t), con il tasso calibrato sui livelli INPS
    # osservati (gradini storici) e poi tenuto costante: cosi' il NUMERO di nuovi
    # pensionati scende in proporzione agli occupati (coerenza demografica).
    # flusso_dinamico=0 ripristina i gradini fissi (comportamento pre-Task4).
    flusso_dinamico: int = 1
    flusso1: float = 0.550               # mln/anno fino al 2030 (calibrazione tasso)
    flusso2: float = 0.600               # 2031-2037
    flusso3: float = 0.480               # dal 2038
    switch1: int = 2031
    switch2: int = 2038

    # --- Biforcazione contributiva ---------------------------------------------
    # (inizio_t, quota_totale, quota_brdp). Quota espressa in PUNTI DI SALARIO
    # LORDO: scavata dentro il 33% IVS (33 -> 33-quota_totale al PAYG; quota
    # totale ai fondi, di cui quota_brdp in BRDP e il resto nei fondi di mercato).
    # Fase I = 6% (4% fondi + 2% BRDP). Scatti OGNI DIECI ANNI (scelta utente
    # 2026-08): 2027, 2037, 2047, 2057 — cadenza regolare e leggibile, nello
    # spirito del Save More Tomorrow (escalation 6 -> 8 -> 10 -> 12).
    fasi: tuple = ((0, 0.06, 0.02), (10, 0.08, 0.02), (20, 0.10, 0.02), (30, 0.12, 0.0))
    # RAMPA FASI: ogni scatto di fase si raggiunge linearmente in N anni invece
    # che in un anno secco
    rampa_fasi: int = 5
    rf: float = 0.025                    # rendimento reale fondi (base, deterministico)

    # --- RENDIMENTO STOCASTICO (Task 9, 2026-07) --------------------------------
    # rf_path: sentiero ANNUO di rendimenti reali dei fondi. Se None, si usa il
    # valore scalare rf (run deterministico, comportamento storico del modello).
    # Se e' un array di lunghezza >= anni, ogni anno usa il proprio rendimento:
    # la capitalizzazione dei montanti e la dinamica del fondo diventano
    # path-dependent. E' cio' che serve per il Monte Carlo e i fan chart.
    # NB SEQUENCE-OF-RETURNS RISK: con un sentiero, l'ORDINE dei rendimenti conta
    # (un crollo poco prima del pensionamento fa piu' danno di uno all'inizio):
    # e' un rischio reale dei sistemi a capitalizzazione che il modello
    # deterministico non puo' rappresentare.
    rf_path: object = None               # np.ndarray | None (escluso da confronti)
    # Volatilita' annua del rendimento reale di un portafoglio 60/40 (~9%):
    # usata dal generatore Monte Carlo. Fonte: dati storici azioni/obbligazioni.
    rf_sigma: float = 0.09
    mc_n_sim: int = 2000                 # numero di simulazioni Monte Carlo
    mc_seed: int = 20260726              # seme (riproducibilita' accademica)
    r_brdp: float = 0.015                # spread reale BRDP su HICP
    brdp_bullet: int = 15                # rimborso bullet dopo 15 anni

    # --- Chi entra nella biforcazione ------------------------------------------
    # Soglia UNICA di eta' nell'anno di avvio: chi ha al piu' part_t0_eta anni
    # entra subito, gli altri mai. 45 anni nel 2027 significa circa 22 anni di
    # accumulazione residua, il minimo perche' il montante possa compensare la
    # minore pensione pubblica. part_t5/part_t5_eta restano solo per la replica
    # congelata, che usa la curva di partecipazione aggregata.
    part_t0_eta: int = 45
    part_t5_eta: int = 50                # solo replica congelata
    part_t5: int = 5                     # solo replica congelata

    # --- Pilastro 0 -------------------------------------------------------------
    p0_base: float = 650.0               # €/mese
    p0_cap: float = 1200.0               # €/mese: P0=0 oltre
    # P0 SOLO per le coorti che hanno biforcato (dal 2051): chi va in pensione
    # prima resta nel regime assistenziale attuale (dentro H_assist), niente
    # P0 e niente credito safety net. Evita la quasi-doppia pensione dei primi
    # decenni (P0 sopra una P1 ancora piena). 0 = P0 a tutti i nuovi (vecchio).
    p0_solo_biforcati: int = 1
    safety_net: float = 32.5             # €mld assorbiti da P0 
    p0_forfait_v14: float = 13.0         # solo replica v14
    p0_phasein_v14: int = 35             # solo replica v14

    # --- Fiscalità (C5) ----------------------------------------------------------
    irpef_contributi: float = 0.27       # Leva B (TEE sui contributi biforcati)
    irpef_pensioni: float = 0.15         # aliquota media IRPEF sulle pensioni
    regime_p2: str = "TEE"               # "TEE" (default, come v14) o "EET"

    # --- Distribuzione redditi per il taper P0 (profili x peso) ------------------
    profili: tuple = (0.5, 1.0, 1.5)
    pesi_profili: tuple = (0.50, 0.35, 0.15)

    # --- COERENZA PREVIDENZIALE / ASSISTENZIALE (Task 7, 2026-07) ----------------
    # Il modello sovrapponeva TRE split diversi. Ora sono distinti:
    #  1) quota_contrib (72.8%)  = quota di spesa FINANZIATA da contributi
    #     (concetto di FINANZIAMENTO: il resto e' coperto da fiscalita' generale);
    #  2) quota_pensionati_previdenziali = quota dei PENSIONATI che percepiscono
    #     una prestazione contributiva (gli altri hanno solo prestazioni
    #     assistenziali: assegno sociale, invalidita' civile);
    #  3) safety_net (32.5 €mld) = prestazioni assistenziali che il P0 sostituisce.
    # BUG CORRETTO: la calibrazione kappa divideva la spesa CONTRIBUTIVA per TUTTI
    # i pensionati (inclusi i soli-assistenziali, che una pensione contributiva non
    # ce l'hanno), sottostimando la pensione contributiva media di ~20% e con essa
    # dP1 e la Leva F. Ora numeratore e denominatore sono coerenti.
    # NB: la voce "H x (1-quota_contrib)" NON e' la spesa assistenziale (28.5 €mld
    # per INPS) ma la quota non coperta da contributi: rietichettata nell'output.
    quota_pensionati_previdenziali: float = 0.82

    # --- PARTECIPAZIONE: una sola fonte di verita' (Task 7) ----------------------
    # 1 = la quota di lavoratori che biforcano e' DERIVATA dalle coorti
    # (anno_adesione), 0 = usa la curva aggregata di v14 (partecipazione_L).
    # Le due divergevano fino a 7pp (33% vs 28% nel 2027, 85% vs 92% nel 2047):
    # i contributi si raccoglievano da una platea diversa da quella accreditata,
    # distorcendo il P2 pro-capite. Ora L(t) discende dalla stessa regola.
    L_da_coorti: int = 1

    # --- LIVELLO SALARIALE PER COORTE (2026-08) ---------------------------------
    # 1 = il salario di inizio carriera di ogni coorte e' w0 rivalutato fino
    # all'anno in cui quella carriera comincia; 0 = tutte le coorti partono da w0.
    # Con 0 le pensioni restano identiche in termini reali per ogni generazione,
    # mentre i contributi ai fondi crescono col salario medio dell'anno: le due
    # gambe del modello divergono e il risparmio dP1, costante in reale, si
    # erode rispetto a un PIL che cresce (la spesa pubblica in %PIL risale
    # artificiosamente sull'orizzonte lungo). La replica congelata usa sempre 0.
    salario_per_coorte: int = 1

    # --- REVERSIBILITA' (Task 7) -------------------------------------------------
    # LATO PRIVATO: il montante P2 e' CAPITALE INDIVIDUALE del lavoratore, quindi
    # passa INTEGRALMENTE al superstite (rev_alpha_p2 = 1.0): non e' una
    # prestazione pubblica reversibile al 60%, e' proprieta' che si trasmette.
    # La rendita iniziale scende di conseguenza (conversione a due teste).
    # LATO PUBBLICO: l'aliquota di reversibilita' sulla P1 scende da 60% a 45%
    # per i NUOVI pensionati — sostenibile proprio perche' il superstite eredita
    # il P2. Il risparmio e' ora ENDOGENO (dentro dP1), non piu' il forfait di
    # Leva H: per questo levaH e' azzerata di default (v. sotto).
    rev_attiva: int = 1
    rev_quota: float = 0.40       # quota di pensionati che lascia un coniuge superstite
    rev_durata: float = 6.0       # anni attesi di erogazione al superstite (gap longevita' M/F)
    rev_alpha_p2: float = 1.0     # quota del montante P2 trasferita al superstite
    rev_pub_sq: float = 0.60      # aliquota reversibilita' pubblica, status quo (legge)
    rev_pub_rif: float = 0.45     # aliquota reversibilita' pubblica, riforma (nuovi)
    # IMPOSTA UNA TANTUM DI SUCCESSIONE sul capitale P2 trasferito al superstite.
    # Razionale: in regime TEE il montante e' gia' stato tassato a monte (Leva B),
    # quindi il trasferimento al coniuge e' un evento SUCCESSORIO, non reddito.
    # Scaglioni progressivi (soglia_superiore, aliquota) sul capitale trasferito.
    rev_tax_attiva: int = 1
    # Scaglioni (soglia_superiore, aliquota marginale) sul capitale trasferito.
    # Scelta utente 2026-08: franchigia ampia (50.000 €) e aliquota marginale
    # massima contenuta (15%). La franchigia esclude di fatto i montanti piccoli
    # — carriere brevi, redditi bassi — cosi' il prelievo colpisce solo le
    # successioni di importo rilevante.
    rev_tax_scaglioni: tuple = ((50000.0, 0.00), (100000.0, 0.05),
                                (200000.0, 0.10), (1e18, 0.15))
    # BASE IMPONIBILE (scelta utente 2026-07: "montante"):
    #  "montante" = intero montante individuale accumulato -> e' l'eredita' del
    #               CONTO individuale che passa al coniuge (~160k, gettito ~€mld);
    #  "riserva"  = valore attuale della sola rendita reversibile (~47k): e'
    #               l'entitlement effettivamente trasferito se il prodotto e' una
    #               rendita a due teste. Gettito quasi nullo.
    # TENSIONE DICHIARATA: con rev_alpha_p2=1.0 la rendita PROSEGUE al superstite
    # (non c'e' liquidazione del capitale), quindi "montante" tassa un valore
    # nozionale piu' ampio dell'entitlement strettamente trasferito. E' una scelta
    # di policy (trattare il conto come massa ereditaria), non un'identita'
    # attuariale: va dichiarata nel paper.
    rev_tax_base: str = "montante"

    # --- Leve parametriche v15 (razionalizzazione R2, 2026-07) -------------------
    # A  Imposta una-tantum sulla conversione dello STOCK dei fondi pensione
    #    esistenti (262 €mld COVIP Relazione 2026, fine 2025; 30% converte;
    #    imposta 10%; finestra 3 anni).
    # G  Sotto-indicizzazione progressiva delle pensioni medio-alte (perequazione,
    #    entro sentenza C.Cost. 167/2025); decade -1,5%/anno perche' lo stock di
    #    pensioni retributive alte si estingue per mortalita'.
    # H  Revisione delle pensioni di reversibilita' ai superstiti (aggancio
    #    attuariale ai contributi del de cuius): ~5 mld/anno.
    # I  ETA' PENSIONABILE (assorbe la ex Leva C "67->69"): risparmio del wedge
    #    di +2 anni OLTRE il sentiero indicizzato alla speranza di vita, che il
    #    baseline RGS gia' incorpora. Rampa lineare in levaI_rampa anni fino a
    #    levaI_regime, poi COSTANTE (il wedge persiste: entrambi i sentieri,
    #    baseline e riforma, continuano a indicizzarsi). Con baseline_politico=1
    #    si aggiunge la vecchia componente "anti-sterilizzazione" (ripristino
    #    degli adeguamenti sospesi/sterilizzati dal 2019: 2.75 + 0.14/biennio),
    #    legittima SOLO dichiarando un baseline politico in cui le
    #    sterilizzazioni continuano (default OFF per prudenza accademica).
    # J  CHIUSURA DEROGHE E ANTICIPAZIONI RESIDUE (assorbe la ex componente
    #    "deroghe" della Leva C e la vecchia Leva J): Quota 103/APe/Opzione
    #    Donna: 3.5 €mld dal 2027 con decadimento lineare a zero in 10 anni
    #    (vs legislazione vigente le deroghe scadono da sole: il risparmio e'
    #    l'effetto-coorte transitorio, esaurito entro il ~2037).
    # ELIMINATE dal motore v15 (R2): Leva C (doppio conteggio col baseline
    #    indicizzato; assorbita da I e J) e Leva E (crediti natalita': costo
    #    puro senza calcolo del "delta figli", nessun ritorno modellato).
    # (Nomi completi in LEVE_NOMI, usato per le intestazioni dell'output.)
    levaA_stock: float = 262.0; levaA_convers: float = 0.30   # A: conversione stock fondi
    levaA_tax: float = 0.10;   levaA_anni: int = 3            # A: imposta 10%, finestra 3 anni
    levaG: float = 6.0; levaG_decad: float = 0.015           # G: sotto-indicizzazione (perequazione)
    # H AZZERATA (Task 7): il risparmio da revisione della reversibilita' e' ora
    # ENDOGENO — sta dentro dP1, come differenza fra la reversibilita' pubblica
    # dello status quo (60% di una P1 piena) e quella della riforma (45% di una
    # P1 ridotta). Tenere anche il forfait di 5 €mld sarebbe doppio conteggio.
    # Rimettere a 5.0 solo se si disattiva rev_attiva.
    levaH: float = 0.0; levaH_avvio: int = 1                 # H: ora endogena (v. reversibilita')
    # I AZZERATA (2026-07): il sentiero RGS a legislazione vigente porta l'eta' a
    # 68a11m nel 2050 e a 70 nel 2067 da solo, ed e' GIA' dentro H(t). Il nostro
    # sentiero modellato (67->69, cap 69) coincide con quello del baseline fino al
    # 2050 e resta SOTTO dopo: rivendicare 9 €mld/anno per "aver alzato l'eta'"
    # significa contare un risparmio gia' incorporato nella linea di confronto.
    # Rimettere a 9.0 solo dichiarando un baseline politico diverso (eta' congelata).
    levaI_regime: float = 0.0            # I: €mld a regime (wedge oltre il sentiero indicizzato)
    levaI_rampa: int = 15                # I: anni di phase-in (ex levaC_anni)
    baseline_politico: int = 0           # I: 1 = aggiunge componente anti-sterilizzazione
    levaJ_iniziale: float = 3.5          # J: €mld iniziali (ex levaC_deroghe)
    levaJ_decad_anni: int = 10           # J: anni di esaurimento coorti aventi diritto

    # --- Parametri leve CONGELATI, usati SOLO dalla replica v14 (formule v14) ----
    # (levaI/levaI_step/levaI_avvio sono riusati anche dalla componente
    #  anti-sterilizzazione della nuova Leva I quando baseline_politico=1.)
    levaC_deroghe: float = 3.5; levaC_eta: float = 9.0; levaC_anni: int = 15
    levaE_nascite: float = 0.369; levaE_credito: float = 1550.0; levaE_anni: int = 5
    levaI: float = 2.75; levaI_step: float = 0.14; levaI_avvio: int = 1
    levaJ: float = 1.0; levaJ_anni: int = 3

    # --- Ricalibrazione esplicita P1 (canale 4) -----------------------------------
    # quota della rendita P2 che riduce ulteriormente la P1 (clawback di policy).
    # 0.7 (scelta utente 2026-07): il pensionato a regime resta ~al livello SQ
    # (~1.270 € vs 1.179 SQ, coorte 2080); il dividendo rf>g va a spesa pubblica
    # (-0.8pp PIL al 2080) e, via Leva K, a taglio del cuneo agganciato al calo P1.
    # Framing: ricalibrazione del coefficiente NDC per le coorti biforcate.
    # 0.45 (scelta utente 2026-07, dopo la correzione della kappa): con la P1
    # calibrata correttamente (+22%), il vecchio 0.7 tagliava la pensione totale
    # dell'8.8% sotto lo status quo. Con 0.45 la riforma tornava ~neutrale
    # sull'adeguatezza, al prezzo di un debito di transizione piu' alto.
    # 0.3727 (scelta utente 2026-08): l'aliquota e' la REGOLA — "si preleva il
    # 37,27% della rendita P2" — e il vantaggio del pensionato ne e' la
    # CONSEGUENZA, non il target. Calibrata perche' la coorte a regime (2080)
    # risulti ~50 €/mese sopra lo status quo. Insieme alla salvaguardia
    # (clawback_floor) da': nessuna coorte sotto lo status quo, +50 a regime,
    # profilo basso +20,5%, profili medio e alto +2,8%.
    # NB: senza salvaguardia questa stessa aliquota lascerebbe le coorti
    # 2051-2063 fino a 23 €/mese sotto lo status quo, perche' il loro montante
    # e' ancora piccolo (carriera solo in parte biforcata).
    clawback_p2: float = 0.3727

    # --- MODO DI RICALIBRAZIONE DELLA P1 ------------------------------------------
    # "costante": si preleva sempre la quota fissa clawback_p2 della rendita P2.
    #   Semplice, ma con un'aliquota unica NESSUN valore soddisfa insieme le due
    #   condizioni "nessuna coorte sotto lo status quo" e "nessuna coorte molto
    #   sopra": le coorti di transizione hanno un montante piccolo (carriera solo
    #   in parte biforcata) e la stessa aliquota che le lascia neutre regala
    #   ~170 €/mese alle coorti a regime.
    # "target": la P1 residua e' fissata in modo che la pensione complessiva del
    #   lavoratore di profilo medio sia pari alla pensione di status quo piu'
    #   clawback_target euro al mese; il resto del dividendo della capitalizzazione
    #   resta alle finanze pubbliche. Il prelievo non puo' mai essere negativo (lo
    #   Stato non integra) ne' superare la P1 residua: e' un TETTO al vantaggio,
    #   non una garanzia sul minimo. Se il rendimento e' basso e il montante non
    #   compensa la minore P1, la pensione resta sotto il target e la differenza
    #   resta a carico del lavoratore.
    clawback_modo: str = "costante"
    clawback_target: float = 20.0        # €/mese sopra lo status quo (modo "target")
    # CLAUSOLA DI SALVAGUARDIA (modo "costante"): il prelievo si riduce quanto
    # basta perche' la pensione complessiva del profilo medio non scenda sotto
    # lo status quo maggiorato di clawback_floor euro. Serve alle coorti di
    # TRANSIZIONE: chi va in pensione nei primi anni ha un montante ancora
    # piccolo (carriera solo in parte biforcata), quindi un'aliquota calibrata
    # sulle coorti a regime le lascerebbe sotto. Con la salvaguardia l'aliquota
    # resta la regola generale e il vantaggio a regime emerge da solo.
    # None disattiva la salvaguardia (aliquota nuda).
    clawback_floor: float | None = 0.0   # €/mese minimi sopra lo status quo

    # --- LEVA K — riduzione del cuneo contributivo (RIDISEGNO Task 3, 2026-07) ---
    # NUOVA logica: il taglio del cuneo e' ancorato al CALO DEL MONTE P1 PUBBLICO,
    # misurato come riduzione_p1(t) = dP1_lordo(t) / H_contrib(t) — la frazione
    # del pilastro contributivo pubblico che le coorti biforcate hanno gia'
    # sostituito. Economia: se il pilastro 1 eroga meno pensioni contributive,
    # la pressione contributiva puo' scendere in proporzione (dividendo della
    # riforma restituito ai lavoratori). sconto_pp(t) = min(levaK_max,
    # levaK_coeff x riduzione_p1(t)).
    # VECCHIO difetto rimosso: non si aggancia piu' al surplus di bilancio (che
    # azzerava meccanicamente il saldo, rendendolo illeggibile). Il taglio e'
    # ora un costo di policy vero, finanziato AL PIU' da una quota del dividendo
    # netto di P1 dell'anno (dP1 - IRPEF persa): cosi' non fa esplodere il debito
    # e il saldo totale resta informativo. Si contabilizza solo la perdita di
    # gettito, NON il minor accrual NDC futuro (risparmio prudenzialmente non
    # contato). Il driver e' monotono e liscio -> nessuna instabilita' numerica.
    # Opzione non attivata (default): un lieve AUMENTO del cuneo prima del 2040
    # (quando riduzione_p1=0) e' politicamente oneroso e qui volutamente escluso.
    # DISATTIVATA (scelta utente 2026-08). Ragione: il taglio dell'aliquota
    # contributiva non e' una conseguenza endogena della riforma ma una decisione
    # fiscale autonoma, che il modello non e' in grado di giustificare da sola;
    # e costa molto (25,7 €mld al 2080, cioe' quasi tutto il disavanzo di quell'
    # anno). Tenerla accesa faceva apparire come fallimento della riforma
    # ("non si autofinanzia") cio' che era invece una scelta di destinazione del
    # dividendo. Rimettere a 1 solo dichiarando esplicitamente quella scelta.
    levaK_attiva: int = 0
    levaK_coeff: float = 0.15           # pp di cuneo per punto di riduzione frazionale del monte P1
    levaK_max: float = 0.06            # taglio massimo: 6 punti (33% -> 27%)
    levaK_quota_dividendo: float = 0.5  # quota del dividendo netto P1 destinata al taglio (resto: riduce il debito)

    # --- LEVA L — recupero tax gap contributivo (era in v14 ma FUORI dal saldo) --
    # MEF-RGS Relazione evasione 2025: gap contributivo 10 €mld/anno (2.4 lavoratore
    # + 7.6 datore). Recuperabile in 5-10 anni con riforma riscossione/incrocio
    # banche dati. Default OFF nel base (prudenza); ON nel pacchetto anticipato.
    levaL_attiva: int = 0
    levaL_gap: float = 10.0              # €mld/anno a regime
    levaL_rampa: int = 10                # anni per il pieno recupero

    # --- Finanziamento -----------------------------------------------------------
    r_debito: float = 0.015              # tasso reale sul debito di transizione

    def anni_range(self):
        return np.arange(self.anni)

    def anno(self, t):
        return self.anno0 + t


def params_pessimistico() -> Params:
    """Scenario pessimistico: crescita PIL 0.3%, salari fermi, fondi a 1.25%."""
    p = Params()
    p.g = 0.003
    p.g_w = 0.0
    p.rf = 0.0125
    return p


def params_v14() -> Params:
    """Parametri CONGELATI di v14, usati SOLO dalla replica di validazione.
    Non ereditano le modifiche fatte ai default di Params: la replica deve
    sempre confrontarsi con il file v14 originale, qualunque calibrazione
    si scelga per v15. (Il run dell'utente con fasi modificate mostrava
    max|diff|=9.4 su Leva F proprio perche' la replica usava i nuovi fasi.)"""
    p = Params()
    # Fase III q_brdp = 0.0 (in v10/v14 la Fase III ha "0 BRDP + 10 P2"): il
    # valore 0.2 presente in una revisione intermedia era un refuso (20% del
    # salario in BRDP) che rompeva la validazione della colonna AE.
    p.fasi = ((0, 0.06, 0.02), (10, 0.08, 0.02), (20, 0.10, 0.0), (30, 0.12, 0.0))
    p.rampa_fasi = 0                     # v14: fasi a gradino
    p.p0_cap = 1250.0
    # R1/Task7: congela i valori v14 dei parametri modificati nel 2026-07.
    # levaH resta 5.0 (in v15 e' azzerata perche' la reversibilita' e' endogena);
    # rev_attiva/L_da_coorti spenti: v14 non li aveva.
    p.levaH = 5.0
    p.rev_attiva = 0
    p.rev_tax_attiva = 0
    p.L_da_coorti = 0
    p.pensionati0 = 17.0
    p.pensioni_vigenti0 = 22.3
    p.coeff69 = 0.0594
    p.levaA_stock = 244.0
    p.rf = 0.025
    p.r_brdp = 0.015
    p.g = 0.008
    p.g_w = 0.002
    p.pil0 = 2237.0
    p.w0 = 26700.0
    p.occupati0 = 17.5
    p.n_occ = -0.005
    p.aliquota = 0.33
    p.irpef_contributi = 0.27
    p.p0_forfait_v14 = 13.0
    p.p0_phasein_v14 = 35
    p.brdp_bullet = 15
    p.carriera = 42
    return p


# ------------------------------------------------------------------------------
# 2. BLOCCHI COMUNI (macro, biforcazione, leve parametriche)
# ------------------------------------------------------------------------------

def quota_bif(p: Params, t: int) -> tuple[float, float]:
    """(quota totale, quota BRDP) biforcata all'anno t.
    Con rampa_fasi>0 ogni fase converge linearmente al target in N anni
    partendo dal target della fase precedente (0 per la prima fase)."""
    rampa = getattr(p, "rampa_fasi", 0)
    tot, brdp = 0.0, 0.0
    prev_tot, prev_brdp = 0.0, 0.0
    for t0, q_tot, q_brdp in sorted(p.fasi):
        if t < t0:
            break
        f = min(1.0, (t - t0 + 1) / rampa) if rampa > 0 else 1.0
        tot = prev_tot + (q_tot - prev_tot) * f
        brdp = prev_brdp + (q_brdp - prev_brdp) * f
        prev_tot, prev_brdp = q_tot, q_brdp
    return tot, brdp


def macro(p: Params, sq_ratio_v14: bool):
    """Serie macro. sq_ratio_v14=True replica v14: estrapolazione lineare del
    ratio SQ e formula immigrazione v14 (0.116 mln con fattore (1-n)^(t-1) che
    compone il flusso). In v15 (C9): flusso migratorio costante e decomposto
    (176k x 70% eta' lavorativa x 65% occupazione = 80k occupati/anno)."""
    T = p.anni
    t = np.arange(T)
    pil = p.pil0 * (1 + p.g) ** t
    w = p.w0 * (1 + p.g_w) ** t
    occ = np.empty(T); occ[0] = p.occupati0
    for i in range(1, T):
        if sq_ratio_v14:
            occ[i] = occ[i-1] * (1 + p.n_occ) + p.migrazione_v14 * (1 - p.n_occ) ** (i - 1)
        else:
            occ[i] = occ[i-1] * (1 + p.n_occ) + p.migrazione
    # stock pensioni vigenti: crescita fino al picco, poi lieve declino (Task 4.4)
    tpk = p.anno_picco_pens - p.anno0
    pens_vig = np.empty(T)
    for i in range(T):
        if i <= tpk:
            pens_vig[i] = p.pensioni_vigenti0 * (1 + p.g_pensioni) ** i
        else:
            pens_vig[i] = (p.pensioni_vigenti0 * (1 + p.g_pensioni) ** tpk
                           * (1 + p.g_pensioni_post) ** (i - tpk))

    if sq_ratio_v14:
        # sentiero v14 (SOLO replica): lineare fino al picco 2040, poi discesa
        # lineare via 14.1% al 2060 ESTRAPOLATA oltre l'orizzonte (senza fonte)
        tp = p.anno_picco_v14 - p.anno0
        t60 = 2060 - p.anno0
        ratio = np.empty(T)
        for i in range(T):
            if i <= tp:
                ratio[i] = p.ratio0_v14 + (p.ratio_picco_v14 - p.ratio0_v14) * i / tp
            else:
                ratio[i] = (p.ratio_picco_v14 +
                            (p.ratio_2060_v14 - p.ratio_picco_v14) * (i - tp) / (t60 - tp))
    else:
        # C1/R1: interpolazione lineare tra le ancore ufficiali DFP 2026/RGS.
        # np.interp tiene COSTANTE il ratio dopo l'ultima ancora (14.0% dal 2070).
        anni_anc = np.array([a for a, _ in p.sq_anchors], dtype=float)
        val_anc = np.array([v for _, v in p.sq_anchors], dtype=float)
        anni_cal = p.anno0 + t.astype(float)
        ratio = np.interp(anni_cal, anni_anc, val_anc)
        if p.sq_estensione == "trend":
            # PROSECUZIONE DEL TREND oltre l'ultima ancora ufficiale. Il tasso
            # e' quello medio geometrico osservato nella finestra di discesa
            # (sq_trend_da..sq_trend_a), applicato a partire dall'ultima ancora.
            # Motivazione: tenere il ratio costante implicherebbe che la spesa
            # cresca esattamente come il PIL, mentre nel modello i pensionati
            # calano e le pensioni crescono come i salari (g_w < g). Il tasso
            # cosi' stimato coincide quasi con quello implicito in quelle due
            # dinamiche, il che rende l'estrapolazione coerente e non arbitraria.
            ult_anno = float(anni_anc[-1])
            r_da = float(np.interp(p.sq_trend_da, anni_anc, val_anc))
            r_a = float(np.interp(p.sq_trend_a, anni_anc, val_anc))
            n_anni = p.sq_trend_a - p.sq_trend_da
            tasso = (r_a / r_da) ** (1.0 / n_anni) - 1.0 if n_anni > 0 else 0.0
            oltre = anni_cal > ult_anno
            ratio = np.where(oltre,
                             val_anc[-1] * (1.0 + tasso) ** (anni_cal - ult_anno),
                             ratio)
            # pavimento prudenziale: nessun sistema pensionistico pubblico
            # scende sotto una soglia fisiologica, e l'estrapolazione di un
            # trend di transizione su un secolo non e' credibile senza limite.
            ratio = np.maximum(ratio, p.sq_floor)
    H = pil * ratio

    # occupati autonomi (Task 8): stessa dinamica migratoria dei dipendenti ma
    # con trend proprio. Zero se perimetro L1 o replica v14.
    occ_aut = np.zeros(T)
    if not sq_ratio_v14 and p.perimetro == "universale":
        occ_aut[0] = p.occ_autonomi0
        for i in range(1, T):
            occ_aut[i] = occ_aut[i-1] * (1 + p.g_occ_autonomi) + 0.3 * p.migrazione
    return dict(pil=pil, w=w, occ=occ, occ_aut=occ_aut, pens_vig=pens_vig,
                ratio=ratio, H=H)


def partecipazione_L(p: Params, t: int) -> float:
    """Quota lavoratori in P2 (phase-in v14): 33% -> 60% (t=5) -> 100% (t=29)."""
    if t <= 5:
        return 0.33 + (0.60 - 0.33) / 5 * t
    return min(1.0, 0.60 + (1.0 - 0.60) / 24 * (t - 5))


# Nomi economici delle leve ATTIVE nel motore v15 (post R2). Usati nelle
# intestazioni dell'output e nella documentazione. Chiave = lettera della leva.
# (C ed E sono state eliminate — v. docstring R2; restano nella replica v14.)
LEVE_NOMI = {
    "A": "Imposta una-tantum su conversione stock fondi pensione",
    "B": "Recupero IRPEF sui contributi biforcati (regime TEE)",
    "D": "Costo transizione: contributi sottratti alla ripartizione (gap biforcazione)",
    "F": "Risparmio netto sostituzione P1->P0/P2 (pilastro 0 sostitutivo)",
    "G": "Sotto-indicizzazione progressiva pensioni medio-alte (perequazione)",
    "H": "Revisione reversibilita' ai superstiti (aggancio attuariale)",
    "I": "Eta' pensionabile: +2 anni oltre il sentiero indicizzato (ex Leva C eta')",
    "J": "Chiusura deroghe e anticipazioni residue (ex Leva C deroghe + ex J)",
    "K": "Riduzione cuneo contributivo (dividendo rf>g)",
    "L": "Recupero gap contributivo (evasione)",
}


def leve_parametriche_v14(p: Params, T: int):
    """Leve A, C, E, G, H, I, J con le formule ORIGINALI di v14.
    Usata SOLO dalla replica di validazione (con params_v14 congelati)."""
    t = np.arange(T)
    A = np.where(t < p.levaA_anni,
                 p.levaA_stock * p.levaA_convers * p.levaA_tax / p.levaA_anni, 0.0)
    C = np.minimum(p.levaC_deroghe + p.levaC_eta,
                   p.levaC_deroghe + p.levaC_eta * (t + 1) / p.levaC_anni)
    E = -(p.levaE_nascite * p.levaE_credito / 1000.0) * np.minimum(t + 1, p.levaE_anni)
    G = p.levaG * (1 - p.levaG_decad) ** t
    Hlev = np.where(t >= p.levaH_avvio, p.levaH, 0.0)
    I = np.where(t >= p.levaI_avvio,
                 p.levaI + p.levaI_step * ((t - p.levaI_avvio) // 2), 0.0)
    J = np.where(t < p.levaJ_anni, p.levaJ, 0.0)
    return A, C, E, G, Hlev, I, J


def leve_parametriche_v15(p: Params, T: int):
    """Leve parametriche del motore v15 dopo la razionalizzazione R2 (2026-07):
    A, G, H invariate; C ed E ELIMINATE; I e J ridisegnate (v. Params).
    Economia delle nuove leve:
    - I: il baseline RGS gia' indicizza l'eta' alla speranza di vita; il
      risparmio incrementale della riforma e' il WEDGE di +2 anni (69 vs
      sentiero indicizzato), ~4.5 €mld per anno di eta' -> 9 €mld a regime,
      raggiunti con rampa lineare (phase-in graduale come la ex Leva C).
      Componente anti-sterilizzazione aggiunta solo se baseline_politico=1.
    - J: chiudere Quota 103/APe/OD produce un risparmio-coorte transitorio
      (le platee si esauriscono): 3.5 €mld che decadono a zero in 10 anni."""
    t = np.arange(T)
    A = np.where(t < p.levaA_anni,
                 p.levaA_stock * p.levaA_convers * p.levaA_tax / p.levaA_anni, 0.0)
    G = p.levaG * (1 - p.levaG_decad) ** t
    Hlev = np.where(t >= p.levaH_avvio, p.levaH, 0.0)
    I = p.levaI_regime * np.minimum(1.0, (t + 1) / p.levaI_rampa)
    if p.baseline_politico:
        I = I + np.where(t >= p.levaI_avvio,
                         p.levaI + p.levaI_step * ((t - p.levaI_avvio) // 2), 0.0)
    J = p.levaJ_iniziale * np.maximum(0.0, 1.0 - t / p.levaJ_decad_anni)
    return A, G, Hlev, I, J


def flusso_pensionati_step(p: Params, anno: int) -> float:
    """Flusso di nuovi pensionati a gradini (calibrazione INPS: 0.55/0.60/0.48)."""
    if anno < p.switch1:
        return p.flusso1
    if anno < p.switch2:
        return p.flusso2
    return p.flusso3


def costruisci_flussi(p: Params, occ: np.ndarray) -> dict:
    """Flusso di nuovi pensionati per anno di pensionamento (Task 4.2).
    Se flusso_dinamico: flusso(t) = tasso_pens x occ(t), con tasso_pens
    calibrato in modo che nella finestra storica riproduca i gradini INPS e a
    regime resti costante -> il NUMERO di nuovi pensionati evolve con la forza
    lavoro invece di restare piatto per sempre (i gradini fissi erano il difetto
    segnalato). Per gli anni oltre l'orizzonte (coorti che si pensionano dopo il
    2080, servono ad alloca_montanti) si estende l'ultimo occ disponibile.
    Se OFF: torna ai gradini fissi (comportamento pre-Task4)."""
    T = p.anni
    anni = range(p.anno0, p.anno0 + T + p.carriera + 5)
    if not p.flusso_dinamico:
        return {y: flusso_pensionati_step(p, y) for y in anni}
    occ_regime = occ[min(p.switch2 - p.anno0, T - 1)]     # occ all'inizio del regime
    tasso_regime = p.flusso3 / occ_regime                 # tasso di pensionamento a regime
    flussi = {}
    for y in anni:
        i = y - p.anno0
        occ_y = occ[i] if 0 <= i < T else occ[-1]
        if y < p.switch2:
            flussi[y] = flusso_pensionati_step(p, y)      # finestra storica: gradini INPS
        else:
            flussi[y] = tasso_regime * occ_y              # a regime: tasso x occupati
    return flussi


# ------------------------------------------------------------------------------
# 3. MOTORE REPLICA v14 (per validazione)
# ------------------------------------------------------------------------------

def _montante(p: Params, anno_pens: int, rate_fn, r: float, K: int,
              per_coorte: bool = False) -> float:
    """Montante contributivo di una carriera di K anni che termina in anno_pens.

    ECONOMIA: ogni anno k della carriera il lavoratore accantona
    w(k) * aliquota(k); il gettone viene poi rivalutato per gli anni che
    mancano al pensionamento al tasso r. Nel PAYG contributivo (NDC) r e' il
    tasso di rivalutazione legale (qui g, la crescita del PIL: e' la regola
    italiana); nella capitalizzazione r e' il rendimento di mercato.
        montante = SUM_k  w(k) * aliquota(cal_k) * (1+r)^(K-k)

    LIVELLO SALARIALE DELLA COORTE (per_coorte=True): il salario del primo anno
    di carriera non e' w0 per tutte le coorti, ma w0 rivalutato fino all'anno in
    cui quella carriera comincia. Senza questa correzione ogni generazione
    ripartirebbe dallo stesso salario reale e le pensioni sarebbero identiche
    per sempre, mentre i contributi che alimentano i fondi sono calcolati sul
    salario medio dell'anno di calendario, che invece cresce: le due gambe del
    modello userebbero due dinamiche salariali diverse.
    per_coorte=False riproduce la convenzione delle formule congelate (replica).
    """
    k = np.arange(1, K + 1)
    cal = anno_pens - K + k
    rates = np.fromiter((rate_fn(int(c)) for c in cal), dtype=float, count=K)
    w_iniziale = p.w0
    if per_coorte:
        # anni fra l'anno base e l'inizio della carriera (negativo se la
        # carriera e' cominciata prima dell'anno base)
        w_iniziale = p.w0 * (1 + p.g_w) ** (cal[0] - p.anno0)
    return float((w_iniziale * (1 + p.g_w) ** (k - 1) * rates
                  * (1 + r) ** (K - k)).sum())


def carriera_ndc_v14(p: Params, anno_pens: int, rate_fn, carriera: int | None = None,
                     per_coorte: bool = False):
    """Montante NDC (rivalutato a g, regola italiana). carriera=None usa
    p.carriera (default v14); il motore v15 passa la carriera della coorte
    (piu' lunga se l'eta' pensionabile e' salita). per_coorte scala il livello
    salariale iniziale con l'anno in cui la carriera comincia."""
    return _montante(p, anno_pens, rate_fn, p.g,
                     p.carriera if carriera is None else carriera, per_coorte)


def montante_capitalizzato_v14(p: Params, anno_pens: int, rate_fn, r: float,
                               carriera: int | None = None):
    """Montante a CAPITALIZZAZIONE (rivalutato al rendimento di mercato r)."""
    return _montante(p, anno_pens, rate_fn, r,
                     p.carriera if carriera is None else carriera)


# --- ETA' PENSIONABILE DINAMICA e coefficienti per eta' (Task 4.1) -------------

def eta_pensionabile(p: Params, anno_pens: int) -> int:
    """Eta' pensionabile della riforma per la coorte che si pensiona in anno_pens.
    Se eta_dinamica: parte dai 67 anni vigenti e sale di +1 ogni eta_step_anni
    fino al cap eta_pens_max = 69 (eta' obiettivo della riforma). Il salire
    graduale 67->69 e' il "+2 anni" spalmato su decenni; il cap a 69 e' un
    vincolo di proponibilita' politica, non tecnico."""
    if not p.eta_dinamica:
        return p.eta_pens
    incr = max(0, (anno_pens - p.anno0)) // p.eta_step_anni
    return min(p.eta_pens + incr, p.eta_pens_max)


def carriera_pens(p: Params, anno_pens: int) -> int:
    """Carriera della coorte: base + gli anni di eta' guadagnati (chi va in
    pensione piu' tardi ha lavorato piu' a lungo)."""
    return p.carriera + (eta_pensionabile(p, anno_pens) - p.eta_pens)


def coeff_reversibile(p: Params, coeff: float) -> float:
    """Coefficiente di conversione a DUE TESTE (joint-and-survivor).

    ECONOMIA: se la rendita prosegue al superstite, il fondo paga per piu' anni,
    quindi la rendita iniziale deve scendere perche' il valore attuale resti lo
    stesso (operazione attuarialmente NEUTRA per il fondo: non crea ne' distrugge
    ricchezza, la redistribuisce fra titolare e superstite).

        coeff_JS = coeff / (1 + p_rev * alpha * a_rev / a_single),
        con a_single = 1/coeff  (e' il DIVISORE della tabella DM)

    Con p_rev=0.40, alpha=1.0, a_rev=6, a_single=16.6 (eta' 69):
    fattore = 1 + 0.40*1.0*6/16.6 = 1.145 -> coeff 6.024% -> 5.26% (-12.6%).
    """
    if not p.rev_attiva or coeff <= 0:
        return coeff
    a_single = 1.0 / coeff
    return coeff / (1.0 + p.rev_quota * p.rev_alpha_p2 * p.rev_durata / a_single)


def imposta_successione_p2(p: Params, capitale: float) -> float:
    """Imposta UNA TANTUM progressiva sul capitale P2 trasferito al superstite.

    RAZIONALE: in regime TEE i contributi biforcati sono gia' stati tassati
    all'ingresso (Leva B, 27%): la rendita non e' reddito nuovo. Il passaggio del
    montante al coniuge e' percio' un evento SUCCESSORIO, e si tassa come tale,
    per scaglioni progressivi sul capitale trasferito (0% fino a 25k, poi 5/10/
    15/18%). Progressiva = protegge i montanti piccoli, che sono la maggioranza.
    """
    if not p.rev_tax_attiva or capitale <= 0:
        return 0.0
    tot, prec = 0.0, 0.0
    for soglia, aliq in p.rev_tax_scaglioni:
        if capitale <= prec:
            break
        tot += (min(capitale, soglia) - prec) * aliq
        prec = soglia
    return tot


def rf_serie(p: Params, n: int) -> np.ndarray:
    """Serie annua dei rendimenti reali dei fondi, lunga n.

    Se rf_path e' None -> costante al valore scalare p.rf (run deterministico).
    Se e' un array piu' corto di n, si estende con la MEDIA del sentiero (serve
    per le coorti che si pensionano oltre l'orizzonte di simulazione).
    """
    if p.rf_path is None:
        return np.full(n, p.rf, dtype=float)
    path = np.asarray(p.rf_path, dtype=float)
    if path.size >= n:
        return path[:n]
    return np.concatenate([path, np.full(n - path.size, float(path.mean()))])


def fattori_capitalizzazione(rf_ser: np.ndarray) -> np.ndarray:
    """Fattori cumulati C[k] = prod_{j<k} (1+rf[j]), con C[0]=1.

    Capitalizzare un flusso dall'anno i all'anno y costa C[y]/C[i]: con
    rendimento costante si riduce a (1+rf)^(y-i), con un sentiero cattura
    l'effetto dell'ORDINE dei rendimenti (sequence-of-returns risk).
    """
    return np.concatenate([[1.0], np.cumprod(1.0 + rf_ser)])


def fattore_rendita(anni: float, r: float) -> float:
    """Fattore di annualita' a(n,r) = (1-(1+r)^-n)/r: valore attuale di 1 euro
    l'anno per 'anni' anni. Serve a capitalizzare la rendita del superstite nel
    capitale trasferito (base imponibile dell'imposta di successione)."""
    if r <= 0:
        return anni
    return (1.0 - (1.0 + r) ** (-anni)) / r


def partecipazione_da_coorti(p: Params, flussi: dict) -> np.ndarray:
    """Quota di lavoratori che biforcano, DERIVATA dalle coorti (Task 7).

    In ogni anno di calendario sono al lavoro le coorti che si pensioneranno
    entro una carriera; di queste, biforcano quelle che hanno gia' aderito
    (anno_adesione). La quota e' pesata per la numerosita' delle coorti.
    Elimina la doppia fonte di verita' con la curva aggregata di v14, che
    divergeva fino a 7pp e distorceva il P2 pro-capite.
    """
    T = p.anni
    L = np.zeros(T)
    ade_cache = {}
    for i in range(T):
        cal = p.anno0 + i
        num = den = 0.0
        for y in range(cal + 1, cal + p.carriera + 1):
            f = flussi.get(y, 0.0)
            if f <= 0:
                continue
            den += f
            if y not in ade_cache:
                ade_cache[y] = anno_adesione(p, y)
            ade = ade_cache[y]
            if ade is not None and ade <= cal:
                num += f
        L[i] = num / den if den > 0 else 0.0
    return L


def coeff_eta(p: Params, eta: int) -> float:
    """Coefficiente di trasformazione (DM 20/11/2024) all'eta' data.
    Fuori tabella: usa il valore estremo piu' vicino (67 sotto, 71 sopra)."""
    tab = dict(p.coeff_eta_tab)
    if eta in tab:
        return tab[eta]
    eta_min, eta_max = min(tab), max(tab)
    if eta <= eta_min:
        return tab[eta_min]
    if eta >= eta_max:
        return tab[eta_max]
    return tab[eta]


def bif_tot_cal_v14(p: Params, cal: int) -> float:
    """Quota biforcata TOTALE nell'anno di calendario cal (formula AK di v14:
    si applica a chiunque, senza vincolo di partecipazione per eta')."""
    if cal < p.anno0:
        return 0.0
    return quota_bif(p, cal - p.anno0)[0]


def _verifica_params_v14(p: Params) -> None:
    """Guardia contro regressioni sui parametri CONGELATI della replica.

    Il refuso q_brdp = 0.2 nella Fase III (20% del salario in BRDP, invece del
    2% o dello 0 delle fasi finali) e' gia' ricomparso DUE volte nel file,
    rompendo silenziosamente la validazione delle colonne AE/AF: la replica
    produceva numeri plausibili ma sbagliati (primo anno divergente 2047, scarto
    interamente sulla voce AD = rimborso bullet 15 anni dopo). Qui il modello
    fallisce subito e in modo esplicito, invece di stampare risultati errati.
    """
    for t0, q_tot, q_brdp in p.fasi:
        assert 0.0 <= q_brdp <= 0.02, (
            f"params_v14 CORROTTO: quota BRDP = {q_brdp:.2%} nella fase da t={t0}. "
            f"In v10/v14 vale 2% (o 0 nelle fasi finali). Il refuso tipico e' "
            f"0.2 al posto di 0.0/0.02 e rompe la validazione AE/AF.")
        assert 0.0 <= q_brdp <= q_tot <= 0.5, (
            f"params_v14 CORROTTO: fase t={t0} incoerente (q_tot={q_tot}, q_brdp={q_brdp}).")


def run_replica_v14(p: Params) -> dict:
    _verifica_params_v14(p)
    T = p.anni
    m = macro(p, sq_ratio_v14=True)
    pil, w, occ, H = m["pil"], m["w"], m["occ"], m["H"]
    t = np.arange(T)

    L = np.array([partecipazione_L(p, i) for i in t])
    M = L * occ
    bif = np.array([quota_bif(p, i) for i in t])          # (tot, brdp)
    q_tot, q_brdp = bif[:, 0], bif[:, 1]
    q_p2 = q_tot - q_brdp

    A, C, E, G_, Hlev, I_, J_ = leve_parametriche_v14(p, T)
    B = q_p2 * M * w * p.irpef_contributi / 1000.0        # Leva B (TEE)
    Q = -q_tot * M * w / 1000.0                           # Leva D (gap)
    Z = q_brdp * M * w / 1000.0                           # BRDP emessi
    R = -Q - Z                                            # flusso ai fondi

    # NDC per coorte (formule AK/AL/AP, identiche per costruzione a v14)
    AL = carriera_ndc_v14(p, p.anno0, lambda cal: p.aliquota)
    AK = np.array([carriera_ndc_v14(p, p.anno(i),
                   lambda cal: p.aliquota - bif_tot_cal_v14(p, cal)) for i in t])
    AP = np.array([montante_capitalizzato_v14(p, p.anno(i),
                   lambda cal: bif_tot_cal_v14(p, cal), p.rf) for i in t])
    AM = AK / AL

    # EMA a 20 anni (durata pensione fissa B172=20 in v14)
    AN = np.empty(T); AN[0] = 1.0
    for i in range(1, T):
        AN[i] = AN[i-1] * (1 - 1/20) + AM[i] * (1/20)

    AO = np.minimum(p.p0_forfait_v14, p.p0_forfait_v14 * (t + 1) / p.p0_phasein_v14)
    U = H * (1 - AN) - AO                                 # Leva F v14

    # BRDP: stock, rimborso bullet a 15 anni, interessi
    AH_ = np.empty(T); AD = np.zeros(T)
    for i in range(T):
        emesso = Z[:i+1].sum()
        rimborsato = Z[:i+1-p.brdp_bullet].sum() if i >= p.brdp_bullet else 0.0
        AH_[i] = emesso - rimborsato
        if i >= p.brdp_bullet:
            AD[i] = -Z[i - p.brdp_bullet]
    AC_ = -(p.r_brdp * AH_)

    AE = Z + A + B + C + Q + E + U + G_ + Hlev + I_ + J_ + AC_ + AD
    AF = np.cumsum(AE)
    J_spesa = H - AE
    K_ratio = J_spesa / pil

    # calibrazione kappa e pensioni individuali (colonne AQ/AR/AS)
    kappa = (H[0] * 1000 / (p.pensionati0 * p.mensilita)) / (AL * p.coeff69 / p.mensilita)
    AR_ = AK * p.coeff69 / p.mensilita * kappa
    AS_ = AP * p.coeff69 / p.mensilita * kappa

    # fondo v14 (Dinamica): SENZA decumulo — riportato solo per confronto
    fondo = np.empty(T); fondo[0] = R[0]
    for i in range(1, T):
        cedole = AH_[i-1] * p.rf
        rimborso_in = -AD[i]  # il fondo riceve il capitale rimborsato
        fondo[i] = fondo[i-1] * (1 + p.rf) + R[i] + cedole + rimborso_in

    return dict(t=t, anno=p.anno0 + t, pil=pil, w=w, occ=occ, H=H,
                ratio=m["ratio"], L=L, M=M, A=A, B=B, C=C, Q=Q, R=R, E=E,
                U=U, G=G_, Hlev=Hlev, I=I_, J=J_, Z=Z, AH=AH_, AC=AC_, AD=AD,
                AE=AE, AF=AF, spesa_rif=J_spesa, ratio_rif=K_ratio,
                AM=AM, AN=AN, AO=AO, kappa=kappa, AR=AR_, AS=AS_,
                fondo_v14=fondo, fondi_cum=np.cumsum(R))


# ------------------------------------------------------------------------------
# 4. MOTORE v15 — CONTABILITA' PER COORTI
# ------------------------------------------------------------------------------

def anno_adesione(p: Params, anno_pens: int) -> int | None:
    """Anno di calendario in cui la coorte (che si pensiona in anno_pens alla sua
    eta' pensionabile) entra nella biforcazione. None = mai.

    REGOLA (semplificata 2026-08): soglia UNICA di eta' nell'anno di avvio.
    Chi ha al piu' part_t0_eta anni entra subito; gli altri restano interamente
    nella ripartizione.

    Il phase-in a due scaglioni che c'era prima era logicamente rovesciato:
    faceva entrare subito i piu' giovani e cinque anni dopo i quarantenni piu'
    anziani, cioe' proprio quelli che avevano MENO tempo per accumulare. Se una
    coorte e' abbastanza giovane da poter costruire un montante, non c'e' alcuna
    ragione per farla aspettare; se non lo e', non deve entrare affatto.
    """
    eta_2027 = eta_pensionabile(p, anno_pens) - (anno_pens - p.anno0)
    return p.anno0 if eta_2027 <= p.part_t0_eta else None


@dataclass
class Coorte:
    anno_pens: int
    N: float                 # mln di persone
    p1_sq: float             # €/mese, pensione controfattuale status quo
    p1_rif: float            # €/mese, P1 residua post-biforcazione
    p2: float                # €/mese, rendita P2+BRDP (da allocazione coerente)
    p0: float                # €/mese, P0 medio (taper su profili) — COSTO fiscale
    montante_agg: float      # €mld, montante aggregato della coorte al pensionamento
    rendita_agg: float       # €mld/anno, rendita aggregata (annua, a piena coorte)
    eleggibile_p0: bool = True   # False: resta nel regime assistenziale attuale
    # --- stato SUPERSTITE (Task 7) ---
    p0_sup: float = 0.0      # €/mese, P0 del superstite (test su reversibilita' + P2)
    cap_trasf: float = 0.0   # €, capitale P2 trasferito al superstite (base imposta)
    tax_succ: float = 0.0    # €, imposta di successione una tantum per evento
    # P0 del lavoratore MEDIO (profilo m=1), usato per l'ADEGUATEZZA. Distinto da
    # p0, che e' la media pesata sui profili e serve al COSTO aggregato del P0.
    # Sommare p0 (media di popolazione) a p1_rif+p2 (che sono grandezze del
    # lavoratore m=1) accrediterebbe al pensionato medio un P0 che in realta'
    # spetta solo ai profili bassi: due popolazioni diverse nella stessa somma.
    p0_m1: float = 0.0


def alloca_montanti(p: Params, R_fondi: np.ndarray, Z: np.ndarray, flussi: dict,
                    versanti: np.ndarray | None = None):
    """Alloca i flussi AGGREGATI verso i fondi alle coorti di pensionamento.

    ECONOMIA: il modello genera i contributi biforcati a livello aggregato
    (R_fondi verso i fondi di mercato, Z verso i BRDP). Qui si distribuisce il
    flusso di ciascun anno fra le coorti che in quell'anno sono ANCORA AL
    LAVORO e hanno gia' aderito alla biforcazione, in proporzione alla loro
    numerosita' (coorti piu' popolose accantonano di piu'). Ogni quota viene poi
    capitalizzata fino all'anno di pensionamento della coorte: al rendimento di
    mercato rf per la parte fondi, alla cedola r_brdp per la parte BRDP.
    Per costruzione somma(montanti) = somma(flussi capitalizzati): niente doppio
    scaling fra livello aggregato e livello di coorte.

    Implementazione vettorizzata: adesioni e pesi precalcolati una sola volta,
    il loop interno sulle coorti diventa un'operazione su array (era O(T*Y) in
    Python puro con anno_adesione ricalcolata a ogni iterazione).
    """
    T = p.anni
    anni_pens = list(range(p.anno0, p.anno0 + T + p.carriera + (p.eta_pens_max - p.eta_pens)))
    ys = np.array(anni_pens)
    flussi_arr = np.array([flussi[y] for y in anni_pens], dtype=float)
    # anno di adesione per coorte (INF = non aderisce mai): calcolato UNA volta
    INF = 10 ** 9
    ade_arr = np.array([(lambda a: INF if a is None else a)(anno_adesione(p, y))
                        for y in anni_pens], dtype=float)
    # Fattori di capitalizzazione: con rf costante equivalgono a (1+rf)^n; con un
    # sentiero stocastico catturano l'ordine dei rendimenti. L'orizzonte va esteso
    # oltre T perche' le ultime coorti si pensionano fino a T+carriera.
    # carriera MASSIMA: con l'eta' pensionabile che sale, le coorti tarde hanno
    # una carriera piu' lunga di p.carriera e devono poter ricevere quote dei
    # flussi anche negli anni piu' remoti (altrimenti il loro montante e'
    # sistematicamente sottostimato).
    carr_max_a = p.carriera + (p.eta_pens_max - p.eta_pens)
    if versanti is None:                       # retrocompatibilita'
        versanti = np.array([sum(flussi.get(p.anno0 + i + 1 + kk, 0.0)
                                 for kk in range(carr_max_a))
                             for i in range(T)], dtype=float)
    n_ext = T + carr_max_a + 1
    C_rf = fattori_capitalizzazione(rf_serie(p, n_ext))
    mont_arr = np.zeros(len(anni_pens))
    for i in range(T):
        cal = p.anno0 + i
        # coorti al lavoro in 'cal' (si pensionano entro una carriera) e gia'
        # aderenti alla biforcazione
        attive = (ys > cal) & (ys <= cal + carr_max_a) & (ade_arr <= cal)
        if not attive.any():
            continue
        # QUOTA PRO CAPITE, non quota della coorte. Il flusso dell'anno e'
        # diviso per il numero di lavoratori che lo hanno effettivamente
        # versato, e ogni coorte ne riceve una quota pari alla propria
        # numerosita'. Normalizzare invece sulla somma delle coorti attive
        # sarebbe corretto solo se quella somma coincidesse con i versanti: non
        # coincide (la carriera teorica di 42 anni e' piu' lunga di quella
        # implicita nei flussi di pensionamento osservati), e il montante
        # individuale ne risulterebbe distorto — gonfiato o compresso a seconda
        # del segno dello scarto. Cosi' invece il montante pro capite e'
        # esattamente quello di chi ha versato quei contributi.
        vers = versanti[i]
        if vers <= 0:
            continue
        anni_cap = ys[attive] - cal              # anni di capitalizzazione residui
        cap_rf = C_rf[i + anni_cap] / C_rf[i]    # fattore fondi (path-dependent)
        quota_pc = flussi_arr[attive]            # teste della coorte
        mont_arr[attive] += ((R_fondi[i] / vers) * quota_pc * cap_rf
                             + (Z[i] / vers) * quota_pc
                             * (1 + p.r_brdp) ** anni_cap)
    return dict(zip(anni_pens, mont_arr))


def costruisci_coorti(p: Params, kappa_c: float,
                      R_fondi: np.ndarray, Z: np.ndarray,
                      flussi: dict,
                      versanti: np.ndarray | None = None) -> list[Coorte]:
    """Per ogni anno di pensionamento 2027-2080 calcola le grandezze della
    coorte. P1 e' ancorata (kappa_c) alla pensione CONTRIBUTIVA media osservata
    (H0 * quota_contrib / pensionati); P2 deriva dall'allocazione coerente dei
    flussi aggregati (nessun doppio scaling).
    Task 4.1: eta' pensionabile, carriera e coefficiente di trasformazione sono
    coorte-specifici (l'eta' sale con la speranza di vita). Applicati
    SIMMETRICAMENTE a status quo (p1_sq) e riforma (p1_rif): il coefficiente si
    cancella su dP1 via kappa, quindi niente doppio conteggio con la Leva I."""
    coorti = []
    montanti = alloca_montanti(p, R_fondi, Z, flussi, versanti)
    # tasso di sconto per l'attualizzazione della rendita reversibile: con un
    # sentiero stocastico si usa il rendimento MEDIO (l'attualizzazione e' una
    # convenzione attuariale, non deve seguire la volatilita' anno per anno)
    rf_medio = float(rf_serie(p, p.anni).mean())
    # quota biforcata per anno di calendario: precalcolata una sola volta
    # (era ricalcolata dentro il SUMPRODUCT di ogni coorte: ~3.4k chiamate/run)
    carr_max = p.carriera + (p.eta_pens_max - p.eta_pens)
    q_tot_cal = {cal: quota_bif(p, cal - p.anno0)[0]
                 for cal in range(p.anno0 - carr_max, p.anno0 + p.anni + 1)}
    for y in range(p.anno0, p.anno0 + p.anni):
        ade = anno_adesione(p, y)
        carr = carriera_pens(p, y)                         # carriera della coorte
        cf = coeff_eta(p, eta_pensionabile(p, y))          # coeff. all'eta' effettiva

        def bif_eff(cal, ade=ade):
            """Quota biforcata effettiva della coorte nell'anno cal: zero finche'
            la coorte non ha aderito (phase-in per eta')."""
            if ade is None or cal < ade:
                return 0.0
            return q_tot_cal.get(cal, 0.0)

        # AL = montante status quo (aliquota piena 33%); AK = montante NDC
        # residuo dopo la biforcazione. La differenza e' cio' che genera dP1.
        pc = bool(p.salario_per_coorte)
        AL = carriera_ndc_v14(p, y, lambda cal: p.aliquota, carriera=carr,
                              per_coorte=pc)
        AK = carriera_ndc_v14(p, y, lambda cal: p.aliquota - bif_eff(cal),
                              carriera=carr, per_coorte=pc)
        N = flussi[y]

        p1_sq = AL * cf / p.mensilita * kappa_c
        p1_rif = AK * cf / p.mensilita * kappa_c

        # La rendita P2 e' convertita a DUE TESTE: prosegue al superstite (il
        # montante e' capitale individuale), quindi il coefficiente scende.
        cf_p2 = coeff_reversibile(p, cf)
        mont_agg = montanti.get(y, 0.0)                    # €mld della coorte
        rendita_agg = mont_agg * cf_p2                     # €mld/anno
        # pro-capite: €mld -> € : *1e9 ; N in mln -> *1e6
        p2 = (mont_agg * 1e3 / N) * cf_p2 / p.mensilita if N > 0 else 0.0

        # canale 4: ricalibrazione esplicita della P1 (v. clawback_modo)
        if p.clawback_modo == "target":
            # la P1 residua e' portata al livello che rende la pensione totale
            # pari a p1_sq + target; mai sotto zero e mai SOPRA la P1 maturata
            # (lo Stato preleva il surplus, non integra il deficit).
            p1_rif = min(p1_rif, max(0.0, p1_sq + p.clawback_target - p2))
        else:
            p1_ndc = p1_rif                       # P1 maturata, prima del prelievo
            p1_rif = max(0.0, p1_ndc - p.clawback_p2 * p2)
            if p.clawback_floor is not None:
                # salvaguardia: si restituisce quanto basta a non scendere sotto
                # lo status quo + floor, senza mai eccedere la P1 maturata.
                p1_rif = min(p1_ndc, max(p1_rif, p1_sq + p.clawback_floor - p2))

        # P0 con taper per profilo di reddito (pensioni scalano ~linearmente).
        # Se p0_solo_biforcati: le coorti mai entrate nella biforcazione restano
        # nel regime assistenziale attuale (P0=0, nessun credito safety net).
        eleggibile = (ade is not None) or (not p.p0_solo_biforcati)
        taper = p.p0_base / p.p0_cap
        p0 = p0_sup = p0_m1 = 0.0
        if eleggibile:
            # P0 del lavoratore MEDIO: stesso taper, profilo m=1. E' la quota di
            # P0 che spetta davvero a chi ha la pensione p1_rif+p2.
            p0_m1 = max(0.0, p.p0_base - taper * (p1_rif + p2))
            for m, peso in zip(p.profili, p.pesi_profili):
                # TITOLARE: il test di reddito guarda P1 residua + rendita P2
                p0 += peso * max(0.0, p.p0_base - taper * m * (p1_rif + p2))
                # SUPERSTITE: il test guarda la reversibilita' PUBBLICA sulla P1
                # PIU' la rendita P2 ereditata. Includere il P2 e' essenziale:
                # altrimenti chi eredita un buon montante prenderebbe ANCHE il P0
                # (doppia copertura). Cosi' il taper lo esclude da solo.
                reddito_sup = m * (p.rev_pub_rif * p1_rif + p.rev_alpha_p2 * p2)
                p0_sup += peso * max(0.0, p.p0_base - taper * reddito_sup)

        # Base dell'imposta di successione una tantum (v. rev_tax_base):
        #  - "montante": l'intero conto individuale accumulato (massa ereditaria)
        #  - "riserva" : valore attuale della sola rendita reversibile
        if not p.rev_attiva:
            cap_trasf = 0.0
        elif p.rev_tax_base == "montante":
            cap_trasf = (mont_agg * 1e3 / N) if N > 0 else 0.0     # € pro-capite
        else:
            cap_trasf = (p.rev_alpha_p2 * p2 * p.mensilita
                         * fattore_rendita(p.rev_durata, rf_medio))
        tax_succ = imposta_successione_p2(p, cap_trasf)

        coorti.append(Coorte(y, N, p1_sq, p1_rif, p2, p0, mont_agg, rendita_agg,
                             eleggibile, p0_sup, cap_trasf, tax_succ, p0_m1))
    return coorti


def run_v15(p: Params) -> dict:
    T = p.anni
    m = macro(p, sq_ratio_v14=False)
    pil, w, occ, H = m["pil"], m["w"], m["occ"], m["H"]
    t = np.arange(T)

    # --- PERIMETRO DELLA RIFORMA (Task 10) --------------------------------------
    # q_spesa/q_teste isolano il SOTTO-SISTEMA su cui la riforma agisce. Con
    # perimetro="dipendenti" la biforcazione riguarda i soli FPLD+CTPS: spesa,
    # pensionati, flussi e safety net vanno ridotti alla loro quota, altrimenti i
    # contributi verrebbero sottratti a una platea e i risparmi misurati su
    # un'altra (il disallineamento che distorceva le versioni precedenti).
    # NB: le leve PARAMETRICHE (A, G, H, I, J, L — perequazione, reversibilita',
    # eta', deroghe, evasione) restano di SISTEMA: sono misure generali che non
    # dipendono dalla creazione del pilastro a capitalizzazione.
    q_spesa, q_teste = p.quote_perimetro
    H_perim = H * q_spesa                  # spesa pensionistica del perimetro
    H_altri = H - H_perim                  # altre gestioni: non toccate dalla riforma

    # split contributivo / assistenziale (C2), sul perimetro riformato
    H_contrib = H_perim * p.quota_contrib
    H_assist = H_perim * (1 - p.quota_contrib)

    # Forza lavoro COPERTA dalla riforma: dipendenti (+ autonomi se universale).
    # E' la popolazione su cui si calcolano SIA i contributi biforcati SIA le
    # coorti di nuovi pensionati: le due cose devono stare sulla stessa base,
    # altrimenti il montante aggregato si spalma su un numero di teste diverso
    # da quello che l'ha versato e la rendita P2 pro-capite risulta distorta.
    occ_aut = m["occ_aut"]
    occ_coperti = occ + occ_aut

    # flusso di nuovi pensionati per anno (Task 4.2: dinamico, legato agli
    # occupati COPERTI), ridotto alla quota di teste del perimetro: nel
    # sotto-sistema dipendenti solo il 68,5% dei nuovi pensionati proviene da
    # FPLD/CTPS, ed e' su quelle coorti che agisce la riforma.
    # I gradini INPS su cui costruisci_flussi e' calibrata (0.55/0.60/0.48 mln
    # l'anno) sono GIA' i flussi di pensionamento dei dipendenti: 542.853 nuove
    # pensioni dei fondi dei dipendenti nel 2025, non le 792.998 di sistema.
    # Applicarvi di nuovo la quota di teste sarebbe una SECONDA riduzione al
    # perimetro. Il controllo di coerenza: lo stock di pensionati del perimetro
    # (~11,2 mln) diviso la permanenza media (19,5 anni) da' ~0,58 mln l'anno,
    # vicino ai gradini grezzi e lontano dal loro 68,5%.
    flussi = costruisci_flussi(p, occ_coperti)

    # partecipazione alla biforcazione: derivata dalle COORTI (una sola fonte di
    # verita', Task 7) oppure curva aggregata di v14 se L_da_coorti=0
    L = (partecipazione_da_coorti(p, flussi) if p.L_da_coorti
         else np.array([partecipazione_L(p, i) for i in t]))
    M = L * occ                                # lavoratori dipendenti partecipanti
    M_aut = L * occ_aut                        # autonomi partecipanti (0 se L1)
    bif = np.array([quota_bif(p, i) for i in t])
    q_tot, q_brdp = bif[:, 0], bif[:, 1]
    q_p2 = q_tot - q_brdp

    # MONTE SALARI partecipante (Task 8): dipendenti (w) + autonomi (w_autonomi).
    # I punti biforcati (q_tot) sono punti di imponibile, uguali per entrambi;
    # cosi' il costo di transizione (Leva D) e' sull'intero monte salari coperto,
    # coerente con H che copre l'intero sistema. In perimetro L1, M_aut = 0.
    monte_part = (M * w + M_aut * p.w_autonomi) / 1000.0     # €mld
    base_contrib = (occ * w + occ_aut * p.w_autonomi) / 1000.0  # monte salari TOTALE (per Leva K)

    levaD = -q_tot * monte_part
    Z = q_brdp * monte_part
    R_fondi = -levaD - Z                       # afflusso al fondo (netto BRDP)
    levaB = (q_p2 * monte_part * p.irpef_contributi
             if p.regime_p2 == "TEE" else np.zeros(T))

    A, G_, Hlev, I_, J_ = leve_parametriche_v15(p, T)

    # ratio pensionati VIVA (Task 4.4-B): stock di pensioni vigenti / cumulo =
    # numero di pensionati fisici, evoluzione demografica reale. Sostituisce il
    # denominatore statico pensionati0 nei conteggi P0/P2 (share dei pensionati
    # nel nuovo regime). H(t) resta ancorato alla RGS: la ratio governa i NUMERI,
    # non la spesa pubblica aggregata (nessuna divergenza dal sentiero RGS).
    cumulo = p.pensioni_vigenti0 / p.pensionati0
    pensionati_sistema = m["pens_vig"] / cumulo        # tutti i pensionati
    pensionati = pensionati_sistema * q_teste          # quelli del perimetro riformato

    # --- coorti (C3) ------------------------------------------------------------
    # kappa ancorata alla coorte BASE 2027: calibra la pensione contributiva media
    # osservata. Task 7: il denominatore conta SOLO i pensionati che percepiscono
    # una prestazione contributiva — prima si divideva per TUTTI i pensionati,
    # inclusi i soli-assistenziali, sottostimando la media di ~20%.
    # Nel sotto-sistema dipendenti numeratore e denominatore sono entrambi ridotti
    # al perimetro: la pensione media che ne esce e' quella dei dipendenti (piu'
    # alta della media di sistema, perche' i pubblici alzano la media).
    AL0 = carriera_ndc_v14(p, p.anno0, lambda cal: p.aliquota, carriera=p.carriera)
    pensionati_prev0 = p.pensionati0 * q_teste * p.quota_pensionati_previdenziali
    pens_contrib_media = H_perim[0] * p.quota_contrib * 1000 / (pensionati_prev0 * p.mensilita)
    kappa_c = pens_contrib_media / (AL0 * coeff_eta(p, p.eta_pens) / p.mensilita)
    coorti = costruisci_coorti(p, kappa_c, R_fondi, Z, flussi, M)

    # --- MATRICE DI SOPRAVVIVENZA (aggregazione coorti -> anni) ------------------
    # W[i,j] = numero di pensionati della coorte j ancora vivi nell'anno i
    #        = N_j * sopravvivenza(anni trascorsi dal pensionamento)
    # La mortalita' e' lineare: la coorte si estingue in 2*durata_media_pensione
    # anni (durata media 19.5 -> orizzonte 39), cosi' la permanenza MEDIA in
    # pensione e' esattamente la speranza di vita al pensionamento.
    # Precalcolata una volta: ogni aggregato annuo diventa un prodotto
    # matrice-vettore W @ x invece di un doppio loop Python (era il ~50% del
    # tempo di calcolo, con la sopravvivenza ricalcolata 22k volte per run).
    anni_pens_arr = np.array([c.anno_pens for c in coorti], dtype=float)
    N_arr = np.array([c.N for c in coorti], dtype=float)
    eta_da_pens = (p.anno0 + t)[:, None] - anni_pens_arr[None, :]   # (T, C)
    Hor = 2 * p.durata_media_pensione          # orizzonte di estinzione (39 anni)

    def S_lin(x):
        """Sopravvivenza lineare, con 1 prima del pensionamento e 0 dopo Hor."""
        return np.clip(1.0 - x / Hor, 0.0, 1.0)

    S = np.where(eta_da_pens >= 0, S_lin(eta_da_pens), 0.0)
    W = S * N_arr[None, :]                     # (T, C) TITOLARI vivi per coorte

    # --- STATO SUPERSTITE (Task 7) ----------------------------------------------
    # R(a) = S(a - a_rev) - S(a) = quota di coorte il cui titolare e' morto negli
    # ULTIMI rev_durata anni, cioe' i superstiti che stanno ancora incassando.
    # Solo la quota rev_quota lascia un coniuge avente diritto.
    if p.rev_attiva:
        R_sup = np.where(eta_da_pens >= 0,
                         S_lin(eta_da_pens - p.rev_durata) - S_lin(eta_da_pens), 0.0)
        W_sup = R_sup * p.rev_quota * N_arr[None, :]     # (T, C) superstiti vivi
        # NUOVI eventi di decesso nell'anno (per l'imposta di successione):
        # con mortalita' lineare la densita' di decessi e' 1/Hor per anno.
        D_new = np.where((eta_da_pens >= 0) & (eta_da_pens <= Hor),
                         p.rev_quota * N_arr[None, :] / Hor, 0.0)
    else:
        W_sup = np.zeros_like(W)
        D_new = np.zeros_like(W)

    def agg(vals: np.ndarray) -> np.ndarray:
        """Aggrega una grandezza PRO-CAPITE del TITOLARE in serie annua."""
        return W @ np.asarray(vals, dtype=float)

    def agg_sup(vals: np.ndarray) -> np.ndarray:
        """Aggrega una grandezza PRO-CAPITE del SUPERSTITE in serie annua."""
        return W_sup @ np.asarray(vals, dtype=float)

    fattore = p.mensilita / 1000.0             # €/mese * mln persone -> €mld/anno
    v_p1_sq = np.array([c.p1_sq for c in coorti])
    v_p1_rif = np.array([c.p1_rif for c in coorti])
    v_p0 = np.array([c.p0 for c in coorti])
    v_p0_sup = np.array([c.p0_sup for c in coorti])
    v_p2 = np.array([c.p2 for c in coorti])
    v_uno = np.ones(len(coorti))
    v_elig = np.array([1.0 if c.eleggibile_p0 else 0.0 for c in coorti])
    v_tax = np.array([c.tax_succ for c in coorti])

    # pensionati "nuovi" vivi = titolari + superstiti che incassano
    alive_new = agg(v_uno) + agg_sup(v_uno)

    # dP1: risparmio pubblico su P1. DUE componenti:
    #  - titolare: P1 status quo - P1 riformata (biforcazione + clawback);
    #  - superstite: reversibilita' pubblica SQ (60% di P1 piena) meno quella
    #    riformata (45% di P1 ridotta). E' il canale ex-Leva H, ora ENDOGENO.
    # Le due componenti sono tenute SEPARATE e riportate distintamente
    # nell'output: la Leva F aggrega molti canali e senza questo dettaglio non si
    # riesce a leggere quanto pesa la revisione della reversibilita' (ex Leva H).
    dP1_titolari = agg(v_p1_sq - v_p1_rif) * fattore
    dP1_superstiti = agg_sup(p.rev_pub_sq * v_p1_sq
                             - p.rev_pub_rif * v_p1_rif) * fattore
    dP1_lordo = dP1_titolari + dP1_superstiti
    irpef_persa = dP1_lordo * p.irpef_pensioni                    # C5
    p0_lordo = (agg(v_p0) + agg_sup(v_p0_sup)) * fattore          # C6
    # rendite P2 pagate dal fondo: al titolare e, dopo il decesso, al superstite
    rendite_p2 = (agg(v_p2) + agg_sup(p.rev_alpha_p2 * v_p2)) * fattore
    tassa_rendite = (rendite_p2 * p.irpef_pensioni
                     if p.regime_p2 == "EET" else np.zeros(T))
    # imposta UNA TANTUM di successione sui montanti P2 trasferiti nell'anno
    # (€ per evento x mln di eventi -> €mld)
    imposta_succ = (D_new @ v_tax) / 1000.0
    alive_p0 = agg(v_elig) + agg_sup(v_elig)
    # quota dei pensionati (VIVI, ratio reale) gia' passati al nuovo regime P0/P2
    quota_nuovi = np.minimum(1.0, alive_p0 / pensionati)   # Task 4.4-B: denom. vivo
    # SAFETY NET ASSORBITO DAL P0 (C6). Il pilastro zero non e' una prestazione
    # aggiuntiva che si somma all'assistenza esistente: la sostituisce per la
    # platea che copre. Va quindi contabilizzato al netto di cio' che lo Stato
    # gia' spende oggi per gli stessi pensionati (integrazione al minimo,
    # maggiorazioni sociali).
    # VINCOLO STRUTTURALE: l'assorbimento non puo' MAI eccedere il costo del P0.
    # Senza questo tetto il modello registrerebbe un risparmio netto — al 2080
    # il P0 costava 7,8 mld e ne "assorbiva" 13,2 — cioe' accrediterebbe alla
    # riforma un risparmio assistenziale non giustificato: il P0 (650 €/mese,
    # azzeramento a 1.200) e' piu' generoso dell'assegno sociale e
    # dell'integrazione al minimo, quindi non puo' costare meno di cio' che
    # sostituisce. Con il tetto il pilastro zero e' al piu' NEUTRO sul bilancio,
    # mai una fonte di risparmio. NB: il grosso della spesa assistenziale
    # (assegni sociali e invalidita' civili a chi non ha carriera contributiva)
    # riguarda una platea che questo modello non rappresenta e che il P0 non
    # sostituisce: resta invariata dentro H.
    assist_assorbita = np.minimum((p.safety_net * q_teste) * quota_nuovi,
                                  p0_lordo)

    # Leva F v15 = risparmio P1 netto IRPEF - P0 netto safety net + tasse rendite
    levaF = (dP1_lordo - irpef_persa) - (p0_lordo - assist_assorbita) + tassa_rendite

    # --- BRDP e interessi ---------------------------------------------------------
    AH_ = np.empty(T); AD = np.zeros(T)
    for i in range(T):
        emesso = Z[:i+1].sum()
        rimborsato = Z[:i+1-p.brdp_bullet].sum() if i >= p.brdp_bullet else 0.0
        AH_[i] = emesso - rimborsato
        if i >= p.brdp_bullet:
            AD[i] = -Z[i - p.brdp_bullet]

    # --- Leva L: recupero tax gap contributivo (rampa lineare) ---------------------
    levaL = (p.levaL_gap * np.minimum(1.0, (t + 1) / p.levaL_rampa)
             if p.levaL_attiva else np.zeros(T))

    # --- Saldo primario (C7: NIENTE emissioni/rimborsi dentro il saldo) ------------
    # (post R2: niente Leva C ne' Leva E. Task 7: + imposta di successione sui
    #  montanti P2 trasferiti ai superstiti; il risparmio da revisione della
    #  reversibilita' e' dentro levaF, non piu' nel forfait Leva H)
    saldo_prim = A + levaB + levaD + levaF + G_ + Hlev + I_ + J_ + levaL + imposta_succ

    # --- Spesa pubblica riformata (per confronto con RGS) --------------------------
    # Il perimetro riformato risparmia dP1 e paga il P0 netto; le leve
    # parametriche (G, H, I, J) sono misure di sistema e agiscono sul totale.
    # H_altri (le gestioni non riformate) rientra invariato, cosi' il ratio resta
    # confrontabile con il sentiero RGS di spesa pensionistica complessiva.
    spesa_perim_rif = H_perim - dP1_lordo + (p0_lordo - assist_assorbita)
    spesa_rif = spesa_perim_rif + H_altri - (G_ + Hlev + I_ + J_)
    ratio_rif = spesa_rif / pil
    ratio_perim = spesa_perim_rif / pil        # solo il sotto-sistema riformato

    # quota di pensioni pagata dai fondi (indicatore, riportato nella Dinamica)
    quota_fondi = rendite_p2 / (spesa_rif + rendite_p2)

    # DRIVER della LEVA K ridisegnata (Task 3): frazione del monte P1 pubblico
    # gia' sostituita dalle coorti biforcate = dP1_lordo / H_contrib. E' 0 finche'
    # le coorti biforcate non si pensionano (~2051), poi cresce liscia e monotona.
    riduzione_p1 = np.where(H_contrib > 0, dP1_lordo / H_contrib, 0.0)

    # debito di transizione: BRDP + finanziamento dei deficit primari.
    # Identita': Delta(debito totale) = -saldo_tot. Le emissioni BRDP (Z) danno
    # cassa allo Stato (riducono l'altro debito), i rimborsi (-AD) lo aumentano.
    # LEVA K dentro il loop: il taglio del cuneo e' ancorato al calo del monte P1
    # (riduzione_p1) e finanziato al piu' dal dividendo netto di P1 dell'anno.
    debito = np.empty(T); interessi = np.empty(T); saldo_tot = np.empty(T)
    levaK_costo = np.zeros(T); levaK_sconto = np.zeros(T)
    altro_debito_prev = 0.0
    for i in range(T):
        brdp_prev = AH_[i-1] if i > 0 else 0.0
        interessi[i] = p.r_brdp * brdp_prev + p.r_debito * altro_debito_prev
        margine = saldo_prim[i] - interessi[i]
        if p.levaK_attiva:
            base_payroll = base_contrib[i]                 # €mld per punto di cuneo (monte salari coperto)
            sconto_pp = min(p.levaK_max, p.levaK_coeff * riduzione_p1[i])
            costo_pot = sconto_pp * base_payroll
            # finanziato al piu' dalla quota del dividendo netto di P1 dell'anno
            dividendo_p1 = max(0.0, dP1_lordo[i] - irpef_persa[i])
            levaK_costo[i] = min(costo_pot, p.levaK_quota_dividendo * dividendo_p1)
            levaK_sconto[i] = levaK_costo[i] / base_payroll if base_payroll > 0 else 0.0
        saldo_tot[i] = margine - levaK_costo[i]
        altro_debito = altro_debito_prev - saldo_tot[i] - Z[i] + (-AD[i])
        debito[i] = AH_[i] + altro_debito
        altro_debito_prev = altro_debito
    saldo_cum = np.cumsum(saldo_tot)

    # --- SALDO DI CASSA (richiesta utente 2026-08) ---------------------------------
    # Terza lettura del saldo, accanto a primario e totale: il debito BRDP entra
    # in CASSA quando viene emesso (Z = entrata in t) ed esce quando viene
    # rimborsato (AD = uscita bullet in t+15, gia' negativa); gli interessi sono
    # gia' dentro saldo_tot. E' il profilo di cassa effettivo dello Stato.
    # Identita' (verificata): saldo_cassa_cum = saldo_cum + stock BRDP vivo (AH);
    # quando tutti i titoli sono rimborsati le due cumulate coincidono — il
    # debito non crea ne' distrugge risorse, le sposta nel tempo al costo degli
    # interessi.
    saldo_cassa = saldo_tot + Z + AD
    saldo_cassa_cum = np.cumsum(saldo_cassa)

    # ALIQUOTA CONTRIBUTIVA IVS effettiva dopo la Leva K (33% - sconto). NB: e'
    # l'aliquota PENSIONISTICA (9.19 lav.+23.81 dat.), NON il cuneo fiscale totale.
    # Il cuneo OCSE (~45.8% nel 2025, Taxing Wages) somma anche IRPEF e altri
    # contributi; il taglio dell'IVS vi si trasmette ~1:1 (voce additiva).
    aliquota_ivs = p.aliquota - levaK_sconto              # aliquota IVS effettiva
    cuneo_ocse_impl = p.cuneo_ocse_base - levaK_sconto    # cuneo totale implicito
    cuneo_totale = aliquota_ivs   # alias retrocompatibile (era mal chiamato "cuneo")

    # --- DEBITO PENSIONISTICO CUMULATO dal 2027: SQ vs riforma ---------------------
    # Lo status quo NON e' gratis: la spesa supera i contributi (quota_contrib=72.8%
    # nel 2027) e il resto e' coperto da fiscalita'/debito. Qui cumuliamo dal 2027
    # il deficit del sistema (spesa - gettito contributivo, gettito che evolve con
    # occupati x salari) con interessi r_debito. Riforma = SQ + debito transizione.
    gettito_contrib = (p.quota_contrib * H[0]) * (base_contrib / base_contrib[0])
    deficit_sq = H - gettito_contrib
    debito_pens_sq = np.empty(T)
    st = 0.0
    for i in range(T):
        st = st * (1 + p.r_debito) + deficit_sq[i]
        debito_pens_sq[i] = st
    debito_pens_rif = debito_pens_sq + debito

    # --- Fondo con DECUMULO (C4) ----------------------------------------------------
    # Con rf_path lo stock si rivaluta al rendimento DELL'ANNO: e' qui che si
    # manifesta il sequence-of-returns risk (un crollo quando lo stock e' grande
    # pesa molto piu' di uno quando e' piccolo).
    fondo = np.empty(T)
    fondo_in = R_fondi.copy()
    fondo_out = rendite_p2.copy()
    rf_ann = rf_serie(p, T)
    stato = 0.0
    for i in range(T):
        cedole = AH_[i-1] * p.r_brdp if i > 0 else 0.0
        rimborso_in = -AD[i]
        stato = stato * (1 + rf_ann[i]) + fondo_in[i] + cedole + rimborso_in - fondo_out[i]
        fondo[i] = stato

    # pensione del nuovo pensionato dell'anno (adeguatezza)
    p1_new = np.array([c.p1_rif for c in coorti])
    p2_new = np.array([c.p2 for c in coorti])
    p0_new = np.array([c.p0 for c in coorti])          # media pesata: COSTO del P0
    p0_m1_new = np.array([c.p0_m1 for c in coorti])    # profilo m=1: ADEGUATEZZA
    p1_sq_new = np.array([c.p1_sq for c in coorti])
    # PENSIONE DEL NUOVO PENSIONATO (indicatore di adeguatezza): tutte le
    # componenti si riferiscono allo STESSO lavoratore, quello di profilo medio.
    pensione_new = p1_new + p2_new + p0_m1_new

    # ADEGUATEZZA PER PROFILO DI REDDITO: P1, P2 e la pensione di status quo
    # scalano col salario (moltiplicatore m); il P0 no, perche' e' means-tested
    # ed e' proprio la sua non-proporzionalita' a redistribuire. Serve a leggere
    # il risultato distributivo che l'indicatore medio, da solo, nasconde.
    taper_p0 = p.p0_base / p.p0_cap
    pens_profilo, pens_sq_profilo = {}, {}
    for m_, peso_ in zip(p.profili, p.pesi_profili):
        p0_prof = np.where(p0_m1_new + p1_new + p2_new > 0,
                           np.maximum(0.0, p.p0_base - taper_p0 * m_ * (p1_new + p2_new)),
                           0.0)
        p0_prof = np.where(np.array([c.eleggibile_p0 for c in coorti]), p0_prof, 0.0)
        pens_profilo[m_] = m_ * (p1_new + p2_new) + p0_prof
        pens_sq_profilo[m_] = m_ * p1_sq_new
    # media di popolazione (pesi dei profili): e' la pensione MEDIA erogata,
    # diversa dalla pensione del lavoratore medio perche' i profili non sono
    # simmetrici attorno a m=1 (peso 50% sul profilo basso).
    pensione_media_pop = sum(peso_ * pens_profilo[m_]
                             for m_, peso_ in zip(p.profili, p.pesi_profili))
    pensione_sq_media_pop = sum(peso_ * pens_sq_profilo[m_]
                                for m_, peso_ in zip(p.profili, p.pesi_profili))

    return dict(p=p, t=t, anno=p.anno0 + t, pil=pil, w=w, occ=occ,
                pens_vig=m["pens_vig"], ratio=m["ratio"], H=H,
                H_contrib=H_contrib, H_assist=H_assist, L=L, M=M,
                levaA=A, levaB=levaB, levaD=levaD,
                levaF=levaF, levaG=G_, levaH=Hlev, levaI=I_, levaJ=J_,
                Z=Z, AH=AH_, AD=AD, R_fondi=R_fondi,
                dP1_lordo=dP1_lordo, dP1_titolari=dP1_titolari,
                dP1_superstiti=dP1_superstiti, irpef_persa=irpef_persa,
                p0_lordo=p0_lordo, assist_assorbita=assist_assorbita,
                tassa_rendite=tassa_rendite, rendite_p2=rendite_p2,
                imposta_succ=imposta_succ,
                alive_new=alive_new, quota_nuovi=quota_nuovi,
                saldo_prim=saldo_prim, interessi=interessi,
                saldo_tot=saldo_tot, saldo_cum=saldo_cum,
                saldo_cassa=saldo_cassa, saldo_cassa_cum=saldo_cassa_cum,
                debito=debito,
                levaK_costo=levaK_costo, levaK_sconto=levaK_sconto,
                cuneo_totale=cuneo_totale, aliquota_ivs=aliquota_ivs,
                cuneo_ocse_impl=cuneo_ocse_impl, riduzione_p1=riduzione_p1,
                quota_fondi=quota_fondi, pensionati=pensionati,
                pensionati_sistema=pensionati_sistema, H_perim=H_perim,
                H_altri=H_altri, spesa_perim_rif=spesa_perim_rif,
                ratio_perim=ratio_perim,
                levaL=levaL, deficit_sq=deficit_sq,
                debito_pens_sq=debito_pens_sq, debito_pens_rif=debito_pens_rif,
                spesa_rif=spesa_rif, ratio_rif=ratio_rif,
                fondo=fondo, fondo_in=fondo_in, fondo_out=fondo_out,
                coorti=coorti, kappa_c=kappa_c,
                p1_new=p1_new, p2_new=p2_new, p0_new=p0_new,
                p0_m1_new=p0_m1_new, pensione_new=pensione_new,
                pens_profilo=pens_profilo, pens_sq_profilo=pens_sq_profilo,
                pensione_media_pop=pensione_media_pop,
                pensione_sq_media_pop=pensione_sq_media_pop,
                p1_sq_new=p1_sq_new)


# ------------------------------------------------------------------------------
# 5. VALIDAZIONE CONTRO v14
# ------------------------------------------------------------------------------

def valida_replica(rep: dict, file_v14: str | None):
    """Confronta la replica (con parametri v14 CONGELATI) con i valori salvati
    nel file v14. Ritorna (righe_diff, serie_v14) dove serie_v14 contiene le
    colonne lette dal file per il confronto side-by-side con v15."""
    if not file_v14:
        return None, None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_v14, data_only=True)
        ws = wb["Modello"]
    except Exception as e:
        print(f"[validazione] impossibile leggere {file_v14}: {e}")
        print("[validazione] AVVERTIMENTO: il file v14 serve SOLO come check della "
              "replica — tutti i parametri sono nel codice (params_v14). "
              "Validazione saltata, l'esecuzione continua normalmente.")
        return None, None
    colmap = {"H": "H", "U": "U", "AE": "AE", "AF": "AF", "AN": "AN", "AS": "AS"}
    righe = []
    serie_v14 = {}
    for col_xl, chiave in colmap.items():
        serie = rep[chiave]
        vals = []
        diffs = []
        for i in range(rep["t"].size):
            v14 = ws[f"{col_xl}{5+i}"].value
            vals.append(v14 if isinstance(v14, (int, float)) else np.nan)
            if isinstance(v14, (int, float)):
                diffs.append(abs(serie[i] - v14))
        serie_v14[col_xl] = np.array(vals, dtype=float)
        if diffs:
            righe.append((col_xl, max(diffs), float(np.mean(diffs))))
    return righe, serie_v14


# ------------------------------------------------------------------------------
# 5bis. SIMULAZIONE MONTE CARLO SUL RENDIMENTO DEI FONDI (Task 9)
# ------------------------------------------------------------------------------

# Serie di cui si riportano i percentili nei fan chart. Chiave = nome nel dict
# di run_v15; valore = (etichetta, se rapportare al PIL).
MC_SERIE = {
    "ratio_rif":  ("Spesa pubblica (%PIL)", False),
    "fondo":      ("Fondo P2 (%PIL)", True),
    "rendite_p2": ("Rendite P2 pagate (%PIL)", True),
    "saldo_cum":  ("Posizione netta vs SQ (%PIL)", True),
    "saldo_tot":  ("Saldo annuo (€mld)", False),
    "pensione_tot": ("Pensione nuovo pensionato (€/mese)", False),
}
MC_PERCENTILI = (5, 25, 50, 75, 95)


def simula_montecarlo(p: Params, n_sim: int | None = None,
                      seed: int | None = None, verbose: bool = True) -> dict:
    """Monte Carlo sul rendimento reale dei fondi (rf) — fan chart.

    IPOTESI (scelta dichiarata): rendimenti annui i.i.d. ~ Normale(rf, rf_sigma),
    con rf = media ARITMETICA (2.5% base) e rf_sigma = 9% (volatilita' reale di
    un portafoglio 60/40). Ogni simulazione e' un RUN COMPLETO del modello con
    quel sentiero: montanti capitalizzati anno per anno, fondo che si rivaluta al
    rendimento dell'anno, rendite e clawback che ne discendono.

    DUE EFFETTI ECONOMICI che il run deterministico NON puo' mostrare:
    1) VOLATILITY DRAG: con rendimenti i.i.d. di media aritmetica mu e varianza
       s^2, la crescita composta attesa e' ~ mu - s^2/2 (qui 2.5% - 0.4% ~ 2.1%).
       La MEDIANA delle simulazioni e' percio' sotto il run deterministico a
       rf=2.5%: non e' un bug, e' il costo della volatilita'.
    2) SEQUENCE-OF-RETURNS RISK: conta l'ORDINE dei rendimenti, perche' un
       crollo quando lo stock e' grande (vicino al pensionamento) fa molto piu'
       danno di uno quando lo stock e' piccolo.

    Ritorna: percentili per anno di ogni serie in MC_SERIE, il run deterministico
    di riferimento e le statistiche di sintesi.
    """
    n_sim = int(n_sim if n_sim is not None else p.mc_n_sim)
    rng = np.random.default_rng(p.mc_seed if seed is None else seed)
    T = p.anni
    # orizzonte esteso: le coorti oltre l'orizzonte capitalizzano fino a T+carriera
    n_ext = T + p.carriera + 1

    det = run_v15(p)                       # riferimento deterministico
    campioni = {k: np.empty((n_sim, T)) for k in MC_SERIE}
    for s in range(n_sim):
        path = rng.normal(p.rf, p.rf_sigma, size=n_ext)
        p_s = dataclasses.replace(p, rf_path=path)
        d = run_v15(p_s)
        for k in MC_SERIE:
            if k == "pensione_tot":
                serie = d["pensione_new"]
            else:
                serie = d[k]
            _, su_pil = MC_SERIE[k]
            campioni[k][s] = serie / d["pil"] if su_pil else serie
        if verbose and n_sim >= 200 and (s + 1) % max(1, n_sim // 4) == 0:
            print(f"    Monte Carlo: {s + 1}/{n_sim} simulazioni")

    perc = {k: {q: np.percentile(v, q, axis=0) for q in MC_PERCENTILI}
            for k, v in campioni.items()}
    # riferimento deterministico sulle stesse serie
    det_serie = {}
    for k, (_, su_pil) in MC_SERIE.items():
        serie = det["pensione_new"] if k == "pensione_tot" else det[k]
        det_serie[k] = serie / det["pil"] if su_pil else serie
    return dict(perc=perc, campioni=campioni, det=det, det_serie=det_serie,
                n_sim=n_sim, anno=det["anno"], p=p)


def audit_anno_base(p: Params, d: dict) -> list[tuple]:
    """Confronto ANNO BASE fra modello e dati osservati (INPS / Itinerari / RGS).

    Serve a rendere VISIBILI gli scostamenti di calibrazione invece di lasciarli
    impliciti: un modello che non riproduce il 2027 osservato non e' affidabile
    sul 2080. Ritorna righe (voce, modello, osservato, fonte, giudizio).
    """
    q_spesa, q_teste = p.quote_perimetro
    monte = p.occupati0 * p.w0 / 1000.0
    if p.perimetro == "universale":     # il perimetro copre anche gli autonomi
        monte += p.occ_autonomi0 * p.w_autonomi / 1000.0
    gettito = p.aliquota * monte
    # riferimenti osservati (2024-2026), scalati al perimetro quando serve
    oss_spesa_tot = 353.0        # INPS Osservatorio 1.1.2026 (importo complessivo)
    oss_spesa_dip = 244.5        # Itinerari XIII Rapporto 2024: privati 150,9 + pubblici 93,6
    oss_contrib_dip = 209.7      # Itinerari XIII: contributi privati 162,5 + pubblici 47,3
    # Monte imponibile osservato bottom-up (INPS Osservatori 2024): privati
    # 17,7 mln x 24.486 € = 433,4 + pubblici 3,74 mln x 35.350 € = 132,1.
    oss_monte_dip = 565.5
    oss_monte_uni = 565.5 + 168.0    # + autonomi (6,0 mln x ~28.000 €)
    perim_dip = p.perimetro != "universale"
    righe = [
        ("Spesa pensionistica totale (€mld)", d["H"][0], oss_spesa_tot,
         "INPS Osservatorio 1.1.2026", "ok"),
        ("Spesa del perimetro riformato (€mld)", d["H_perim"][0],
         oss_spesa_dip if perim_dip else oss_spesa_tot,
         "Itinerari XIII Rapporto (dati 2024)", "ok"),
        ("Pensionati totali (mln)", d["pensionati_sistema"][0], 16.3,
         "INPS / Itinerari", "ok"),
        ("Pensionati del perimetro (mln)", d["pensionati"][0],
         16.3 * q_teste, "quota flussi INPS 2025", "ok"),
        ("Pensione contributiva media (€/mese)", d["p1_sq_new"][0],
         1555 if perim_dip else 1441, "derivato spesa/pensionati", "ok"),
        ("Occupati del perimetro (mln)", p.occupati0, 17.5, "INPS FPLD+CTPS", "ok"),
        ("Monte salari imponibile (€mld)", monte,
         oss_monte_dip if perim_dip else oss_monte_uni,
         "INPS Osservatori 2024 (priv. 433,4 + pubbl. 132,1)", "ok"),
        ("Gettito contributivo IVS (€mld)", gettito,
         oss_contrib_dip if perim_dip else 260.6,
         "Itinerari XIII (2024); residuo = contribuzione figurativa GIAS", "SCOSTAMENTO"),
        ("Copertura contributiva della spesa", gettito / d["H_perim"][0],
         oss_contrib_dip / oss_spesa_dip if perim_dip else 0.788,
         "contributi / prestazioni", "SCOSTAMENTO"),
        ("Spesa/PIL status quo", d["ratio"][0], 0.154, "DFP 2026 / RGS", "ok"),
    ]
    return righe


def sintesi_montecarlo(mc: dict, anno: int = 2080) -> dict:
    """Statistiche di rischio all'anno indicato: mediana, intervallo 5-95,
    scarto della mediana dal deterministico (volatility drag) e probabilita' di
    esiti sfavorevoli (posizione netta negativa, pensione sotto lo status quo)."""
    p = mc["p"]
    i = anno - p.anno0
    out = {}
    for k, (lbl, _) in MC_SERIE.items():
        c = mc["campioni"][k][:, i]
        out[k] = dict(label=lbl, mediana=float(np.median(c)),
                      p5=float(np.percentile(c, 5)), p95=float(np.percentile(c, 95)),
                      det=float(mc["det_serie"][k][i]),
                      drag=float(np.median(c) - mc["det_serie"][k][i]))
    out["prob_pos_netta_negativa"] = float((mc["campioni"]["saldo_cum"][:, i] < 0).mean())
    sq = mc["det"]["p1_sq_new"][i]
    out["prob_pensione_sotto_sq"] = float((mc["campioni"]["pensione_tot"][:, i] < sq).mean())
    out["anno"] = anno
    return out


# ------------------------------------------------------------------------------
# 6. FOGLIO SOSTITUZIONE (micro, tassi di sostituzione)
# ------------------------------------------------------------------------------

def tabella_sostituzione(p: Params, d: dict, anno_regime: int | None = None):
    """Tassi di sostituzione lordi per profilo di reddito, coorte a regime.

    DERIVATA DAL MOTORE, non ricalcolata per una via parallela: prende la coorte
    dal run passato in `d`, cosi' incorpora automaticamente tutto cio' che il
    motore applica (salario per coorte, allocazione del montante aggregato,
    aliquota di ricalibrazione E clausola di salvaguardia). Una tabella
    calcolata a parte divergerebbe dal resto del documento — ed e' esattamente
    cio' che accadeva prima: con rendimenti bassi indicava una perdita del 13%
    per il profilo medio mentre il motore, grazie alla salvaguardia, lo teneva
    al livello dello status quo.

    Il salario finale della coorte che si pensiona nell'anno c coincide con il
    salario medio dell'anno c: la carriera parte da w0 rivalutato all'anno di
    inizio e cresce di g_w per K-1 anni, cioe' w0*(1+g_w)^(c-anno0).

    Ritorna righe (profilo, regime, etichetta_rendimento, salario_finale,
    pensione_mensile_su_12, tasso_di_sostituzione).
    """
    if anno_regime is None:
        anno_regime = p.anno0 + 42          # coorte con carriera interamente riformata
    i = anno_regime - p.anno0
    eta_r = eta_pensionabile(p, anno_regime)
    w_fin = p.w0 * (1 + p.g_w) ** (anno_regime - p.anno0)
    p1, p2 = d["p1_new"][i], d["p2_new"][i]
    p1_sq = d["p1_sq_new"][i]
    taper = p.p0_base / p.p0_cap
    righe = []
    for m, nome in zip(p.profili, ("Basso", "Medio", "Alto")):
        sal = m * w_fin
        # status quo: nessun pilastro zero, la pensione scala col salario
        pens_sq = m * p1_sq
        righe.append((f"{nome} ({nf_it(m)}x)", "Status quo", "-", sal,
                      pens_sq * p.mensilita / 12,
                      pens_sq * p.mensilita / sal))
        p0 = max(0.0, p.p0_base - taper * m * (p1 + p2))
        tot = m * (p1 + p2) + p0
        righe.append((f"{nome} ({nf_it(m)}x)", f"Riforma ({eta_r} anni)",
                      f"rf={p.rf:.1%}", sal,
                      tot * p.mensilita / 12,
                      tot * p.mensilita / sal))
    return righe


def nf_it(x: float) -> str:
    """0.5 -> '0,5' (usato nelle etichette delle tabelle)."""
    return f"{x:.1f}".replace(".", ",")


# ------------------------------------------------------------------------------
# 7. SCRITTURA EXCEL
# ------------------------------------------------------------------------------

def stili_excel() -> dict:
    """Stili condivisi da tutti i fogli. Creati lazy: openpyxl e' una dipendenza
    solo dell'output, non del motore di calcolo (il modello gira senza)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return dict(
        blu=Font(bold=True, color="1F4E78"),
        hdr_fill=PatternFill("solid", fgColor="D9E1F2"),
        giallo=PatternFill("solid", fgColor="FFF2CC"),
        verde=PatternFill("solid", fgColor="E2EFDA"),
        rosso_f=Font(color="C00000"),
        bordo=Border(bottom=Side(style="thin", color="B0B0B0")),
    )


def intesta_foglio(ws, riga: int, headers: list, widths=None, st: dict = None):
    """Scrive una riga di intestazione formattata e imposta le larghezze colonna."""
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=riga, column=j, value=h)
        c.font = Font(bold=True, size=9)
        c.fill = st["hdr_fill"]
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = st["bordo"]
        if widths and j - 1 < len(widths):
            ws.column_dimensions[get_column_letter(j)].width = widths[j-1]


def scrivi_excel(out_path: str, rep: dict, base: dict, pess: dict,
                 claw: dict, pack: dict, valid, serie_v14, sens: dict,
                 lungo: dict, lungo_pack: dict, p: Params,
                 griglia: dict | None = None, griglia_pess: dict | None = None,
                 mc: dict | None = None):
    import openpyxl
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    st = stili_excel()
    blu, hdr_fill = st["blu"], st["hdr_fill"]
    giallo, verde = st["giallo"], st["verde"]
    rosso_f, bordo = st["rosso_f"], st["bordo"]

    def intesta(ws, riga, headers, widths=None):
        intesta_foglio(ws, riga, headers, widths, st)

    # ---------------- Parametri -------------------------------------------------
    ws = wb.active; ws.title = "Parametri"
    ws["A1"] = "MODELLO v15 — PARAMETRI (motore Python: modello_v15.py)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("Celle gialle = ipotesi chiave. Il file e' GENERATO: modificare i "
                "parametri in modello_v15.py e rilanciare, non editare qui.")
    intesta(ws, 4, ["Parametro", "Valore", "Nota"], [46, 14, 90])
    note = {
        "anno0": "Anno di partenza simulazione", "anni": "Orizzonte 2027-2080",
        "pil0": "R1: PIL 2027 in volume, prezzi ~2025 (ISTAT post revisione set. 2025 + DFP 2026)",
        "g": "Crescita PIL reale (DEF 2025 / AR2024)",
        "w0": "Imponibile medio ponderato privati+pubblici (INPS Osservatori 2024)",
        "g_w": "Crescita salari reali (Eurostat ILC_DI03)",
        "occupati0": "Occupati perimetro L1 (FPLD+CTPS)",
        "migranti_netti": "C9: migranti netti/anno (RGS 26/2025: 176k)",
        "quota_eta_lav": "C9: quota migranti in eta' lavorativa (Moressa ~70%)",
        "tasso_occ_migr": "C9: tasso di occupazione migranti (~65%) — filtro ASSENTE in v14",
        "migrazione_v14": "Solo replica: B18 di v14 (0.116 = 165k x 70%, incoerente con B16=176k)",
        "aliquota": "Aliquota IVS dipendenti (33%)",
        "sq_anchors": "C1/R1: ancore (anno, spesa/PIL) DFP 2026 + RGS magg. 2026 — 15.4% 2027, "
                      "picco 17.1% 2041 (plateau 2044), 16.2% 2050, 14.0% 2070 poi costante; "
                      "interpolazione lineare tra le ancore",
        "ratio0_v14": "SOLO replica v14: 16.0% al 2027 (interpolazione AR2024, perimetro UE + PIL pre-revisione)",
        "quota_contrib": "C2: quota contributiva della spesa (Itinerari Prev. 2024)",
        "flusso1": "Flusso nuovi pensionati fino 2030 (mln) — RICOLLEGATO (morto in v14)",
        "flusso2": "Flusso 2031-2037", "flusso3": "Flusso dal 2038",
        "rf": "Rendimento reale fondi P2 (OECD, 60/40 net of fees)",
        "r_brdp": "Spread reale BRDP su HICP",
        "p0_base": "P0 base €/mese", "p0_cap": "Cap reddito: P0=0 oltre",
        "safety_net": "C6: safety net assorbito da P0 (€mld) — era B112, MORTO in v14",
        "irpef_pensioni": "C5: aliquota media IRPEF sulle pensioni (persa su P1 ridotta)",
        "irpef_contributi": "Leva B: tassa TEE sui contributi biforcati",
        "regime_p2": "TEE (come v14) o EET (rendite tassate, niente Leva B)",
        "durata_media_pensione": "Speranza di vita al pensionamento (ISTAT)",
        "clawback_p2": "Ricalibrazione P1: quota della rendita P2 prelevata dalla pensione pubblica (regola di policy; il vantaggio del pensionato ne e' la conseguenza)",
        "clawback_modo": "'costante' (default): aliquota fissa clawback_p2 + salvaguardia; 'target': P1 fissata perche' la pensione sia SQ + clawback_target",
        "clawback_target": "€/mese di vantaggio imposto sullo status quo (solo nel modo 'target')",
        "clawback_floor": "Salvaguardia: €/mese minimi sopra lo status quo sotto cui il prelievo si riduce (None = disattivata)",
        "levaI_regime": "R2: wedge +2 anni oltre il sentiero indicizzato, €mld a regime (ex levaC_eta)",
        "levaI_rampa": "R2: anni di phase-in del wedge eta' (ex levaC_anni)",
        "baseline_politico": "R2: 1 = SQ politico (sterilizzazioni adeguamenti ripetute) -> Leva I include anche 2.75+0.14/biennio",
        "levaJ_iniziale": "R2: chiusura deroghe Quota103/APe/OD, €mld iniziali (ex levaC_deroghe)",
        "levaJ_decad_anni": "R2: anni di esaurimento delle coorti aventi diritto alle deroghe",
        "pensionati0": "R1: INPS/ISTAT ~16.3 mln stabili (v10 assumeva crescita a 17.0 irrealistica)",
        "pensioni_vigenti0": "R1: INPS Osservatorio 1.1.2026 = 21.26 mln (+0.6%/anno)",
        "levaA_stock": "R1: COVIP Relazione 2026 — patrimonio fondi pensione fine 2025",
        "coeff69": "R1: DM 20/11/2024 tabella ufficiale (6.024%, divisore 16.600)",
        "r_debito": "Tasso reale sul debito di transizione accumulato",
        "pesi_profili": "Pesi profili reddito 0.5x/1x/1.5x per il taper P0",
    }
    r = 5
    for f_ in dataclasses.fields(p):
        val = getattr(p, f_.name)
        if isinstance(val, tuple):
            val = str(val)
        ws.cell(row=r, column=1, value=f_.name)
        c = ws.cell(row=r, column=2, value=val)
        if f_.name in note:
            ws.cell(row=r, column=3, value=note[f_.name])
            c.fill = giallo
        r += 1

    # ---------------- Modello (v15 base) ----------------------------------------
    def foglio_modello(nome, d):
        ws = wb.create_sheet(nome)
        pp = d.get("p", p)
        ws["A1"] = (f"{nome} — proiezione 2027-2080, €mld reali 2027. Positivo = risparmio/entrata. "
                    f"TASSI REALI del run: rf={pp.rf:.2%} (fondi P2), r_BRDP={pp.r_brdp:.2%}, "
                    f"r_debito={pp.r_debito:.2%}.")
        ws["A1"].font = Font(bold=True, size=12)
        headers = ["Anno", "t", "PIL reale (€mld)", "Occupati L1 (mln)",
                   "Salario medio (€)", "Spesa PAYG SQ (€mld)", "SQ %PIL",
                   "Spesa SQ del PERIMETRO riformato (€mld)",
                   "Spesa SQ altre gestioni, non riformate (€mld)",
                   "di cui contributiva (perimetro)", "di cui assistenza (perimetro)",
                   "Pensionati del perimetro (mln)", "Pensionati di sistema (mln)",
                   "%Lav in P2", "Lav. in P2 (mln)",
                   "Leva A — Imposta conv. stock fondi (€mld)",
                   "Leva B — IRPEF contributi biforcati TEE (€mld)",
                   "Leva D — Costo transizione: gap biforcazione (€mld)",
                   "Flusso ai fondi (€mld)",
                   "ΔP1 lordo (coorti)",
                   "  di cui: titolari (biforcazione)",
                   "  di cui: reversibilità 60%→45% (ex Leva H, endogena)",
                   "− IRPEF persa su P1",
                   "− P0 lordo", "+ assist. assorbita", "+ tassa rendite (EET)",
                   "LEVA F v15 — Risparmio netto sostituz. P1→P0/P2 (€mld)",
                   "Leva G — Sotto-indicizzazione perequazione (€mld)",
                   "Leva H — Revisione reversibilità superstiti (€mld)",
                   "Leva I — Età +2 anni oltre sentiero indicizzato (€mld)",
                   "Leva J — Chiusura deroghe e anticipazioni (€mld)",
                   "Imposta successione P2 (€mld)",
                   "BRDP emessi", "Stock BRDP", "Rendite P2 pagate (decumulo)",
                   "Pens. nuovi vivi (mln)",
                   "SALDO PRIMARIO", "Interessi (BRDP+debito)",
                   "Leva K costo (€mld)", "Leva K sconto cuneo (pp)",
                   "Aliquota IVS effettiva (33%-K)", "SALDO TOTALE",
                   "SALDO DI CASSA (+emissioni BRDP, -rimborsi)",
                   "Saldo di cassa cumulato (€mld)",
                   "RISPARMIO CUM. vs SQ (€mld)", "Risparmio cum. %PIL",
                   "Posizione netta vs SQ (+=risparmio, €mld)", "Posizione netta/PIL",
                   "Spesa rif. (€mld)", "Spesa rif./PIL", "Δ spesa vs SQ (pp PIL)",
                   "P1 nuovo pens. (€/m)", "P2 nuovo pens. (€/m)",
                   "P0 medio pesato — COSTO (€/m)",
                   "P0 del profilo medio — ADEGUATEZZA (€/m)",
                   "PENSIONE nuovo pens. = P1+P2+P0(m=1) (€/m)",
                   "P1 SQ contrf. (€/m)"]
        intesta(ws, 3, headers, [7, 4] + [11]*(len(headers)-2))
        cols = [d["anno"], d["t"], d["pil"], d["occ"], d["w"], d["H"], d["ratio"],
                d["H_perim"], d["H_altri"],
                d["H_contrib"], d["H_assist"],
                d["pensionati"], d["pensionati_sistema"], d["L"], d["M"],
                d["levaA"], d["levaB"], d["levaD"], d["R_fondi"],
                d["dP1_lordo"], d["dP1_titolari"], d["dP1_superstiti"],
                -d["irpef_persa"], -d["p0_lordo"],
                d["assist_assorbita"], d["tassa_rendite"], d["levaF"],
                d["levaG"], d["levaH"], d["levaI"], d["levaJ"],
                d["imposta_succ"],
                d["Z"], d["AH"], d["rendite_p2"], d["alive_new"],
                d["saldo_prim"], -d["interessi"],
                -d["levaK_costo"], d["levaK_sconto"], d["cuneo_totale"],
                d["saldo_tot"],
                d["saldo_cassa"], d["saldo_cassa_cum"],
                d["saldo_cum"], d["saldo_cum"] / d["pil"],
                d["saldo_cum"], d["saldo_cum"] / d["pil"],   # posizione netta (+=risparmio)
                d["spesa_rif"], d["ratio_rif"], d["ratio_rif"] - d["ratio"],
                d["p1_new"], d["p2_new"], d["p0_new"], d["p0_m1_new"],
                d["pensione_new"], d["p1_sq_new"]]
        for i in range(d["t"].size):
            for j, col in enumerate(cols, start=1):
                v = col[i]
                v = float(v) if isinstance(v, (np.floating, np.integer)) else v
                c = ws.cell(row=4 + i, column=j, value=v)
                if j <= 2:
                    c.number_format = "0"
                elif headers[j-1] in ("SQ %PIL", "Spesa rif./PIL", "Posizione netta/PIL",
                                      "%Lav in P2", "Risparmio cum. %PIL",
                                      "Δ spesa vs SQ (pp PIL)",
                                      "Leva K sconto cuneo (pp)", "Aliquota IVS effettiva (33%-K)"):
                    c.number_format = "0.00%"
                elif "€/m" in headers[j-1] or headers[j-1] == "Salario medio (€)":
                    c.number_format = "#,##0"
                else:
                    c.number_format = "0.00"
                if headers[j-1].startswith(("LEVA F v15", "SALDO PRIMARIO", "SALDO TOTALE")):
                    c.font = Font(bold=True)
        ws.freeze_panes = "C4"
        return ws

    foglio_modello("Modello", base)
    foglio_modello("Modello_Pess", pess)
    foglio_modello("Modello_C4_OFF", claw)   # counterfattuale senza clawback

    # ---------------- Dinamica (fondo con decumulo) ------------------------------
    ws = wb.create_sheet("Dinamica")
    ws["A1"] = ("Dinamica fondo pensione aggregato — €mld reali. CORREZIONE C4: il fondo "
                "DECUMULA: le rendite P2 dei pensionati escono dallo stock (in v14 non uscivano mai). "
                f"TASSI REALI: rendimento fondi rf={p.rf:.2%}, cedola BRDP r_BRDP={p.r_brdp:.2%}.")
    ws["A1"].font = Font(bold=True, size=12)
    intesta(ws, 3, ["Anno", "Afflussi (contributi P2+BRDP)", "Cedole BRDP",
                    "Rimborsi BRDP al fondo", "Rendimenti maturati (rf x stock)",
                    "Rendite P2 pagate (uscita)",
                    "Stock fondo (base rf={:.1%})".format(p.rf),
                    "Stock fondo %PIL",
                    "Pensionati con P2 (mln)", "Rendita P2 media (€/mese)",
                    "Pensioni totali erogate nell'anno (P1 rif + P0 + P2, €mld)",
                    "Quota pagata dai FONDI (P2 / totale)",
                    "Peso pubblico: spesa rif. %PIL", "Peso SQ %PIL",
                    "BRDP detenuti dal fondo (€mld)",
                    "PATRIMONIO TOTALE lavoratori (fondo + BRDP)",
                    "Stock SENZA decumulo (metodo v14)", "Sovrastima v14 (€mld)"],
            [7, 15, 11, 13, 15, 14, 15, 11, 12, 13, 20, 13, 12, 11, 13, 16, 16, 13])
    # ricostruzione cedole/rimborsi
    T = p.anni
    cedole = np.zeros(T); rimb = np.zeros(T)
    for i in range(T):
        cedole[i] = base["AH"][i-1] * p.r_brdp if i > 0 else 0.0
        rimb[i] = -base["AD"][i]
    fondo_v14_style = np.empty(T); s = 0.0
    for i in range(T):
        s = s * (1 + p.rf) + base["fondo_in"][i] + cedole[i] + rimb[i]
        fondo_v14_style[i] = s
    # pensionati con P2 vivi (solo coorti con montante > 0)
    pens_p2 = np.zeros(T)
    for i in range(T):
        tot = 0.0
        for c_ in base["coorti"]:
            a = p.anno0 + i - c_.anno_pens
            if a >= 0 and c_.montante_agg > 1e-9:
                tot += c_.N * max(0.0, 1.0 - a / (2 * p.durata_media_pensione))
        pens_p2[i] = tot
    for i in range(T):
        rendimenti = (base["fondo"][i-1] * p.rf) if i > 0 else 0.0
        rendita_media = (base["rendite_p2"][i] * 1000 / (pens_p2[i] * p.mensilita)
                         if pens_p2[i] > 1e-9 else 0.0)
        pens_totali = base["spesa_rif"][i] + base["rendite_p2"][i]
        quota_fondi = base["rendite_p2"][i] / pens_totali if pens_totali > 0 else 0.0
        vals = [int(base["anno"][i]), base["fondo_in"][i], cedole[i], rimb[i],
                rendimenti, base["fondo_out"][i], base["fondo"][i],
                base["fondo"][i] / base["pil"][i],
                pens_p2[i], rendita_media, pens_totali, quota_fondi,
                base["ratio_rif"][i], base["ratio"][i],
                base["AH"][i], base["fondo"][i] + base["AH"][i],
                fondo_v14_style[i], fondo_v14_style[i] - base["fondo"][i]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=4 + i, column=j, value=float(v))
            if j == 1:
                c.number_format = "0"
            elif j in (8, 12, 13, 14):
                c.number_format = "0.0%"
            elif j == 10:
                c.number_format = "#,##0"
            else:
                c.number_format = "0.00"
    ws.freeze_panes = "B4"

    # sensibilita' al rendimento rf
    r0s = 6 + T
    ws.cell(row=r0s, column=1,
            value="Sensibilita' al rendimento reale dei fondi (rf) — scenario base, tutto ricalcolato:").font = Font(bold=True)
    intesta(ws, r0s + 1, ["rf reale", "Fondo 2060 (€mld)", "Fondo 2080 (€mld)",
                          "Fondo 2080 %PIL", "Rendite P2 2080 (€mld)",
                          "P2 nuovo pens. 2080 (€/m)", "Leva F 2080 (€mld)",
                          "Saldo tot 2080 (€mld)"],
            [10, 14, 14, 12, 15, 16, 13, 14])
    for k, (rf_val, d_) in enumerate(sorted(sens.items())):
        i60, i80 = 2060 - p.anno0, 2080 - p.anno0
        vals = [rf_val, d_["fondo"][i60], d_["fondo"][i80],
                d_["fondo"][i80] / d_["pil"][i80], d_["rendite_p2"][i80],
                d_["p2_new"][i80], d_["levaF"][i80], d_["saldo_tot"][i80]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r0s + 2 + k, column=j, value=float(v))
            c.number_format = "0.00%" if j == 1 else ("0.0%" if j == 4 else ("#,##0" if j == 6 else "0.0"))

    # ---------------- Dinamica_rf: fondo anno per anno per ogni rf ----------------
    ws = wb.create_sheet("Dinamica_rf")
    ws["A1"] = ("Dinamica del fondo ANNO PER ANNO per rendimento reale rf — tutto ricalcolato "
                "(montanti, rendite, Leva F, saldo). €mld reali 2027. "
                f"r_BRDP={p.r_brdp:.2%} fisso in tutte le colonne; per la griglia completa "
                "r_BRDP x rf vedi Dinamica_scenari e la tabella in fondo al foglio Output.")
    ws["A1"].font = Font(bold=True, size=12)
    rf_sorted = sorted(sens.keys())
    hdr = ["Anno"]
    for rf_val in rf_sorted:
        hdr += [f"Stock rf={rf_val:.2%}", f"Rendite rf={rf_val:.2%}",
                f"Saldo tot rf={rf_val:.2%}"]
    intesta(ws, 3, hdr, [7] + [13] * (len(hdr) - 1))
    for i in range(T):
        ws.cell(row=4 + i, column=1, value=int(base["anno"][i])).number_format = "0"
        col = 2
        for rf_val in rf_sorted:
            d_ = sens[rf_val]
            for v in (d_["fondo"][i], d_["rendite_p2"][i], d_["saldo_tot"][i]):
                c = ws.cell(row=4 + i, column=col, value=float(v))
                c.number_format = "0.0"
                col += 1
    ws.freeze_panes = "B4"

    # ---------------- Dinamica_scenari: griglia r_BRDP x rf (struttura v14) -------
    # Replica le sezioni I-III del foglio Dinamica di v14 (anni in colonna),
    # ma ogni cella e' un run COMPLETO v15: fondo CON decumulo, cedole a r_BRDP,
    # reinvestimento a rf, interessi sul debito coerenti.
    def foglio_scenari(nome, gr, d0, d_pess_sez3=None):
        ws = wb.create_sheet(nome)
        rb_vals = sorted({k[0] for k in gr})
        rf_vals = sorted({k[1] for k in gr})
        anni = [int(a) for a in d0["anno"]]
        n = len(anni)
        ws.column_dimensions["A"].width = 52
        ws.column_dimensions["B"].width = 9
        ws.column_dimensions["C"].width = 9
        for j in range(n):
            ws.column_dimensions[get_column_letter(4 + j)].width = 9

        def titolo_sez(riga, testo):
            c = ws.cell(row=riga, column=1, value=testo)
            c.font = blu

        def testata(riga, titolo):
            c = ws.cell(row=riga, column=1, value=titolo)
            c.font = Font(bold=True, size=9); c.fill = hdr_fill
            for col, lbl in ((2, "r_BRDP_r"), (3, "rf_r")):
                cc = ws.cell(row=riga, column=col, value=lbl)
                cc.font = Font(bold=True, size=9); cc.fill = hdr_fill
            for j, a in enumerate(anni, start=4):
                cc = ws.cell(row=riga, column=j, value=a)
                cc.font = Font(bold=True, size=9); cc.fill = hdr_fill

        def riga_serie(riga, label, serie, rb=None, rf_=None, fmt="0.00"):
            ws.cell(row=riga, column=1, value=label).font = Font(size=9)
            for col, v in ((2, rb), (3, rf_)):
                cc = ws.cell(row=riga, column=col, value=v if v is not None else "—")
                cc.font = Font(size=9)
                if v is not None:
                    cc.number_format = "0.00%"
            for j in range(n):
                c = ws.cell(row=riga, column=4 + j, value=float(serie[j]))
                c.number_format = fmt

        def lbl_scen(rb, rf_):
            return f"r_BRDP={rb:.2%}, rf={rf_:.3%}"

        ws["A1"] = (f"{nome} — griglia r_BRDP x rf (struttura Dinamica v14). "
                    "€mld reali 2027. Fondo CON decumulo (C4); ogni riga = run completo v15.")
        ws["A1"].font = Font(bold=True, size=12)

        r = 3
        titolo_sez(r, "SEZIONE I — Montante fondi pensione P2, CON decumulo (€mld reali €2027)")
        r += 1
        testata(r, "Ia: Stock fondo P2 (scenario r_BRDP_r, rf_r)")
        r += 1
        for rb in rb_vals:
            for rf_ in rf_vals:
                riga_serie(r, lbl_scen(rb, rf_), gr[(rb, rf_)]["fondo"], rb, rf_)
                r += 1
        r += 1
        testata(r, "Ib: Stock BRDP detenuto dai fondi P2 (€mld reali €2027)")
        r += 1
        riga_serie(r, "Stock BRDP (= AH; non dipende da r_BRDP/rf)", d0["AH"])
        r += 2
        testata(r, "Ic: Portafoglio totale P2 = fondo + stock BRDP (€mld reali €2027)")
        r += 1
        for rb in rb_vals:
            for rf_ in rf_vals:
                riga_serie(r, lbl_scen(rb, rf_),
                           gr[(rb, rf_)]["fondo"] + d0["AH"], rb, rf_)
                r += 1

        r += 2
        titolo_sez(r, "SEZIONE II — Flussi BRDP verso i fondi P2 (€mld reali €2027)")
        r += 1
        testata(r, "IIa: Cedole BRDP ai fondi = stock(t-1) x r_BRDP_r")
        r += 1
        ced_per_rb = {}
        for rb in rb_vals:
            ced = np.zeros(n)
            for i in range(1, n):
                ced[i] = d0["AH"][i-1] * rb
            ced_per_rb[rb] = ced
            riga_serie(r, f"r_BRDP={rb:.2%}", ced, rb)
            r += 1
        r += 1
        testata(r, "IIb: Rimborsi capitale BRDP (inflow ai fondi P2)")
        r += 1
        rimb = -d0["AD"]
        riga_serie(r, "Rimborsi (= -AD; bullet 15 anni)", rimb)
        r += 2
        testata(r, "IIc: Flusso BRDP totale verso P2 = cedole + rimborsi")
        r += 1
        for rb in rb_vals:
            riga_serie(r, f"r_BRDP={rb:.2%}", ced_per_rb[rb] + rimb, rb)
            r += 1
        r += 1
        testata(r, "IId: Cedole cumulate BRDP (€mld reali €2027)")
        r += 1
        for rb in rb_vals:
            riga_serie(r, f"r_BRDP={rb:.2%}", np.cumsum(ced_per_rb[rb]), rb)
            r += 1

        if d_pess_sez3 is not None:
            r += 2
            titolo_sez(r, "SEZIONE III — Rendimento implicito PAYG: decomposizione contributiva/fiscale")
            r += 1
            for d_, tag in ((d0, "BASE"), (d_pess_sez3, "PESSIMISTA")):
                gettito = d_["H"] - d_["deficit_sq"]
                testata(r, f"IIIa: Spesa PAYG coperta da contributi — {tag} (€mld)")
                riga_serie(r + 1, tag, gettito)
                testata(r + 2, f"IIIb: Spesa PAYG coperta da fiscalita' generale — {tag} (€mld)")
                riga_serie(r + 3, tag, d_["deficit_sq"])
                testata(r + 4, f"IIIc: Quota spesa coperta da contributi — {tag} (%)")
                riga_serie(r + 5, tag, gettito / d_["H"], fmt="0.0%")
                r += 7
            testata(r, "IIId: Rendimento implicito PAYG — tasso di Aaron (crescita monte salari, % p.a.)")
            r += 1
            for d_, tag in ((d0, "BASE"), (d_pess_sez3, "PESSIMISTA")):
                monte = d_["w"] * d_["occ"]
                aaron = np.zeros(n)
                for i in range(1, n):
                    aaron[i] = monte[i] / monte[i-1] - 1.0
                aaron[0] = aaron[1]
                riga_serie(r, tag, aaron, fmt="0.00%")
                r += 1

        r += 2
        titolo_sez(r, "SEZIONE V — Effetti di equilibrio generale per scenario (solo v15: "
                      "in v14 la griglia era a bilancio pubblico invariato)")
        r += 1
        hdr5 = ["Scenario", "r_BRDP_r", "rf_r", "Fondo 2060", "Fondo 2080",
                "Fondo 2080 %PIL", "Rendite P2 2080", "Saldo tot 2080",
                "Debito trans. 2080 %PIL"]
        for j, h in enumerate(hdr5, start=1):
            c = ws.cell(row=r, column=j, value=h)
            c.font = Font(bold=True, size=9); c.fill = hdr_fill
            c.alignment = Alignment(wrap_text=True, vertical="center")
        r += 1
        i60 = min(2060 - p.anno0, n - 1); i80 = min(2080 - p.anno0, n - 1)
        for rb in rb_vals:
            for rf_ in rf_vals:
                d_ = gr[(rb, rf_)]
                vals = [lbl_scen(rb, rf_), rb, rf_, d_["fondo"][i60], d_["fondo"][i80],
                        d_["fondo"][i80] / d_["pil"][i80], d_["rendite_p2"][i80],
                        d_["saldo_tot"][i80], d_["debito"][i80] / d_["pil"][i80]]
                for j, v in enumerate(vals, start=1):
                    c = ws.cell(row=r, column=j, value=v)
                    c.font = Font(size=9)
                    if j in (2, 3):
                        c.number_format = "0.00%"
                    elif j in (6, 9):
                        c.number_format = "0.0%"
                    elif j > 3:
                        c.number_format = "0.0"
                r += 1

        r += 1
        ws.cell(row=r, column=1, value=(
            "Nota: la sezione IV di v14 (BRDP esteri) non e' replicata: il motore v15 "
            "non modella detentori esteri — tutti i BRDP sono detenuti dai fondi P2.")).font = Font(italic=True, size=9)
        ws.freeze_panes = "D2"

    if griglia:
        foglio_scenari("Dinamica_scenari", griglia, base, pess)
    if griglia_pess:
        foglio_scenari("Dinamica_scenari_Pess", griglia_pess, pess)

    # ---------------- Grafici ------------------------------------------------------
    ws = wb.create_sheet("Grafici")
    ws["A1"] = "GRAFICI — scenario base (dati nelle colonne A-L, grafici ancorati a destra)"
    ws["A1"].font = Font(bold=True, size=12)
    gh = ["Anno", "Spesa SQ %PIL", "Spesa riformata %PIL",
          "Sistema totale (pubbl.+P2) %PIL", "Saldo totale (€mld)",
          "Posizione netta vs SQ %PIL (+=risparmio)", "Rendite P2 %PIL", "Aliquota IVS effettiva (33%-K)"] + \
         [f"Fondo %PIL rf={rf_val:.2%}" for rf_val in rf_sorted] + \
         [f"Spesa PUBBLICA %PIL rf={rf_val:.2%}" for rf_val in rf_sorted]
    intesta(ws, 3, gh, [7] + [13] * (len(gh) - 1))
    col_sist0 = 9 + len(rf_sorted)          # 1a colonna del ventaglio "spesa pubblica"
    for i in range(T):
        sistema_tot = (base["spesa_rif"][i] + base["rendite_p2"][i]) / base["pil"][i]
        row = [int(base["anno"][i]), base["ratio"][i], base["ratio_rif"][i],
               sistema_tot, base["saldo_tot"][i],
               base["saldo_cum"][i] / base["pil"][i],
               base["rendite_p2"][i] / base["pil"][i], base["cuneo_totale"][i]] + \
              [sens[rf_val]["fondo"][i] / sens[rf_val]["pil"][i] for rf_val in rf_sorted] + \
              [sens[rf_val]["ratio_rif"][i] for rf_val in rf_sorted]
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=4 + i, column=j, value=float(v))
            c.number_format = "0" if j == 1 else ("0.0" if j == 5 else "0.0%")

    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties

    def linea(titolo, y_title, cols, anchor, y_fmt="0%",
              y_min=None, y_max=None, stili=None):
        """stili: lista per serie di (colore_hex, tratteggio|None, spessore_pt).
        y_min/y_max: scala esplicita dell'asse Y (senza, Excel parte da 0 e
        schiaccia le curve — era la causa del grafico 'diverso' dalla figura)."""
        ch = LineChart()
        ch.title = titolo
        ch.style = 2
        ch.height, ch.width = 9, 17
        ch.y_axis.title = y_title
        ch.y_axis.numFmt = y_fmt
        if y_min is not None:
            ch.y_axis.scaling.min = y_min
        if y_max is not None:
            ch.y_axis.scaling.max = y_max
        ch.x_axis.title = "Anno"
        for cmin in cols:
            ref = Reference(ws, min_col=cmin, min_row=3, max_row=3 + T)
            ch.add_data(ref, titles_from_data=True)
        if stili:
            for s, (colore, dash, spess) in zip(ch.series, stili):
                ln = LineProperties(solidFill=colore, w=int(spess * 12700))
                if dash:
                    ln.prstDash = dash
                s.graphicalProperties = GraphicalProperties(ln=ln)
                s.smooth = False
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + T)
        ch.set_categories(cats)
        ws.add_chart(ch, anchor)

    # SQ rosso pieno; poi il VENTAGLIO DELLA SPESA PUBBLICA per rendimento reale
    # dei fondi (piu' i fondi rendono, piu' il clawback abbatte P1 e piu' scende
    # la spesa pubblica) e infine il sistema totale del caso base.
    # NB: il sistema totale e' la somma di due movimenti opposti quasi uguali e
    # quindi appare piatto per ogni rf: ventagliare QUELLO nasconderebbe
    # l'effetto del rendimento. La variabile giusta e' la spesa PUBBLICA.
    _stili_rf_xl = {0.005: ("9E9E9E", "sysDot"), 0.0125: ("F9A825", "sysDashDot"),
                    0.025: ("1F4E78", None), 0.04: ("2E7D32", "lgDash")}
    cols_sist = list(range(col_sist0, col_sist0 + len(rf_sorted)))
    stili_sist = []
    for rf_val in rf_sorted:
        col, dash = _stili_rf_xl.get(rf_val, ("555555", "dash"))
        stili_sist.append((col, dash, 2.5 if abs(rf_val - p.rf) < 1e-12 else 1.5))
    linea("Spesa pensionistica %PIL: quanto i fondi abbattono la spesa PUBBLICA (per rendimento)",
          "% PIL", [2] + cols_sist + [4], "N3", y_fmt="0.0%", y_min=0.10, y_max=0.175,
          stili=[("C00000", None, 2.5)] + stili_sist + [("7030A0", "dash", 2.0)])
    linea("Saldo totale della riforma (€mld reali)", "€mld", [5], "N22", y_fmt="0")
    linea("Posizione netta della riforma vs status quo, % PIL (positivo = risparmio cumulato)",
          "% PIL", [6], "N41")
    linea("Fondo pensione in % del PIL per rendimento reale rf", "% PIL",
          list(range(9, 9 + len(rf_sorted))), "N60")
    linea("Leva K: aliquota contributiva IVS effettiva (33% - sconto)", "% salario",
          [8], "W3", y_fmt="0.0%", y_min=0.25, y_max=0.34,
          stili=[("2E7D32", None, 2.5)])

    # bar chart P0/P1/P2 del nuovo pensionato: 2030, 2055, 2080
    r_bar = 4 + T + 2
    ws.cell(row=r_bar, column=1,
            value="Composizione pensione del NUOVO pensionato (€/mese) — riforma vs status quo").font = Font(bold=True)
    intesta(ws, r_bar + 1, ["Scenario", "P1 (€/m)", "P2 (€/m)", "P0 (€/m)"], [16, 10, 10, 10])
    bar_rows = []
    for anno_b in (2030, 2055, 2080):
        i = anno_b - p.anno0
        bar_rows.append((f"{anno_b} riforma", base["p1_new"][i], base["p2_new"][i], base["p0_m1_new"][i]))
        bar_rows.append((f"{anno_b} status quo", base["p1_sq_new"][i], 0.0, 0.0))
    for k, (lbl, v1, v2, v3) in enumerate(bar_rows):
        ws.cell(row=r_bar + 2 + k, column=1, value=lbl)
        for j, v in enumerate((v1, v2, v3), start=2):
            ws.cell(row=r_bar + 2 + k, column=j, value=float(v)).number_format = "#,##0"
    ch = BarChart()
    ch.type = "col"; ch.grouping = "stacked"; ch.overlap = 100
    ch.title = "P1 + P2 + P0 del nuovo pensionato (€/mese): 2030, 2055, 2080"
    ch.height, ch.width = 10, 17
    ch.y_axis.title = "€/mese"
    data = Reference(ws, min_col=2, max_col=4, min_row=r_bar + 1, max_row=r_bar + 1 + len(bar_rows))
    cats = Reference(ws, min_col=1, min_row=r_bar + 2, max_row=r_bar + 1 + len(bar_rows))
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, "N79")

    # ---- DEBITO PENSIONISTICO cumulato dal 2027: SQ vs riforma (fino al 2100) ----
    r_deb = r_bar + 12
    ws.cell(row=r_deb, column=1,
            value=("Debito pensionistico CUMULATO dal 2027 in %PIL (deficit spesa-contributi, "
                   "con interessi) — senza riforma vs con riforma, fino al 2100")).font = Font(bold=True)
    intesta(ws, r_deb + 1, ["Anno", "SENZA riforma %PIL", "CON riforma %PIL",
                            "Differenza (riforma - SQ) %PIL"], [7, 16, 16, 18])
    TL = lungo["t"].size
    for i in range(TL):
        vals = [int(lungo["anno"][i]),
                lungo["debito_pens_sq"][i] / lungo["pil"][i],
                lungo["debito_pens_rif"][i] / lungo["pil"][i],
                lungo["debito"][i] / lungo["pil"][i]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r_deb + 2 + i, column=j, value=float(v))
            c.number_format = "0" if j == 1 else "0.0%"
    ch = LineChart()
    ch.title = "Debito pensionistico cumulato dal 2027 (%PIL): senza riforma vs con riforma"
    ch.height, ch.width = 10, 17
    ch.y_axis.title = "% PIL"; ch.y_axis.numFmt = "0%"; ch.x_axis.title = "Anno"
    for cmin in (2, 3):
        ch.add_data(Reference(ws, min_col=cmin, min_row=r_deb + 1,
                              max_row=r_deb + 1 + TL), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=r_deb + 2, max_row=r_deb + 1 + TL))
    ws.add_chart(ch, "W3")

    # ---- Fondo in €mld per rf (con decumulo/estrazioni) ---------------------------
    r_f = r_deb + TL + 4
    ws.cell(row=r_f, column=1,
            value="Fondo pensione in €mld reali per rendimento rf (estrazioni per rendite INCLUSE)").font = Font(bold=True)
    intesta(ws, r_f + 1, ["Anno"] + [f"Fondo €mld rf={rf_val:.2%}" for rf_val in rf_sorted]
            + [f"Rendite €mld rf={rf_val:.2%}" for rf_val in rf_sorted],
            [7] + [14] * (2 * len(rf_sorted)))
    for i in range(T):
        vals = [int(base["anno"][i])] + \
               [sens[rf_val]["fondo"][i] for rf_val in rf_sorted] + \
               [sens[rf_val]["rendite_p2"][i] for rf_val in rf_sorted]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r_f + 2 + i, column=j, value=float(v))
            c.number_format = "0" if j == 1 else "0.0"
    ch = LineChart()
    ch.title = "Fondo pensione (€mld reali) per rendimento rf — con estrazioni per il pagamento delle rendite"
    ch.height, ch.width = 10, 17
    ch.y_axis.title = "€mld"; ch.x_axis.title = "Anno"
    for k in range(len(rf_sorted)):
        ch.add_data(Reference(ws, min_col=2 + k, min_row=r_f + 1,
                              max_row=r_f + 1 + T), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=r_f + 2, max_row=r_f + 1 + T))
    ws.add_chart(ch, "W22")
    ch = LineChart()
    ch.title = "Rendite P2 pagate dai fondi (€mld reali) per rendimento rf"
    ch.height, ch.width = 10, 17
    ch.y_axis.title = "€mld"; ch.x_axis.title = "Anno"
    for k in range(len(rf_sorted)):
        ch.add_data(Reference(ws, min_col=2 + len(rf_sorted) + k, min_row=r_f + 1,
                              max_row=r_f + 1 + T), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=r_f + 2, max_row=r_f + 1 + T))
    ws.add_chart(ch, "W41")

    # ---------------- Canali di recupero ------------------------------------------
    ws = wb.create_sheet("Canali")
    ws["A1"] = ("Decomposizione LEVA F v15 nei canali (perche' in v14 'non c'era risparmio'): "
                "ogni colonna e' un canale separato, attivabile/disattivabile nei parametri.")
    ws["A1"].font = Font(bold=True, size=12)
    intesta(ws, 3, ["Anno", "Leva F v14 (replica, EMA+forfait)",
                    "ΔP1 lordo coorti", "IRPEF persa su P1 (C5)",
                    "P0 lordo calcolato (C6)", "Safety net assorbito (C6)",
                    "Tassa rendite P2 (se EET)", "LEVA F v15 netta",
                    "Rendite P2 ai pensionati (reddito privato)",
                    "Delta pensione nuovo pens. vs SQ (€/m)"],
            [7, 18, 13, 15, 14, 15, 14, 14, 18, 18])
    for i in range(T):
        delta_pens = (base["pensione_new"][i]
                      - base["p1_sq_new"][i])
        vals = [int(base["anno"][i]), rep["U"][i], base["dP1_lordo"][i],
                -base["irpef_persa"][i], -base["p0_lordo"][i],
                base["assist_assorbita"][i], base["tassa_rendite"][i],
                base["levaF"][i], base["rendite_p2"][i], delta_pens]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=4 + i, column=j, value=float(v))
            c.number_format = "0" if j == 1 else ("#,##0" if j == 10 else "0.00")
    ws.freeze_panes = "B4"

    # ---------------- Sostituzione --------------------------------------------------
    ws = wb.create_sheet("Sostituzione")
    ws["A1"] = ("Tassi di sostituzione LORDI — coorte a regime (carriera 42 anni interamente "
                "riformata), scala kappa contributiva. P0 con taper coerente col macro (stessa formula).")
    ws["A1"].font = Font(bold=True, size=12)
    intesta(ws, 3, ["Profilo", "Variante", "rf", "Salario finale (€/anno)",
                    "Pensione lorda (€/mese eq. 12m)", "TS lordo"],
            [14, 30, 10, 16, 18, 10])
    for i, riga in enumerate(tabella_sostituzione(p, base)):
        for j, v in enumerate(riga, start=1):
            c = ws.cell(row=4 + i, column=j, value=v)
            if j == 4 or j == 5:
                c.number_format = "#,##0"
            if j == 6:
                c.number_format = "0.0%"

    # ---------------- Output ---------------------------------------------------------
    ws = wb.create_sheet("Output")
    ws["A1"] = "Sintesi — anni chiave (scenario base e pessimistico)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (f"TASSI REALI dello scenario base: rf={p.rf:.2%} (fondi P2), "
                f"r_BRDP={p.r_brdp:.2%} (cedola reale BRDP), r_debito={p.r_debito:.2%} "
                f"(debito di transizione). Tutti i valori in €mld reali, prezzi 2027, "
                f"salvo diversa indicazione. La griglia completa r_BRDP x rf e' in fondo "
                f"a questo foglio e, anno per anno, nei fogli Dinamica_scenari(_Pess).")
    ws["A2"].font = Font(italic=True, size=9)
    anni_chiave = [2030, 2035, 2040, 2045, 2050, 2055, 2060, 2065, 2070, 2075, 2080]
    intesta(ws, 3, ["Indicatore"] + [str(a) for a in anni_chiave], [40] + [9]*len(anni_chiave))
    def riga_out(r, label, serie, d, fmt="0.00"):
        ws.cell(row=r, column=1, value=label).font = Font(size=9)
        for j, a in enumerate(anni_chiave, start=2):
            i = a - p.anno0
            c = ws.cell(row=r, column=j, value=float(serie[i]))
            c.number_format = fmt
        return r + 1
    r = 4
    for lbl, key, fmt, d in [
        ("BASE — Saldo primario (€mld)", "saldo_prim", "0.0", base),
        ("BASE — Saldo totale con interessi (€mld)", "saldo_tot", "0.0", base),
        ("BASE — Saldo di cassa (+emissioni BRDP, -rimborsi, €mld)", "saldo_cassa", "0.0", base),
        ("BASE — Saldo totale (%PIL)", None, "0.00%", base),
        ("BASE — Leva F v15 (€mld)", "levaF", "0.0", base),
        ("BASE — Posizione netta vs SQ (%PIL, +=risparmio)", None, "0.0%", base),
        ("BASE — Fondo pensione (%PIL, con decumulo)", None, "0.0%", base),
        ("BASE — Spesa SQ (%PIL)", "ratio", "0.0%", base),
        ("BASE — Spesa PUBBLICA riformata, sistema (%PIL)", "ratio_rif", "0.0%", base),
        ("BASE — di cui PERIMETRO riformato (%PIL)", "ratio_perim", "0.0%", base),
        ("BASE — Riduzione monte P1 (driver Leva K)", "riduzione_p1", "0.0%", base),
        ("BASE — Quota pensionati nel nuovo regime", "quota_nuovi", "0.0%", base),
        ("BASE — Rendite P2 dai fondi (%PIL)", None, "0.00%", base),
        ("BASE — SISTEMA TOTALE pubblico+P2 (%PIL)", None, "0.0%", base),
        ("BASE — Aliquota IVS effettiva 33%-K (Leva K)", "aliquota_ivs", "0.00%", base),
        ("BASE — Cuneo fiscale OCSE implicito (45.8%-K)", "cuneo_ocse_impl", "0.00%", base),
        ("BASE — Pensionati fisici (mln, ratio viva)", "pensionati", "0.0", base),
        ("BASE — Quota pensioni pagata dai fondi", "quota_fondi", "0.0%", base),
        ("BASE — Pensione tot. nuovo pens. vs SQ (€/mese)", None, "#,##0", base),
        ("PESS — Saldo totale (€mld)", "saldo_tot", "0.0", pess),
        ("PESS — Fondo pensione (%PIL)", None, "0.0%", pess),
        ("C4 OFF (senza clawback) — Leva F v15 (€mld)", "levaF", "0.0", claw),
        ("C4 OFF (senza clawback) — Saldo totale (€mld)", "saldo_tot", "0.0", claw),
        ("C4 OFF (senza clawback) — Posizione netta vs SQ (%PIL)", None, "0.0%", claw),
        ("C4 OFF (senza clawback) — Pens. tot. nuovo pens. vs SQ (€/mese)", None, "#,##0", claw),
        ("PACCHETTO ANTICIPO (clawback+taxgap) — Saldo totale (€mld)", "saldo_tot", "0.0", pack),
        ("PACCHETTO ANTICIPO — Posizione netta vs SQ (%PIL)", None, "0.0%", pack),
        ("PACCHETTO ANTICIPO — Pens. tot. nuovo pens. vs SQ (€/mese)", None, "#,##0", pack),
    ]:
        if key is not None:
            r = riga_out(r, lbl, d[key], d, fmt)
        elif "totale (%PIL)" in lbl:
            r = riga_out(r, lbl, d["saldo_tot"] / d["pil"], d, fmt)
        elif "Posizione netta" in lbl:
            r = riga_out(r, lbl, d["saldo_cum"] / d["pil"], d, fmt)
        elif "Fondo" in lbl:
            r = riga_out(r, lbl, d["fondo"] / d["pil"], d, fmt)
        elif "Rendite P2" in lbl:
            r = riga_out(r, lbl, d["rendite_p2"] / d["pil"], d, fmt)
        elif "SISTEMA TOTALE" in lbl:
            r = riga_out(r, lbl, (d["spesa_rif"] + d["rendite_p2"]) / d["pil"], d, fmt)
        elif "vs SQ" in lbl:
            delta = d["pensione_new"] - d["p1_sq_new"]
            r = riga_out(r, lbl, delta, d, fmt)

    # ---- orizzonte lungo: picco e pareggio del debito di transizione
    r += 1
    for lg, nome_lg in ((lungo, "BASE"), (lungo_pack, "PACCHETTO ANTICIPO")):
        if lg is None:
            continue
        i_picco = int(np.argmax(lg["debito"]))
        anni_zero = [int(lg["anno"][i]) for i in range(i_picco, lg["t"].size)
                     if lg["debito"][i] <= 0]
        pareggio = str(anni_zero[0]) if anni_zero else "oltre l'orizzonte"
        ws.cell(row=r, column=1, value=(
            f"PROIEZIONE LUNGA {nome_lg} (a {int(lg['anno'][-1])}): picco debito transizione nel "
            f"{int(lg['anno'][i_picco])} a {lg['debito'][i_picco]:.0f} €mld "
            f"({lg['debito'][i_picco]/lg['pil'][i_picco]:.1%} PIL); ritorno a debito ZERO nel {pareggio}; "
            f"al {int(lg['anno'][-1])}: saldo {lg['saldo_tot'][-1]:+.0f} €mld/anno, "
            f"debito {lg['debito'][-1]:.0f} €mld ({lg['debito'][-1]/lg['pil'][-1]:.1%} PIL).")).font = Font(bold=True, size=10)
        r += 1

    # ---- TAVOLA RIEPILOGATIVA: il peso di tutto al 2060 e al 2080 (scenario base)
    r += 2
    ws.cell(row=r, column=1, value="TAVOLA RIEPILOGATIVA — il peso di ogni componente (scenario base)").font = Font(bold=True, size=12)
    r += 1
    intesta(ws, r, ["Componente", "2060 (€mld)", "2080 (€mld)", "2080 (%PIL)",
                    "Segno", "Spiegazione"], [42, 11, 11, 10, 8, 70])
    i60, i80 = 2060 - p.anno0, 2080 - p.anno0
    d = base
    pil80 = d["pil"][i80]
    tavola = [
        ("Leva A — imposta conversione TEE", d["levaA"], "+", "Una tantum, solo 2027-2029"),
        ("Leva B — IRPEF sui contributi biforcati (TEE)", d["levaB"], "+",
         "27% sui contributi che vanno a P2: gettito immediato e permanente"),
        ("Leva D — contributi persi dal PAYG (biforcazione)", d["levaD"], "−",
         "IL COSTO STRUTTURALE: ogni euro deviato ai fondi manca alla cassa INPS"),
        ("  di cui ΔP1 titolari (biforcazione + clawback)", d["dP1_titolari"], "info",
         "La componente che dipende dalla biforcazione e dal clawback"),
        ("  di cui Δ reversibilità 60%→45% (ex Leva H, ENDOGENA)", d["dP1_superstiti"], "info",
         "Il risparmio da revisione della reversibilità: era il forfait di Leva H, ora calcolato"),
        ("  ΔP1 lordo (P1 ridotte delle nuove coorti)", d["dP1_lordo"], "+",
         "IL RISPARMIO STRUTTURALE: matura solo quando le coorti biforcate si pensionano (dal 2051)"),
        ("  − IRPEF persa sulle P1 ridotte", -d["irpef_persa"], "−", "15% del ΔP1"),
        ("  − P0 lordo erogato ai nuovi pensionati", -d["p0_lordo"], "−",
         "Taper 54% su P1+P2; costa di piu' nei primi decenni (P1 ancora piena)"),
        ("  + Safety net assorbito da P0", d["assist_assorbita"], "+",
         "Assegno sociale/integrazione che P0 sostituisce (32.5 su quota nuovi pensionati)"),
        ("LEVA F v15 (netta) [somma 4 righe sopra]", d["levaF"], "±", "Risparmio P1 netto IRPEF, meno P0 netto safety net"),
        ("Leva G — sotto-indicizzazione", d["levaG"], "+", "Decade -1.5%/anno"),
        ("Leva H — reversibilita'", d["levaH"], "+", "5 €mld costanti"),
        ("Leva I — eta' +2 anni oltre sentiero indicizzato", d["levaI"], "+",
         "Rampa 15 anni fino a 9 €mld, poi costante (assorbe ex Leva C eta'; R2)"),
        ("Leva J — chiusura deroghe/anticipazioni", d["levaJ"], "+",
         "3.5 €mld decadenti a zero in 10 anni (assorbe ex Leva C deroghe; R2)"),
        ("Leva L — recupero gap contributivo", d["levaL"], "+",
         "10 €mld/anno a regime; OFF nello scenario base, ON nel pacchetto"),
        ("Imposta successione su montanti P2", d["imposta_succ"], "+",
         "Una tantum progressiva (0-18%) sul conto individuale trasferito al superstite"),
        ("SALDO PRIMARIO", d["saldo_prim"], "=", "Somma leve attive (A,B,D,F,G,H,I,J,+L)"),
        ("Interessi (BRDP + debito transizione)", -d["interessi"], "−",
         "Negativi se il cumulato e' debito; positivi se e' risparmio investito"),
        ("Leva K — taglio del cuneo (costo)", -d["levaK_costo"], "−",
         "Sconto ancorato al calo del monte P1 (riduzione_p1), finanziato al piu' "
         "dalla quota levaK_quota_dividendo del dividendo netto P1 dell'anno"),
        ("SALDO TOTALE", d["saldo_tot"], "=", "Primario - interessi - Leva K"),
        ("SALDO DI CASSA", d["saldo_cassa"], "=",
         "Saldo totale + emissioni BRDP (entrata) - rimborsi bullet (uscita). "
         "Cumulato = posizione netta + stock BRDP vivo: a titoli estinti coincidono."),
        ("POSIZIONE NETTA vs status quo (+=risparmio)", d["saldo_cum"], "=",
         "Somma cumulata dei saldi totali. Positivo = la riforma ha ACCUMULATO risparmio; "
         "negativo = ha richiesto finanziamento (debito di transizione). E' il -debito."),
        ("Fondo pensione (stock, con decumulo)", d["fondo"], "info",
         "Ricchezza PRIVATA dei lavoratori/pensionati, non dello Stato"),
        ("Rendite P2 pagate dai fondi nell'anno", d["rendite_p2"], "info",
         "Pensioni che lo Stato NON deve piu' pagare: escono dal fondo, non dal bilancio pubblico"),
    ]
    for k, (lbl, serie, segno, spieg) in enumerate(tavola):
        rr = r + 1 + k
        ws.cell(row=rr, column=1, value=lbl).font = Font(size=9, bold=lbl[:4].isupper() or "LEVA F" in lbl)
        ws.cell(row=rr, column=2, value=float(serie[i60])).number_format = "0.0"
        ws.cell(row=rr, column=3, value=float(serie[i80])).number_format = "0.0"
        ws.cell(row=rr, column=4, value=float(serie[i80]) / pil80).number_format = "0.00%"
        ws.cell(row=rr, column=5, value=segno)
        ws.cell(row=rr, column=6, value=spieg).font = Font(size=8, italic=True)
    rr = r + 2 + len(tavola)
    ws.cell(row=rr, column=1,
            value=("Lettura della posizione netta: e' GIA' la differenza rispetto al modello senza riforma "
                   "(il modello calcola solo flussi differenziali). Positivo = la riforma ha ACCUMULATO risparmio; "
                   "negativo = ha richiesto finanziamento (il 'debito di transizione'). Non e' monotona: le leve "
                   "parametriche generano surplus nei primi anni (2027-2042), poi la Leva D (contributi persi) cresce "
                   "mentre il ΔP1 delle coorti biforcate non e' ancora maturato (gobba 2045-2075: la posizione netta "
                   "scende), e il risparmio strutturale arriva solo a regime.")).font = Font(size=9, italic=True)

    # ---- GRIGLIA r_BRDP x rf (come il foglio Output/Dinamica di v10): ogni riga
    # dichiara la coppia di tassi reali a cui e' calcolata; ogni riga e' un run
    # COMPLETO v15 (fondo con decumulo, cedole a r_BRDP, debito coerente).
    for gr_, nome_gr in ((griglia, "BASE"), (griglia_pess, "PESSIMISTICO")):
        if not gr_:
            continue
        rr += 3
        ws.cell(row=rr, column=1, value=(
            f"GRIGLIA SCENARI r_BRDP x rf — {nome_gr} (16 run completi v15; "
            f"€mld reali 2027; dettaglio anno per anno nel foglio "
            f"Dinamica_scenari{'_Pess' if nome_gr != 'BASE' else ''})")).font = Font(bold=True, size=12)
        rr += 1
        intesta(ws, rr, ["Scenario (tassi reali dichiarati)", "r_BRDP_r", "rf_r",
                         "Fondo 2060", "Fondo 2080", "Fondo 2080 %PIL",
                         "Rendite P2 2080", "Spesa pubbl. 2080 %PIL",
                         "Sistema tot. 2080 %PIL", "Saldo tot 2080",
                         "Debito trans. 2080 %PIL", "Pens. tot nuovo pens. 2080 (€/mese)"],
                [30, 9, 9, 11, 11, 12, 12, 13, 13, 11, 13, 16])
        rb_vals = sorted({k[0] for k in gr_})
        rf_vals = sorted({k[1] for k in gr_})
        i60o, i80o = 2060 - p.anno0, 2080 - p.anno0
        for rb in rb_vals:
            for rf_ in rf_vals:
                rr += 1
                d_ = gr_[(rb, rf_)]
                pens80 = float(d_["pensione_new"][i80o])
                vals = [f"r_BRDP={rb:.2%}, rf={rf_:.3%}", rb, rf_,
                        float(d_["fondo"][i60o]), float(d_["fondo"][i80o]),
                        float(d_["fondo"][i80o] / d_["pil"][i80o]),
                        float(d_["rendite_p2"][i80o]),
                        float(d_["ratio_rif"][i80o]),
                        float((d_["spesa_rif"][i80o] + d_["rendite_p2"][i80o]) / d_["pil"][i80o]),
                        float(d_["saldo_tot"][i80o]),
                        float(d_["debito"][i80o] / d_["pil"][i80o]), pens80]
                fmts = [None, "0.00%", "0.000%", "0.0", "0.0", "0.0%", "0.0",
                        "0.0%", "0.0%", "0.0", "0.0%", "#,##0"]
                for j, (v, fm) in enumerate(zip(vals, fmts), start=1):
                    c = ws.cell(row=rr, column=j, value=v)
                    c.font = Font(size=9)
                    if fm:
                        c.number_format = fm
                    if rb == p.r_brdp and rf_ == p.rf and nome_gr == "BASE":
                        c.fill = verde        # evidenzia la combinazione base
        rr += 1
        ws.cell(row=rr, column=1, value=(
            "La riga evidenziata in verde e' la combinazione dello scenario base "
            f"(r_BRDP={p.r_brdp:.2%}, rf={p.rf:.2%})." if nome_gr == "BASE" else
            "Scenario pessimistico: g=0.3%, salari fermi; stessi 16 incroci di tassi.")).font = Font(size=9, italic=True)

    # ---------------- Audit anno base --------------------------------------------------
    ws = wb.create_sheet("Audit_anno_base")
    ws["A1"] = "AUDIT DELL'ANNO BASE — il modello riproduce il 2027 osservato?"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("Un modello che non riproduce l'anno base non e' affidabile sul lungo periodo. "
                "Le righe marcate SCOSTAMENTO sono limiti DICHIARATI, non errori nascosti: "
                "vanno lette prima di usare i risultati.")
    ws["A2"].font = Font(italic=True, size=9)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    intesta(ws, 4, ["Voce", "Modello", "Osservato", "Scarto %", "Fonte del dato osservato",
                    "Giudizio"], [40, 13, 13, 11, 44, 15])
    for i, (voce, mod, oss, fonte, giud) in enumerate(audit_anno_base(p, base)):
        rr = 5 + i
        pct = "Copertura" in voce or "PIL" in voce
        ws.cell(row=rr, column=1, value=voce).font = Font(size=9)
        ws.cell(row=rr, column=2, value=float(mod)).number_format = "0.0%" if pct else "#,##0.0"
        ws.cell(row=rr, column=3, value=float(oss)).number_format = "0.0%" if pct else "#,##0.0"
        c = ws.cell(row=rr, column=4, value=float(mod) / float(oss) - 1 if oss else 0)
        c.number_format = "+0.0%;-0.0%"
        ws.cell(row=rr, column=5, value=fonte).font = Font(size=8, italic=True)
        cg = ws.cell(row=rr, column=6, value=giud)
        cg.font = rosso_f if giud != "ok" else Font(size=9)
    rr = 6 + len(audit_anno_base(p, base))
    ws.cell(row=rr, column=1, value=(
        "MONTE SALARI (correzione 2026-08). w0 e' ora l'imponibile medio PONDERATO privati+pubblici "
        "(32.300 €), derivato bottom-up dagli Osservatori INPS 2024: privati 17,7 mln x 24.486 € "
        "(433,4 €mld) + pubblici 3,74 mln x 35.350 € (132,1 €mld) = 565,5 €mld su 17,5 mln posizioni. "
        "Il vecchio 26.700 € era l'imponibile dei soli privati FPLD e produceva un monte ~21% sotto "
        "l'osservato, sottostimando la Leva D (costo di transizione). Il gettito IVS resta ~11% sotto "
        "i 209,7 €mld di Itinerari: il residuo e' contribuzione FIGURATIVA a carico GIAS (CIG, "
        "malattia, maternita') che non transita dal monte salari effettivo — sottostima prudente "
        "della base contributiva, dichiarata.")).font = Font(size=9, italic=True)
    ws.cell(row=rr, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 40
    ws.row_dimensions[rr].height = 90

    # ---------------- Stocastico (Monte Carlo su rf) -----------------------------------
    if mc is not None:
        ws = wb.create_sheet("Stocastico")
        ws["A1"] = (f"MONTE CARLO sul rendimento reale dei fondi — {mc['n_sim']} simulazioni, "
                    f"rf annuo i.i.d. ~ Normale(media {p.rf:.2%}, dev.st. {p.rf_sigma:.1%}).")
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = ("Ogni simulazione e' un run COMPLETO del modello con quel sentiero di rendimenti "
                    "(montanti capitalizzati anno per anno, fondo rivalutato al rendimento dell'anno). "
                    "La MEDIANA sta sotto il run deterministico per VOLATILITY DRAG: con rendimenti "
                    "i.i.d. la crescita composta attesa e' circa mu - sigma^2/2. Il ventaglio misura "
                    "il rischio di mercato che la riforma trasferisce ai pensionati. "
                    "ATTENZIONE ALLA LETTURA: i percentili sono calcolati SERIE PER SERIE (marginali) "
                    "— la colonna '95°' NON e' un unico scenario. Per il fondo il 95° percentile viene "
                    "dai sentieri con rendimenti ALTI, per la spesa pubblica dai sentieri con rendimenti "
                    "BASSI (meno clawback => P1 piu' alta => piu' spesa): per questo 'fondo alto' e "
                    "'spesa alta' compaiono nella stessa colonna pur venendo da simulazioni OPPOSTE. "
                    "Fra fondo e spesa pubblica la correlazione e' NEGATIVA, come atteso.")
        ws["A2"].font = Font(italic=True, size=9)
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        r_mc = 4
        for k, (lbl, _) in MC_SERIE.items():
            ws.cell(row=r_mc, column=1, value=lbl).font = blu
            intesta(ws, r_mc + 1,
                    ["Anno"] + [f"{q}° perc." for q in MC_PERCENTILI] + ["Deterministico"],
                    [7] + [12] * (len(MC_PERCENTILI) + 1))
            pct = "%PIL" in lbl or "€/mese" not in lbl and "€mld" not in lbl
            for i in range(T):
                vals = ([int(mc["anno"][i])]
                        + [float(mc["perc"][k][q][i]) for q in MC_PERCENTILI]
                        + [float(mc["det_serie"][k][i])])
                for j, v in enumerate(vals, start=1):
                    c = ws.cell(row=r_mc + 2 + i, column=j, value=v)
                    c.number_format = ("0" if j == 1 else
                                       ("0.00%" if pct else
                                        ("#,##0" if "€/mese" in lbl else "0.0")))
            r_mc += T + 4
        # sintesi di rischio al 2080
        sint = sintesi_montecarlo(mc, 2080)
        ws.cell(row=r_mc, column=1, value="SINTESI DI RISCHIO AL 2080").font = Font(bold=True, size=12)
        intesta(ws, r_mc + 1, ["Indicatore", "5° perc.", "Mediana", "95° perc.",
                               "Deterministico", "Drag (mediana - determ.)"],
                [36, 12, 12, 12, 14, 22])
        rr_mc = r_mc + 2
        for k, (lbl, su_pil) in MC_SERIE.items():
            d_ = sint[k]
            fmt = "0.00%" if (su_pil or "%PIL" in lbl) else ("#,##0" if "€/mese" in lbl else "0.0")
            ws.cell(row=rr_mc, column=1, value=lbl).font = Font(size=9)
            for j, key in enumerate(("p5", "mediana", "p95", "det", "drag"), start=2):
                ws.cell(row=rr_mc, column=j, value=float(d_[key])).number_format = fmt
            rr_mc += 1
        rr_mc += 1
        for testo, val in (
                ("Probabilita' di posizione netta NEGATIVA al 2080 (riforma a debito)",
                 sint["prob_pos_netta_negativa"]),
                ("Probabilita' che la pensione del nuovo pensionato sia SOTTO lo status quo",
                 sint["prob_pensione_sotto_sq"])):
            ws.cell(row=rr_mc, column=1, value=testo).font = Font(size=9, bold=True)
            ws.cell(row=rr_mc, column=2, value=float(val)).number_format = "0.0%"
            rr_mc += 1

    # ---------------- Validazione v14 --------------------------------------------------
    ws = wb.create_sheet("Validazione_v14")
    ws["A1"] = ("Validazione: il motore Python replica il foglio 'Modello' di v14 "
                "(stesse formule) e confronta con i valori salvati nel file v14.")
    ws["A1"].font = Font(bold=True, size=12)
    intesta(ws, 3, ["Colonna v14", "Max |diff| (€mld o unita')", "Media |diff|", "Esito"],
            [16, 22, 16, 12])
    if valid:
        for i, (col, mx, mean) in enumerate(valid):
            ws.cell(row=4+i, column=1, value=col)
            ws.cell(row=4+i, column=2, value=mx).number_format = "0.0000"
            ws.cell(row=4+i, column=3, value=mean).number_format = "0.0000"
            ok = "OK" if mx < 0.05 else "VERIFICARE"
            c = ws.cell(row=4+i, column=4, value=ok)
            if ok != "OK":
                c.font = rosso_f
    else:
        ws["A4"] = "File v14 non trovato: validazione non eseguita."
    r0 = 6 + (len(valid) if valid else 1)
    ws.cell(row=r0, column=1,
            value=("CONFRONTO v14 vs v15 anno per anno — la replica usa i parametri CONGELATI di v14 "
                   "(fasi 6/8/10/12, P0 forfait); v15 usa i parametri correnti (incl. tue modifiche).")).font = Font(bold=True)
    intesta(ws, r0+1, ["Anno",
                       "Leva F v14 (file)", "Leva F replica", "Leva F v15",
                       "Saldo cassa v14 (file)", "Saldo primario v15", "Saldo totale v15",
                       "Saldo cum v14 (file)", "Risparmio cum v15",
                       "Fondo v14 (senza decumulo)", "Fondo v15 (con decumulo)"],
            [7, 12, 12, 12, 14, 13, 13, 13, 13, 16, 16])
    for i in range(T):
        u14 = serie_v14["U"][i] if serie_v14 is not None else np.nan
        ae14 = serie_v14["AE"][i] if serie_v14 is not None else np.nan
        af14 = serie_v14["AF"][i] if serie_v14 is not None else np.nan
        for j, v in enumerate([int(rep["anno"][i]), u14, rep["U"][i], base["levaF"][i],
                               ae14, base["saldo_prim"][i], base["saldo_tot"][i],
                               af14, base["saldo_cum"][i],
                               rep["fondo_v14"][i], base["fondo"][i]], start=1):
            if isinstance(v, float) and np.isnan(v):
                continue
            c = ws.cell(row=r0+2+i, column=j, value=float(v))
            c.number_format = "0" if j == 1 else "0.00"

    # ---------------- Note metodologiche -----------------------------------------------
    ws = wb.create_sheet("Note")
    ws.column_dimensions["A"].width = 130
    note_txt = [
        "NOTE METODOLOGICHE — v15 (generato da modello_v15.py)",
        "",
        "Differenze rispetto a v14 (ognuna e' una correzione documentata):",
        "C1 (agg. R1 2026-07) Status quo per ANCORE ufficiali DFP 2026/RGS: 15.4% (2027), 15.5% (2028-29), "
        "picco 17.1% (2041, plateau fino al 2044), 16.2% (2050), 14.0% (2070) e poi costante. Sostituisce sia la "
        "discesa lineare infinita di v14 sia il livello 16.0% al 2027 (interpolazione AR2024: perimetro UE piu' "
        "ampio e PIL pre-revisione ISTAT set. 2025).",
        "C2 La spesa e' divisa in quota contributiva (72.8%) e assistenziale: la riduzione di P1 da "
        "biforcazione si applica solo alla quota contributiva. In v14 la Leva F si applicava a tutta la spesa "
        "(sovrastima ~20 €mld al 2080 su quel lato).",
        "C3 Contabilita' per coorti: ogni coorte di pensionamento ha P1 residua NDC, montante P2 (quota fondi "
        "a rf, quota BRDP a HICP+1.5%), P0 con taper, mortalita' lineare (durata media 19.5 anni). "
        "La partecipazione segue il phase-in per eta': le coorti che si pensionano prima del 2051 non hanno "
        "mai biforcato (in v14 la formula AK biforcava tutti dal 2027, incoerente con la colonna %Lav in P2).",
        "C4 Il fondo DECUMULA: le rendite P2 escono dallo stock. Il foglio Dinamica mostra anche lo stock "
        "'metodo v14' e la sovrastima (centinaia di €mld a fine orizzonte).",
        "C5 Fiscalita' onesta: la riduzione di P1 fa perdere IRPEF (aliquota media 15%); regime TEE come v14 "
        "(Leva B tassa i contributi biforcati al 27%) oppure EET (rendite tassate, Leva B=0) — parametro regime_p2.",
        "C6 P0 calcolato: taper 54% (=650/1200) sul reddito P1+P2 per tre profili di reddito (0.5x/1x/1.5x, pesi 50/35/15), "
        "al netto del safety net assorbito (32.5 €mld — parametro che in v14 esisteva ma non era collegato). "
        "Sostituisce il forfait hard-coded di 13 €mld.",
        "C7 Saldo primario separato dal finanziamento: le emissioni BRDP non sono entrate e i rimborsi non sono "
        "costi; interessi su BRDP (HICP+1.5% reale) e sul debito di transizione accumulato (1.5% reale). "
        "In v14 il 'saldo annuo' era un flusso di cassa che contava il debito emesso come entrata. "
        "DAL 2026-08 il modello espone TRE saldi: PRIMARIO (leve, senza finanziamento), TOTALE "
        "(primario - interessi - Leva K) e DI CASSA (totale + emissioni BRDP - rimborsi bullet): "
        "nel saldo di cassa il debito e' prima un'entrata (emissione) e poi un'uscita (interessi anno "
        "per anno e capitale al rimborso). Cumulata di cassa = posizione netta + stock BRDP vivo.",
        "C8 (storica) Sovrapposizione Leva C / Leva J: superata da R2 (Leva C eliminata).",
        "",
        "R1 (2026-07) DATI AGGIORNATI: pensioni vigenti 21.4 mln e pensionati 16.4 mln (INPS Osservatorio "
        "1.1.2026: 21.26 mln, +0.6%/anno); stock fondi 262 €mld (COVIP Relazione 2026); coefficiente a 69 anni "
        "6.024% (DM 20/11/2024, divisore 16.600); PIL base 2247 €mld in volume a prezzi ~2025.",
        "R2 (2026-07) LEVE RAZIONALIZZATE: Leva E eliminata (costo puro, nessun ritorno demografico modellato). "
        "Leva C eliminata: il baseline RGS a legislazione vigente include GIA' l'indicizzazione dell'eta' alla "
        "speranza di vita (>20pp PIL cumulati al 2060 per MEF/RGS), quindi contare per intero '67->69 + deroghe' "
        "era un DOPPIO CONTEGGIO. Nuova Leva I = wedge di +2 anni oltre il sentiero indicizzato (rampa 15 anni "
        "fino a 9 €mld, poi costante); nuova Leva J = chiusura deroghe/anticipazioni (3.5 €mld decadenti a zero "
        "in 10 anni: le coorti aventi diritto si esauriscono entro il ~2037). Flag baseline_politico (default "
        "OFF): aggiunge alla Leva I la componente anti-sterilizzazione (2.75+0.14/biennio), legittima solo "
        "dichiarando un baseline politico in cui le sterilizzazioni post-2019 continuano.",
        "C9 Immigrazione corretta e decomposta: 176k netti (RGS) x 70% eta' lavorativa x 65% tasso di occupazione "
        "= ~80k occupati/anno, flusso costante. In v14: 0.116 mln (=165k x 70%, incoerente col parametro 176k), "
        "SENZA filtro occupazionale, e con un fattore (1-n)^(t-1) che faceva crescere il flusso dello 0.5%/anno "
        "senza giustificazione. Effetto: v15 ha ~36k occupati/anno in meno da migrazione (piu' prudente).",
        "",
        "GRIGLIA r_BRDP x rf (fogli Dinamica_scenari e Dinamica_scenari_Pess): replica la struttura del "
        "foglio Dinamica di v14 (16 combinazioni 4x4, anno per anno), ma ogni scenario e' un run COMPLETO "
        "del motore v15: fondo con decumulo, cedole a r_BRDP reinvestite a rf, montanti BRDP capitalizzati "
        "a r_BRDP, interessi sul debito coerenti. In v14 la griglia variava solo il fondo a bilancio "
        "pubblico invariato; qui la Sezione V mostra anche gli effetti di equilibrio generale (saldo, debito).",
        "",
        "REPLICA v14: usa parametri CONGELATI (params_v14: fasi 6/8/10/12, cap P0 1250, ecc.) — resta valida "
        "anche se si ricalibrano i parametri di v15. Le fasi di biforcazione correnti di v15 sono una SCELTA "
        "dell'utente (4/6/8/12: Fase II da t=10, III da t=25, IV da t=40, rampe di 5 anni) e possono differire da v14.",
        "",
        "Cosa NON e' cambiato rispetto a v14: leve A (salvo stock aggiornato), B, G, H (stessi importi e "
        "profili), BRDP (2% con bullet a 15 anni), aliquota 33%, phase-in %Lav in P2. Cambiate con fonte: "
        "sentiero SQ (R1), leve C/E/I/J (R2), coeff. 69 anni e platee (R1).",
        "",
        "COME ENTRA P0 (definizione operativa): P0 spetta SOLO ai nuovi pensionati (i vecchi restano nel regime "
        "attuale). Per ogni coorte, P0 pro-capite = max(0; 650 - taper x reddito pensionistico), taper = 650/cap, "
        "calcolato su tre profili di reddito (0.5x/1x/1.5x della pensione media, pesi 50/35/15). Il costo lordo "
        "aggregato = somma su coorti vive; il costo NETTO sottrae il safety net assorbito (32.5 €mld x quota "
        "nuovi pensionati su 17 mln). P0 e' pagato dalla FISCALITA' (non dai contributi): per questo e' un costo "
        "del saldo, mentre P1 resta contributiva — nessuna doppia contabilizzazione.",
        "",
        "COME P2 FA RISPARMIARE (definizione operativa): il risparmio NON e' il fondo. Il canale e': (i) i "
        "contributi biforcati riducono la P1 NDC delle nuove coorti (DeltaP1 = P1_SQ - P1_rif, matura dal 2051); "
        "(ii) al netto dell'IRPEF persa su quelle P1; (iii) il costo simmetrico e' la Leva D (contributi persi "
        "subito). Le rendite P2 NON entrano nel bilancio pubblico: escono dal fondo (Dinamica) e sostituiscono "
        "reddito pensionistico che lo Stato non paga. Riducono pero' il P0 via taper e, in regime EET, generano "
        "gettito. Il vantaggio sistemico r>g si vede nella colonna 'Quota pagata dai FONDI' della Dinamica.",
        "",
        "RAMPA FASI (correzione): gli scatti di fase della biforcazione convergono al target in 5 anni "
        "(rampa_fasi=5) invece che in un anno secco: lo scalino 6%->10% costava 23 €mld di Leva D in un solo "
        "anno, un profilo di policy implausibile che creava 'crolli' artificiali nel saldo.",
        "",
        "TIMING DI P0 (correzione su input utente): P0 spetta SOLO alle coorti che hanno partecipato alla "
        "biforcazione (prime pensioni nel 2051). Chi si pensiona prima resta nel regime assistenziale attuale "
        "(dentro la quota assistenza della spesa): niente P0 e niente credito safety net. Cosi' P0 nasce "
        "insieme al sistema a capitalizzazione e sparisce la quasi-doppia pensione dei primi decenni.",
        "",
        "P1 E DEMOGRAFIA: la pensione del nuovo pensionato e' costante in termini reali per convenzione "
        "(come v14). In un NDC reale P1 si muove con la demografia attraverso tre canali: rivalutazione del "
        "montante al PIL (qui g fisso 0.8%, che gia' incorpora la demografia media), coefficienti di "
        "trasformazione aggiornati all'aspettativa di vita (qui fissi: prudente, li taglierebbe), e carriere/"
        "salari. L'effetto demografico AGGREGATO e' invece dentro il sentiero RGS della spesa H(t). "
        "Raffinamento possibile: coefficienti dinamici per coorte.",
        "",
        "CUNEO CONTRIBUTIVO E LEVA K (ridisegno Task 3, 2026-07): la biforcazione e' DENTRO il 33% (33-bif al "
        "PAYG, bif ai fondi): la pressione contributiva totale NON aumenta mai. La Leva K taglia il cuneo TOTALE "
        "sotto il 33% in base al CALO DEL MONTE P1 pubblico: sconto = min(levaK_max=6pp, levaK_coeff x "
        "dP1_lordo/H_contrib). Man mano che le coorti biforcate sostituiscono quelle full-PAYG, il pilastro 1 "
        "eroga meno pensioni contributive e il cuneo scende in proporzione (dividendo restituito ai lavoratori). "
        "Il taglio e' finanziato AL PIU' dalla quota levaK_quota_dividendo (50%) del dividendo netto di P1 "
        "dell'anno (dP1 - IRPEF persa): resta un costo di policy vero (puo' incidere sul saldo, non lo azzera per "
        "costruzione come nel vecchio aggancio al surplus) senza far esplodere il debito. Il driver e' monotono e "
        "liscio -> nessuna instabilita'. Prudenza: si contabilizza la perdita di gettito ma NON il minor accrual "
        "NDC futuro (risparmio non contato). Il cuneo effettivo e' nel foglio Modello e nel grafico dedicato.",
        "",
        "Canale 4 (ricalibrazione esplicita di P1 oltre l'effetto meccanico NDC): parametro clawback_p2, "
        "BASE = 0.7 (scelta 2026-07): il 70% della rendita P2 riduce la P1, cosi' la pensione totale a "
        "regime resta ~al livello dello status quo (~1.270 vs 1.179 €/mese, coorte 2080) e il dividendo "
        "rf>g va a minor spesa pubblica e, via Leva K, a taglio del cuneo. Il foglio Modello_C4_OFF mostra "
        "il counterfattuale senza clawback. Framing giuridico: ricalibrazione del coefficiente NDC per le "
        "coorti biforcate (la P1 e' prestazione pubblica; il montante P2 non si tocca). "
        "Va dichiarata nella riforma, non nascosta nei calcoli.",
        "",
        "Limiti noti (dichiarati): pensione controfattuale costante in termini reali per coorte (come v14); "
        "mortalita' lineare invece di tavole ISTAT; distribuzione redditi a 3 profili; perimetro L1 (17.5 mln "
        "dipendenti) mentre la spesa copre l'intero sistema; nessun effetto di equilibrio generale del fondo "
        "sul PIL (canale 'risollevamento' macro: da modellare separatamente, qui volutamente escluso per prudenza).",
    ]
    for i, txt in enumerate(note_txt, start=1):
        c = ws.cell(row=i, column=1, value=txt)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if i == 1:
            c.font = Font(bold=True, size=13)

    wb.save(out_path)
    return out_path


# ------------------------------------------------------------------------------
# 7bis. FIGURE PNG (matplotlib) — salvate in ./figure_v15/
# ------------------------------------------------------------------------------

def genera_figure(base, pack, sens, lungo, lungo_pack, p, cartella="figure_v15",
                  mc: dict | None = None):
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.ticker import PercentFormatter
    except ImportError:
        print("[figure] matplotlib non installato: pip install matplotlib")
        return []
    import os
    os.makedirs(cartella, exist_ok=True)
    out = []

    def salva(fig, nome):
        path = os.path.join(cartella, nome)
        fig.tight_layout()
        fig.savefig(path, dpi=150)
        plt.close(fig)
        out.append(path)

    # 1) Spesa pensionistica %PIL — VENTAGLIO DELLA SPESA PUBBLICA per rendimento.
    # ATTENZIONE METODOLOGICA: la variabile che misura la "copertura dei fondi" e'
    # la SPESA PUBBLICA (piu' i fondi rendono, piu' il clawback abbatte P1 e piu'
    # la spesa pubblica scende), NON il sistema totale. Il sistema totale e' la
    # SOMMA di due movimenti opposti quasi equivalenti (spesa pubblica -0.76pp,
    # rendite P2 +0.98pp fra rf=0.5% e rf=4% al 2080) e appare percio' piatto:
    # graficarlo da solo NASCONDE proprio l'effetto del rendimento. Qui si
    # ventaglia la spesa pubblica e si riporta il sistema totale solo per il caso
    # base, come riferimento del costo-risorse complessivo.
    fig, ax = plt.subplots(figsize=(11, 6))
    ax.plot(base["anno"], base["ratio"], lw=2.6, color="#C00000",
            label="Status quo (DFP 2026/RGS)")

    stili_rf = {0.005: ("#9E9E9E", ":"), 0.0125: ("#F9A825", "-."),
                0.025: ("#1F4E78", "-"), 0.04: ("#2E7D32", (0, (5, 1)))}
    rf_min, rf_max = min(sens), max(sens)
    for rf_val, d_ in sorted(sens.items()):
        col, dash = stili_rf.get(rf_val, ("#555555", "--"))
        evid = abs(rf_val - p.rf) < 1e-12
        ax.plot(d_["anno"], d_["ratio_rif"], lw=2.6 if evid else 1.7, ls=dash,
                color=col, label=(f"Riforma — spesa PUBBLICA, rf = {rf_val:.2%}"
                                  + (" (BASE)" if evid else "")))
    # fascia = incertezza sulla spesa pubblica dovuta al rendimento dei fondi
    ax.fill_between(base["anno"], sens[rf_max]["ratio_rif"], sens[rf_min]["ratio_rif"],
                    alpha=0.13, color="#1F4E78",
                    label=f"Ventaglio spesa pubblica (rf {rf_min:.2%}–{rf_max:.2%})")
    # sistema totale del solo caso base (costo-risorse: pubblico + rendite P2)
    sist_base = (base["spesa_rif"] + base["rendite_p2"]) / base["pil"]
    ax.plot(base["anno"], sist_base, lw=2.0, ls="--", color="#7B1FA2",
            label="Sistema totale (pubbl. + rendite P2), rf base")
    ax.fill_between(base["anno"], base["ratio_rif"], sist_base, alpha=0.12,
                    color="#7B1FA2", label="Quota pagata dai FONDI (base)")

    ax.set_title("Spesa pensionistica in % del PIL — quanto i fondi abbattono la spesa PUBBLICA")
    ax.set_ylabel("% del PIL")
    ax.set_xlabel(f"Anno   —   in tutte le curve r_BRDP = {p.r_brdp:.2%} (fisso): varia solo il "
                  f"rendimento dei fondi rf. Griglia completa r_BRDP × rf nel foglio Dinamica_scenari.")
    ax.yaxis.set_major_formatter(PercentFormatter(1.0))
    ax.legend(fontsize=8, loc="lower left"); ax.grid(alpha=0.3)
    salva(fig, "2_spesa_pensionistica_pil.png")

    # 3) Fondo per rendimento rf (€mld, con decumulo) + rendite
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5.5))
    colori = ["#9E9E9E", "#F9A825", "#1F4E78", "#2E7D32"]
    for (rf_val, d_), col in zip(sorted(sens.items()), colori):
        ax1.plot(d_["anno"], d_["fondo"], lw=2.2, color=col, label=f"rf = {rf_val:.2%}")
        ax2.plot(d_["anno"], d_["rendite_p2"], lw=2.2, color=col, label=f"rf = {rf_val:.2%}")
    ax1.set_title("Stock del fondo (€mld reali)\nestrazioni per le rendite INCLUSE")
    ax2.set_title("Rendite P2 pagate dai fondi (€mld reali/anno)")
    for ax in (ax1, ax2):
        ax.set_ylabel("€mld"); ax.legend(); ax.grid(alpha=0.3)
    salva(fig, "3_fondo_per_rendimento.png")

    # 4) Saldo totale e debito di transizione: base vs pacchetto
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5.5))
    ax1.axhline(0, color="k", lw=0.8)
    ax1.plot(base["anno"], base["saldo_tot"], lw=2.2, color="#1F4E78", label="Base")
    ax1.plot(pack["anno"], pack["saldo_tot"], lw=2.2, color="#2E7D32", label="Pacchetto")
    ax1.set_title("Saldo totale annuo vs status quo (€mld)"); ax1.set_ylabel("€mld")
    ax2.axhline(0, color="k", lw=0.8)
    ax2.plot(lungo["anno"], lungo["saldo_cum"] / lungo["pil"], lw=2.2, color="#1F4E78", label="Base")
    ax2.plot(lungo_pack["anno"], lungo_pack["saldo_cum"] / lungo_pack["pil"], lw=2.2,
             color="#2E7D32", label="Pacchetto")
    ax2.set_title("Posizione netta vs status quo, % PIL (positivo = risparmio cumulato)")
    ax2.yaxis.set_major_formatter(PercentFormatter(1.0))
    for ax in (ax1, ax2):
        ax.legend(); ax.grid(alpha=0.3)
    salva(fig, "4_saldo_e_posizione_netta.png")

    # 5) Bar chart P1/P2/P0 del nuovo pensionato: 2030, 2055, 2080
    fig, ax = plt.subplots(figsize=(10, 5.5))
    anni_b = (2030, 2055, 2080)
    x = np.arange(len(anni_b))
    larg = 0.35
    p1 = [base["p1_new"][a - p.anno0] for a in anni_b]
    p2 = [base["p2_new"][a - p.anno0] for a in anni_b]
    p0 = [base["p0_m1_new"][a - p.anno0] for a in anni_b]
    sq = [base["p1_sq_new"][a - p.anno0] for a in anni_b]
    ax.bar(x - larg/2, p1, larg, color="#1F4E78", label="P1 (NDC residua)")
    ax.bar(x - larg/2, p2, larg, bottom=p1, color="#2E7D32", label="P2 (rendita fondi)")
    ax.bar(x - larg/2, p0, larg, bottom=np.array(p1)+np.array(p2), color="#F9A825", label="P0 (taper)")
    ax.bar(x + larg/2, sq, larg, color="#C00000", alpha=0.75, label="Status quo (P1 piena)")
    for i in range(len(anni_b)):
        tot = p1[i] + p2[i] + p0[i]
        ax.text(x[i] - larg/2, tot + 15, f"{tot:,.0f}", ha="center", fontsize=9, weight="bold")
        ax.text(x[i] + larg/2, sq[i] + 15, f"{sq[i]:,.0f}", ha="center", fontsize=9)
    ax.set_xticks(x); ax.set_xticklabels([str(a) for a in anni_b])
    ax.set_ylabel("€/mese (reali 2027)")
    ax.set_title("Pensione del NUOVO pensionato: riforma (P1+P2+P0) vs status quo")
    ax.legend(); ax.grid(alpha=0.3, axis="y")
    salva(fig, "5_pensione_nuovo_pensionato.png")

    # 6) FAN CHART Monte Carlo (Task 9) — incertezza sul rendimento dei fondi
    if mc is not None:
        anni_mc = mc["anno"]
        chiavi = ["ratio_rif", "fondo", "saldo_cum", "pensione_tot"]
        fig, axes = plt.subplots(2, 2, figsize=(13, 9))
        for ax, k in zip(axes.ravel(), chiavi):
            lbl, su_pil = MC_SERIE[k]
            pc = mc["perc"][k]
            ax.fill_between(anni_mc, pc[5], pc[95], alpha=0.18, color="#1F4E78",
                            label="5°–95° percentile")
            ax.fill_between(anni_mc, pc[25], pc[75], alpha=0.32, color="#1F4E78",
                            label="25°–75° percentile")
            ax.plot(anni_mc, pc[50], lw=2.4, color="#1F4E78", label="mediana")
            ax.plot(anni_mc, mc["det_serie"][k], lw=1.8, ls="--", color="#C00000",
                    label=f"deterministico (rf={p.rf:.2%})")
            if k == "pensione_tot":      # riferimento: pensione dello status quo
                ax.plot(anni_mc, mc["det"]["p1_sq_new"], lw=1.6, ls=":", color="#2E7D32",
                        label="status quo")
            if k == "saldo_cum":
                ax.axhline(0, color="k", lw=0.8)
            ax.set_title(lbl, fontsize=11)
            ax.set_xlabel("Anno"); ax.grid(alpha=0.3)
            if su_pil or "%PIL" in lbl:
                ax.yaxis.set_major_formatter(PercentFormatter(1.0))
            ax.legend(fontsize=7)
        fig.suptitle(
            f"Fan chart Monte Carlo — {mc['n_sim']} simulazioni, rendimento reale dei fondi "
            f"i.i.d. ~ N({p.rf:.1%}, {p.rf_sigma:.0%}).\n"
            "La MEDIANA sta sotto il deterministico per volatility drag "
            "(crescita composta ≈ μ − σ²/2); il ventaglio misura il rischio di mercato "
            "trasferito ai pensionati.", fontsize=10)
        fig.tight_layout(rect=(0, 0, 1, 0.93))
        salva(fig, "6_fan_chart_montecarlo.png")

    return out


# ------------------------------------------------------------------------------
# 8. MAIN
# ------------------------------------------------------------------------------

_V14_NOME = "modello_dinamico_v14_2027_2080_corretto.xlsx"


def _trova_file_v14() -> str:
    """Percorso del file v14 per la validazione: sta nella cartella corrente
    oppure nella sottocartella storica del progetto."""
    import os
    for cand in (_V14_NOME, os.path.join("Progetto sistema Pensionistico", _V14_NOME)):
        if os.path.exists(cand):
            return cand
    return _V14_NOME


def main(perimetro: str | None = None, out_default: str | None = None,
         figure_dir: str = "figure_v15", etichetta: str = ""):
    """Esecuzione completa. perimetro/out_default/figure_dir sono passati dai due
    runner (modello_v15_totale.py e modello_v15_dip.py); da riga di comando si
    possono comunque forzare con --perimetro/--out."""
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=out_default or "modello_dinamico_v15_2027_2080.xlsx")
    ap.add_argument("--v14", default=_trova_file_v14(),
                    help="file v14 per la validazione della replica")
    ap.add_argument("--perimetro", default=perimetro,
                    choices=["dipendenti", "universale", "L1"],
                    help="dipendenti = sotto-sistema FPLD+CTPS coerente (default); "
                         "universale = anche autonomi e casse; L1 = alias di dipendenti")
    ap.add_argument("--figure", default=figure_dir, help="cartella delle figure PNG")
    ap.add_argument("--mc-sim", type=int, default=None,
                    help="numero di simulazioni Monte Carlo (0 = disattiva)")
    args = ap.parse_args()

    p = Params()
    if args.perimetro:
        p = dataclasses.replace(p, perimetro=args.perimetro)
    if etichetta:
        print("=" * 78)
        print(f"  {etichetta}")
        print(f"  Perimetro: {p.perimetro} | output: {args.out} | figure: {args.figure}")
        print("=" * 78)

    def P(**kw) -> Params:
        """Params di uno scenario: eredita SEMPRE il perimetro del run corrente
        (senza questo, gli scenari secondari tornerebbero al default)."""
        return dataclasses.replace(p, **kw)

    def P_pess(**kw) -> Params:
        """Come P(), ma con la calibrazione pessimistica (g, g_w, rf ridotti)."""
        pp = params_pessimistico()
        return dataclasses.replace(pp, perimetro=p.perimetro, **kw)

    def tassi(pp: Params) -> str:
        """Dichiarazione esplicita dei tassi reali usati nel run."""
        return (f"[tassi reali: rf={pp.rf:.2%}, r_BRDP={pp.r_brdp:.2%}, "
                f"r_debito={pp.r_debito:.2%}]")

    print("== Replica v14 (validazione, parametri v14 CONGELATI) ==")
    rep = run_replica_v14(params_v14())
    for anno, lbl in [(2027, "t=0"), (2040, "picco"), (2060, ""), (2080, "fine")]:
        i = anno - p.anno0
        print(f"  {anno}: LevaF={rep['U'][i]:8.2f}  Saldo={rep['AE'][i]:8.2f}  "
              f"SaldoCum={rep['AF'][i]:8.2f}")

    valid, serie_v14 = valida_replica(rep, args.v14)
    if valid:
        print("== Confronto con file v14 ==")
        for col, mx, mean in valid:
            esito = "OK" if mx < 0.05 else "!! VERIFICARE"
            print(f"  col {col:3s}: max|diff|={mx:10.4f}  media={mean:10.4f}  {esito}")

    print(f"== Modello v15 (base) {tassi(p)} ==")
    base = run_v15(p)
    for anno in (2030, 2040, 2050, 2060, 2070, 2080):
        i = anno - p.anno0
        print(f"  {anno}: LevaF={base['levaF'][i]:7.2f}  SaldoPrim={base['saldo_prim'][i]:7.2f}  "
              f"SaldoTot={base['saldo_tot'][i]:7.2f}  Debito%PIL={base['debito'][i]/base['pil'][i]:6.1%}  "
              f"Fondo%PIL={base['fondo'][i]/base['pil'][i]:6.1%}  "
              f"RenditeP2={base['rendite_p2'][i]:6.1f}")

    p_pess = P_pess()
    print(f"== Modello v15 (pessimistico) {tassi(p_pess)} ==")
    pess = run_v15(p_pess)
    for anno in (2040, 2060, 2080):
        i = anno - p.anno0
        print(f"  {anno}: SaldoTot={pess['saldo_tot'][i]:7.2f}  "
              f"Fondo%PIL={pess['fondo'][i]/pess['pil'][i]:6.1%}")

    # controfattuale SENZA ricalibrazione: il pensionato tiene tutto il dividendo
    # della capitalizzazione (vale per entrambi i modi di clawback)
    p_c4 = P(clawback_p2=0.0, clawback_modo="costante")
    print(f"== Modello v15 (canale 4 OFF: counterfattuale senza clawback) {tassi(p_c4)} ==")
    claw = run_v15(p_c4)
    for anno in (2040, 2060, 2080):
        i = anno - p.anno0
        print(f"  {anno}: LevaF={claw['levaF'][i]:7.2f}  SaldoTot={claw['saldo_tot'][i]:7.2f}  "
              f"Debito%PIL={claw['debito'][i]/claw['pil'][i]:6.1%}")

    # sensibilita' al rendimento fondi
    print(f"== Sensibilita' al rendimento fondi (r_BRDP={p.r_brdp:.2%} fisso) ==")
    sens = {}
    for rf_val in (0.005, 0.0125, 0.025, 0.04):
        p_s = P(rf=rf_val)
        sens[rf_val] = run_v15(p_s)
        i80 = 2080 - p.anno0
        d_ = sens[rf_val]
        print(f"  rf={rf_val:6.2%}: Fondo2080={d_['fondo'][i80]:7.0f} mld "
              f"({d_['fondo'][i80]/d_['pil'][i80]:6.1%} PIL)  "
              f"RenditeP2 2080={d_['rendite_p2'][i80]:6.1f} mld  "
              f"SpesaPubbl2080={d_['ratio_rif'][i80]:6.1%}")

    # griglia 4x4 r_BRDP x rf (struttura Dinamica di v14, sezioni I-II):
    # ogni combinazione e' un run COMPLETO del motore v15 (coorti, decumulo,
    # interessi sul debito coerenti con r_BRDP) — base e pessimistico.
    R_BRDP_GRID = (0.0075, 0.015, 0.025, 0.035)
    RF_GRID = (0.005, 0.0125, 0.025, 0.04)
    griglia, griglia_pess = {}, {}
    for rb in R_BRDP_GRID:
        for rf_val in RF_GRID:
            p_g = P(r_brdp=rb, rf=rf_val)
            griglia[(rb, rf_val)] = run_v15(p_g)
            p_gp = P_pess(r_brdp=rb, rf=rf_val)
            griglia_pess[(rb, rf_val)] = run_v15(p_gp)

    # stampa della griglia (come il foglio Output/Dinamica di v10): ogni riga
    # dichiara ESPLICITAMENTE la coppia di tassi reali a cui e' calcolata.
    i60g, i80g = 2060 - p.anno0, 2080 - p.anno0
    for gr_, nome_gr in ((griglia, "BASE"), (griglia_pess, "PESSIMISTICO")):
        print(f"\n== GRIGLIA r_BRDP x rf — scenario {nome_gr} "
              f"(16 run completi v15; €mld reali 2027) ==")
        print(f"{'r_BRDP':>7} {'rf':>7} {'Fondo60':>8} {'Fondo80':>8} {'F80%PIL':>8} "
              f"{'RendP2_80':>9} {'SpPub80':>8} {'SisTot80':>8} {'Saldo80':>8} {'Deb80%PIL':>9}")
        for rb in R_BRDP_GRID:
            for rf_val in RF_GRID:
                d_ = gr_[(rb, rf_val)]
                sistot = (d_['spesa_rif'][i80g] + d_['rendite_p2'][i80g]) / d_['pil'][i80g]
                print(f"{rb:>7.2%} {rf_val:>7.2%} {d_['fondo'][i60g]:>8.0f} {d_['fondo'][i80g]:>8.0f} "
                      f"{d_['fondo'][i80g]/d_['pil'][i80g]:>8.1%} {d_['rendite_p2'][i80g]:>9.1f} "
                      f"{d_['ratio_rif'][i80g]:>8.1%} {sistot:>8.1%} "
                      f"{d_['saldo_tot'][i80g]:>8.1f} {d_['debito'][i80g]/d_['pil'][i80g]:>9.1%}")

    # PACCHETTO ANTICIPO FRUTTI: clawback base + recupero tax gap (Leva L)
    p_pack = P(levaL_attiva=1)
    print(f"\n== Pacchetto anticipo frutti {tassi(p_pack)} ==")
    pack = run_v15(p_pack)

    # proiezioni lunghe (2100) per picco/pareggio del debito
    lungo = run_v15(P(anni=74))
    p_pl = P(anni=74, levaL_attiva=1)
    lungo_pack = run_v15(p_pl)
    for lg, nome in ((lungo, "BASE"), (lungo_pack, "PACCHETTO")):
        i_picco = int(np.argmax(lg["debito"]))
        anni_zero = [int(lg["anno"][i]) for i in range(i_picco, 74)
                     if lg["debito"][i] <= 0]
        print(f"\n== Proiezione lunga {nome} (2100) ==\n  Picco debito: {int(lg['anno'][i_picco])} "
              f"({lg['debito'][i_picco]:.0f} mld, {lg['debito'][i_picco]/lg['pil'][i_picco]:.1%} PIL)  "
              f"Pareggio: {anni_zero[0] if anni_zero else '>2100'}  "
              f"2100: saldo {lg['saldo_tot'][-1]:+.0f} mld, debito {lg['debito'][-1]/lg['pil'][-1]:.1%} PIL")

    print("\n== DEBITO PENSIONISTICO cumulato dal 2027 (%PIL): senza vs con riforma ==")
    print(f"{'Anno':>6} {'SENZA riforma':>14} {'CON riforma (base)':>19} {'CON rif. (pacchetto)':>21}")
    for anno in (2030, 2040, 2050, 2060, 2070, 2080, 2090, 2100):
        i = anno - p.anno0
        print(f"{anno:>6} {lungo['debito_pens_sq'][i]/lungo['pil'][i]:>14.1%} "
              f"{lungo['debito_pens_rif'][i]/lungo['pil'][i]:>19.1%} "
              f"{lungo_pack['debito_pens_rif'][i]/lungo_pack['pil'][i]:>21.1%}")

    print("\n== Pacchetto anticipo frutti (clawback base + tax gap 10 mld) ==")
    for anno in (2035, 2045, 2055, 2065, 2075, 2080):
        i = anno - p.anno0
        print(f"  {anno}: SaldoTot base={base['saldo_tot'][i]:7.2f}  pacchetto={pack['saldo_tot'][i]:7.2f}  "
              f"DebitoTr base={base['debito'][i]/base['pil'][i]:6.1%}  pacchetto={pack['debito'][i]/pack['pil'][i]:6.1%}")

    print("\n== Spesa pensionistica %PIL (base) ==")
    print(f"{'Anno':>6} {'SQ':>7} {'Pubblica rif.':>14} {'Rendite P2':>11} {'Sistema tot':>12}")
    for anno in (2030, 2040, 2050, 2060, 2070, 2080):
        i = anno - p.anno0
        print(f"{anno:>6} {base['ratio'][i]:>7.1%} {base['ratio_rif'][i]:>14.1%} "
              f"{base['rendite_p2'][i]/base['pil'][i]:>11.2%} "
              f"{(base['spesa_rif'][i]+base['rendite_p2'][i])/base['pil'][i]:>12.1%}")

    # confronto sintetico v14 vs v15 (console)
    print("\n== v14 (file) vs v15 — anni chiave ==")
    print(f"{'Anno':>6} {'LevaF v14':>10} {'LevaF v15':>10} {'Saldo v14':>10} "
          f"{'SaldoTot v15':>13} {'RispCum v15':>12} {'Debito v15':>11}")
    for anno in (2030, 2040, 2050, 2060, 2070, 2080):
        i = anno - p.anno0
        u14 = serie_v14["U"][i] if serie_v14 is not None else float("nan")
        ae14 = serie_v14["AE"][i] if serie_v14 is not None else float("nan")
        print(f"{anno:>6} {u14:>10.2f} {base['levaF'][i]:>10.2f} {ae14:>10.2f} "
              f"{base['saldo_tot'][i]:>13.2f} {base['saldo_cum'][i]:>12.2f} "
              f"{base['debito'][i]:>11.2f}")

    # controlli di coerenza interna
    assert (base["fondo"] >= -1e-6).all(), "fondo negativo: incoerenza decumulo"
    err_id = np.max(np.abs(base["debito"] + base["saldo_cum"]))
    assert err_id < 1e-6, f"identita' debito = -risparmio cumulato violata: {err_id}"
    chk = base["fondo"][-1]
    print(f"\nCheck: fondo 2080 con decumulo = {chk:.0f} €mld "
          f"(v14 senza decumulo: {rep['fondo_v14'][-1]:.0f} €mld)")
    print(f"Check identita': debito transizione = -risparmio cumulato (err={err_id:.2e})")

    # --- AUDIT ANNO BASE: il modello riproduce il 2027 osservato? -----------------
    print("\n== Audit anno base (modello vs osservato) ==")
    print(f"  {'voce':>38} {'modello':>10} {'osservato':>10} {'scarto':>8}")
    for voce, mod, oss, _fonte, giud in audit_anno_base(p, base):
        pct = "Copertura" in voce or "PIL" in voce
        f = (lambda x: f"{x:>9.1%}") if pct else (lambda x: f"{x:>9,.1f}")
        flag = "" if giud == "ok" else "  <-- SCOSTAMENTO DICHIARATO"
        print(f"  {voce:>38} {f(mod)} {f(oss)} {mod/oss-1:>+8.0%}{flag}")

    # --- MONTE CARLO sul rendimento dei fondi (fan chart) -------------------------
    n_sim = p.mc_n_sim if args.mc_sim is None else args.mc_sim
    mc = None
    if n_sim > 0:
        print(f"\n== Monte Carlo su rf: {n_sim} simulazioni, "
              f"rf ~ N({p.rf:.2%}, {p.rf_sigma:.1%}) ==")
        mc = simula_montecarlo(p, n_sim=n_sim)
        sint = sintesi_montecarlo(mc, 2080)
        print(f"  {'serie':>36} {'5° perc.':>10} {'mediana':>10} {'95° perc.':>10} {'determ.':>10}")
        for k, (lbl, su_pil) in MC_SERIE.items():
            d_ = sint[k]
            fm = (lambda x: f"{x:>9.2%}") if (su_pil or "%PIL" in lbl) else (
                 (lambda x: f"{x:>9.0f}") if "€/mese" in lbl else (lambda x: f"{x:>9.1f}"))
            print(f"  {lbl:>36} {fm(d_['p5'])} {fm(d_['mediana'])} {fm(d_['p95'])} {fm(d_['det'])}")
        print(f"  Volatility drag sul fondo 2080 (mediana - deterministico): "
              f"{sint['fondo']['drag']:+.2%} del PIL")
        print(f"  P(posizione netta < 0 al 2080)          = {sint['prob_pos_netta_negativa']:.1%}")
        print(f"  P(pensione < status quo al 2080)        = {sint['prob_pensione_sotto_sq']:.1%}")

    out = scrivi_excel(args.out, rep, base, pess, claw, pack, valid, serie_v14,
                       sens, lungo, lungo_pack, p,
                       griglia=griglia, griglia_pess=griglia_pess, mc=mc)
    print(f"\nScritto: {out}")

    figure = genera_figure(base, pack, sens, lungo, lungo_pack, p,
                           cartella=args.figure, mc=mc)
    for f_ in figure:
        print(f"Figura: {f_}")
    if figure:
        print("(altre 9 figure native sono nel foglio 'Grafici' dell'Excel)")


if __name__ == "__main__":
    # Questo file e' il MOTORE CONDIVISO, non un eseguibile: lanciarlo
    # direttamente produrrebbe un terzo output quasi-duplicato (con il perimetro
    # di default) accanto a quelli dei due runner, ed e' esattamente il tipo di
    # ambiguita' che fa divergere i risultati senza che nessuno se ne accorga.
    raise SystemExit(
        "modello_v15_core.py e' il motore condiviso e non va lanciato direttamente.\n"
        "  Perimetro totale (dipendenti + autonomi):  python3 modello_v15_totale.py\n"
        "  Perimetro dipendenti (17,5 mln FPLD+CTPS): python3 modello_v15_dip.py")
